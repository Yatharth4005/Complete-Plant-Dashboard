import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import calendar
from datetime import date

def generate_performance_excel(department, year, month, records):
    """
    Generates a beautifully styled Excel workbook for the department's monthly performance data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Performance - {department.code}"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Styling helpers
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header_parent = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_header_child = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_total = Font(name="Segoe UI", size=10, bold=True)
    
    fill_title = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid") # Indigo
    fill_date = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate
    fill_plan = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Blue
    fill_plan_light = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Lighter Blue
    fill_actual = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid") # Green
    fill_actual_light = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid") # Lighter Green
    fill_loss_nof = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid") # Orange
    fill_loss_eaf = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid") # Red
    
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    thick_border_bottom = Side(border_style="medium", color="1E293B")
    double_border_bottom = Side(border_style="double", color="1E293B")
    
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=double_border_bottom)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # 1. Title Block
    month_name = calendar.month_name[month]
    ws.merge_cells("A1:O1" if department.code == 'SMS2' else "A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"📈 Performance Metrics Dashboard — {department.name} ({month_name}, {year})"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 40
    
    # Empty space
    ws.row_dimensions[2].height = 10
    
    # Build maps for existing data
    records_by_day = {r.date.day: r for r in records}
    num_days = calendar.monthrange(year, month)[1]
    
    if department.code == 'SMS2':
        # SMS2 Columns config
        # Headers structure:
        # Row 3 (Parent Headers)
        # Row 4 (Child Headers)
        ws.row_dimensions[3].height = 28
        ws.row_dimensions[4].height = 24
        
        # DATE (rowspan=2)
        ws.merge_cells("A3:A4")
        cell_date = ws["A3"]
        cell_date.value = "DATE"
        cell_date.font = font_header_parent
        cell_date.fill = fill_date
        cell_date.alignment = align_center
        
        # PLAN (colspan=6)
        ws.merge_cells("B3:G3")
        cell_plan = ws["B3"]
        cell_plan.value = "PLAN"
        cell_plan.font = font_header_parent
        cell_plan.fill = fill_plan
        cell_plan.alignment = align_center
        
        # PLAN Sub headers
        plan_subs = [
            ("Total Tap", "B4"),
            ("Total Production (MT)", "C4"),
            ("EAF-II (Tap)", "D4"),
            ("EAF-II (Production) (MT)", "E4"),
            ("NOF (Tap)", "F4"),
            ("NOF (Production) (MT)", "G4")
        ]
        for name, cell_ref in plan_subs:
            c = ws[cell_ref]
            c.value = name
            c.font = font_header_child
            c.fill = fill_plan_light
            c.alignment = align_center
            
        # ACTUAL (colspan=6)
        ws.merge_cells("H3:M3")
        cell_actual = ws["H3"]
        cell_actual.value = "ACTUAL"
        cell_actual.font = font_header_parent
        cell_actual.fill = fill_actual
        cell_actual.alignment = align_center
        
        # ACTUAL Sub headers
        actual_subs = [
            ("Total Tap", "H4"),
            ("Total Production (MT)", "I4"),
            ("EAF-II (Tap)", "J4"),
            ("EAF-II (Production) (MT)", "K4"),
            ("NOF (Tap)", "L4"),
            ("NOF (Production) (MT)", "M4")
        ]
        for name, cell_ref in actual_subs:
            c = ws[cell_ref]
            c.value = name
            c.font = font_header_child
            c.fill = fill_actual_light
            c.alignment = align_center
            
        # PRODUCTION LOSS (NOF) (colspan=1, rowspan=2 split)
        ws.merge_cells("N3:N3")
        cell_loss_nof = ws["N3"]
        cell_loss_nof.value = "PRODUCTION LOSS (NOF)"
        cell_loss_nof.font = font_header_child
        cell_loss_nof.fill = fill_loss_nof
        cell_loss_nof.alignment = align_center
        
        cell_loss_nof_sub = ws["N4"]
        cell_loss_nof_sub.value = "Loss (MT)"
        cell_loss_nof_sub.font = font_header_child
        cell_loss_nof_sub.fill = fill_loss_nof
        cell_loss_nof_sub.alignment = align_center
        
        # PRODUCTION LOSS (EAF-II)
        ws.merge_cells("O3:O3")
        cell_loss_eaf = ws["O3"]
        cell_loss_eaf.value = "PRODUCTION LOSS (EAF-II)"
        cell_loss_eaf.font = font_header_child
        cell_loss_eaf.fill = fill_loss_eaf
        cell_loss_eaf.alignment = align_center
        
        cell_loss_eaf_sub = ws["O4"]
        cell_loss_eaf_sub.value = "Loss (MT)"
        cell_loss_eaf_sub.font = font_header_child
        cell_loss_eaf_sub.fill = fill_loss_eaf
        cell_loss_eaf_sub.alignment = align_center
        
        # Style all cells in Header area (even merged ones need borders/fills to prevent rendering gaps)
        for r in range(3, 5):
            for col in range(1, 16):
                cell = ws.cell(row=r, column=col)
                if not cell.fill.fill_type:
                    # Apply corresponding fill
                    if col == 1: cell.fill = fill_date
                    elif col <= 7: cell.fill = fill_plan
                    elif col <= 13: cell.fill = fill_actual
                    elif col == 14: cell.fill = fill_loss_nof
                    elif col == 15: cell.fill = fill_loss_eaf
                cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_border_bottom)

        # Write Data
        row_idx = 5
        for day in range(1, num_days + 1):
            r = records_by_day.get(day)
            ws.row_dimensions[row_idx].height = 20
            
            # Format row elements
            row_data = [
                day,
                r.plan_tap_sms2 if r else 0.0,
                r.plan_prod_sms if r else 0.0,
                r.plan_eaf2 if r else 0.0,
                r.plan_prod_eaf2 if r else 0.0,
                r.plan_neof if r else 0.0,
                r.plan_prod_neof if r else 0.0,
                r.actual_tap_sms2 if r else 0.0,
                r.actual_prod_sms if r else 0.0,
                r.actual_eaf2 if r else 0.0,
                r.actual_prod_eaf2 if r else 0.0,
                r.actual_neof if r else 0.0,
                r.actual_prod_neof if r else 0.0,
                r.prod_loss_nof if r else 0.0,
                r.prod_loss_eaf2 if r else 0.0,
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = font_data
                cell.border = border_data
                
                # Alignments and number formatting
                if col_idx == 1:
                    cell.alignment = align_center
                    cell.font = Font(name="Segoe UI", size=10, bold=True)
                else:
                    cell.alignment = align_center
                    # If columns are production (columns 3, 5, 7, 9, 11, 13, 14, 15), format with commas
                    if col_idx in [3, 5, 7, 9, 11, 13, 14, 15]:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.##"
                
                # Zebra striping
                if day % 2 == 0:
                    cell.fill = fill_zebra
                    
            row_idx += 1
            
        # Totals Row
        ws.row_dimensions[row_idx].height = 24
        total_cell = ws.cell(row=row_idx, column=1)
        total_cell.value = "Total"
        total_cell.font = font_total
        total_cell.alignment = align_center
        total_cell.fill = fill_total
        total_cell.border = border_total
        
        for col_idx in range(2, 16):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = f"=SUM({col_letter}5:{col_letter}{row_idx-1})"
            cell.font = font_total
            cell.alignment = align_center
            cell.fill = fill_total
            cell.border = border_total
            
            if col_idx in [3, 5, 7, 9, 11, 13, 14, 15]:
                cell.number_format = "#,##0"
            else:
                cell.number_format = "#,##0.##"
                
    else:
        # SMS3 Columns Config
        ws.row_dimensions[3].height = 28
        ws.row_dimensions[4].height = 24
        
        # DATE
        ws.merge_cells("A3:A4")
        cell_date = ws["A3"]
        cell_date.value = "DATE"
        cell_date.font = font_header_parent
        cell_date.fill = fill_date
        cell_date.alignment = align_center
        
        # PLAN
        ws.merge_cells("B3:C3")
        cell_plan = ws["B3"]
        cell_plan.value = "PLAN"
        cell_plan.font = font_header_parent
        cell_plan.fill = fill_plan
        cell_plan.alignment = align_center
        
        plan_subs = [
            ("EAF III (Heat Nos)", "B4"),
            ("PROD. EAF III (MT)", "C4")
        ]
        for name, cell_ref in plan_subs:
            c = ws[cell_ref]
            c.value = name
            c.font = font_header_child
            c.fill = fill_plan_light
            c.alignment = align_center
            
        # ACTUAL
        ws.merge_cells("D3:E3")
        cell_actual = ws["D3"]
        cell_actual.value = "ACTUAL"
        cell_actual.font = font_header_parent
        cell_actual.fill = fill_actual
        cell_actual.alignment = align_center
        
        actual_subs = [
            ("EAF III (Heat Nos)", "D4"),
            ("PROD. EAF III (MT)", "E4")
        ]
        for name, cell_ref in actual_subs:
            c = ws[cell_ref]
            c.value = name
            c.font = font_header_child
            c.fill = fill_actual_light
            c.alignment = align_center
            
        # PRODUCTION LOSS
        ws.merge_cells("F3:F3")
        cell_loss = ws["F3"]
        cell_loss.value = "PRODUCTION LOSS"
        cell_loss.font = font_header_child
        cell_loss.fill = fill_loss_eaf
        cell_loss.alignment = align_center
        
        cell_loss_sub = ws["F4"]
        cell_loss_sub.value = "Loss (MT)"
        cell_loss_sub.font = font_header_child
        cell_loss_sub.fill = fill_loss_eaf
        cell_loss_sub.alignment = align_center
        
        for r in range(3, 5):
            for col in range(1, 7):
                cell = ws.cell(row=r, column=col)
                if not cell.fill.fill_type:
                    if col == 1: cell.fill = fill_date
                    elif col <= 3: cell.fill = fill_plan
                    elif col <= 5: cell.fill = fill_actual
                    elif col == 6: cell.fill = fill_loss_eaf
                cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_border_bottom)

        # Write Data
        row_idx = 5
        for day in range(1, num_days + 1):
            r = records_by_day.get(day)
            ws.row_dimensions[row_idx].height = 20
            
            row_data = [
                day,
                r.plan_eaf3_heats if r else 0.0,
                r.plan_prod_eaf3 if r else 0.0,
                r.actual_eaf3_heats if r else 0.0,
                r.actual_prod_eaf3 if r else 0.0,
                r.prod_loss_eaf3 if r else 0.0,
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = font_data
                cell.border = border_data
                
                if col_idx == 1:
                    cell.alignment = align_center
                    cell.font = Font(name="Segoe UI", size=10, bold=True)
                else:
                    cell.alignment = align_center
                    # If columns are production (columns 3, 5, 6), format with commas
                    if col_idx in [3, 5, 6]:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.##"
                        
                if day % 2 == 0:
                    cell.fill = fill_zebra
                    
            row_idx += 1
            
        # Totals Row
        ws.row_dimensions[row_idx].height = 24
        total_cell = ws.cell(row=row_idx, column=1)
        total_cell.value = "Total"
        total_cell.font = font_total
        total_cell.alignment = align_center
        total_cell.fill = fill_total
        total_cell.border = border_total
        
        for col_idx in range(2, 7):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = f"=SUM({col_letter}5:{col_letter}{row_idx-1})"
            cell.font = font_total
            cell.alignment = align_center
            cell.fill = fill_total
            cell.border = border_total
            
            if col_idx in [3, 5, 6]:
                cell.number_format = "#,##0"
            else:
                cell.number_format = "#,##0.##"

    # Auto-adjust column widths with minimum padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            # Skip title row merged cell length to avoid making column A extremely wide
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            # Handle linebreaks in headers by finding max length of splits
            if '\n' in val_str:
                lines = val_str.split('\n')
                val_len = max(len(l) for l in lines)
            else:
                val_len = len(val_str)
                
            if val_len > max_len:
                max_len = val_len
                
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

    return wb
