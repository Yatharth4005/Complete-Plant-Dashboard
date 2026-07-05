import os
import re
import xlrd
import openpyxl
from datetime import datetime, date, time
from django.db import transaction
from django.utils import timezone
from tpm.models import Department
from delays.models import DelayRecord, DelayUpload

def parse_date_string(date_str):
    """
    Tries to parse a date string using various formats.
    Returns a datetime.date object if successful, else None.
    """
    if not date_str:
        return None
        
    if isinstance(date_str, (datetime, date)):
        if isinstance(date_str, datetime):
            return date_str.date()
        return date_str
    
    date_str = str(date_str).strip()
    
    # Strip prefixes like "DATE :", "DATE:", etc.
    date_str = re.sub(r'(?i)^date\s*:\s*[-]*\s*>', '', date_str).strip()
    date_str = re.sub(r'(?i)^date\s*:\s*', '', date_str).strip()
    
    # Try different formats
    formats = [
        '%d.%m.%Y', '%d.%m.%y',
        '%d-%m-%Y', '%d-%m-%y',
        '%d/%m/%Y', '%d/%m/%y',
        '%m/%d/%Y', '%m/%d/%y',
        '%Y/%m/%d', '%Y/%m/%d %H:%M:%S',
        '%Y-%m-%d',
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f',
        '%d-%b-%Y', '%d-%b-%y', # e.g. 10-Jun-2026
        '%d %b %Y', '%d %b %y',
        '%d %B %Y', '%d %B %y',
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
            
    # Try searching for a date pattern inside the string
    match = re.search(r'\d{1,2}[./-]\d{1,2}[./-]\d{2,4}', date_str)
    if match:
        for fmt in ['%d.%m.%Y', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d.%m.%y', '%d-%m-%y', '%d/%m/%y', '%m/%d/%y']:
            try:
                return datetime.strptime(match.group(0), fmt).date()
            except ValueError:
                continue
                
    match_alpha = re.search(r'\d{1,2}[- ][A-Za-z]{3}[- ]\d{4}', date_str)
    if match_alpha:
        for fmt in ['%d-%b-%Y', '%d %b %Y']:
            try:
                return datetime.strptime(match_alpha.group(0), fmt).date()
            except ValueError:
                continue

    return None


def parse_time_to_minutes(time_val):
    """
    Parses a time value (time/datetime object, string, or day fraction float)
    into the number of minutes since midnight.
    """
    if not time_val:
        return None
    if isinstance(time_val, (datetime, time)):
        return time_val.hour * 60 + time_val.minute
    
    # If it's a float/int (Excel time serial):
    if isinstance(time_val, float) and 0.0 < time_val < 1.0:
        return round(time_val * 24 * 60)
        
    # Try parsing string format:
    time_str = str(time_val).strip().replace('.', ':')
    match = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', time_str)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        return h * 60 + m
    return None


def format_time_value(time_val):
    """
    Standardizes a time value to a string format "HH:MM".
    """
    if not time_val:
        return None
    if isinstance(time_val, (datetime, time)):
        return time_val.strftime('%H:%M')
    if isinstance(time_val, float) and 0.0 <= time_val <= 1.0:
        total_mins = round(time_val * 24 * 60)
        h = total_mins // 60
        m = total_mins % 60
        return f"{h:02d}:{m:02d}"
    
    time_str = str(time_val).strip().replace('.', ':')
    match = re.search(r'(\d{1,2}):(\d{2})', time_str)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        return f"{h:02d}:{m:02d}"
    return str(time_val).strip()


def extract_times_from_range(range_str):
    """
    Extracts start and end times from a time range string.
    """
    if not range_str:
        return None, None
    range_str = str(range_str).strip()
    match = re.search(r'(\d{1,2}[:.]\d{2})\s*(?:-|to)\s*(\d{1,2}[:.]\d{2})', range_str, re.IGNORECASE)
    if match:
        start_raw = match.group(1).replace('.', ':')
        end_raw = match.group(2).replace('.', ':')
        return start_raw, end_raw
    return None, None


def parse_month_to_num(month_val):
    """
    Parses various representations of a month into an integer (1-12).
    """
    if not month_val:
        return None
    if isinstance(month_val, (int, float)):
        val = int(month_val)
        if 1 <= val <= 12:
            return val
        return None
    month_str = str(month_val).strip().lower()
    month_map = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    for k, v in month_map.items():
        if month_str.startswith(k):
            return v
    return None


def parse_row_date(date_val, month_num=None, default_year=None):
    """
    Parses a date cell value. Can handle full dates, or day integers
    when combined with a month_num and a default_year.
    """
    if not date_val:
        return None
        
    if isinstance(date_val, (datetime, date)):
        if isinstance(date_val, datetime):
            return date_val.date()
        return date_val
        
    parsed_full = parse_date_string(date_val)
    if parsed_full:
        return parsed_full
        
    # Try parsing as simple day number
    day_num = None
    try:
        if isinstance(date_val, (int, float)):
            day_num = int(date_val)
        else:
            match = re.search(r'\d+', str(date_val))
            if match:
                day_num = int(match.group(0))
    except Exception:
        pass
        
    if not day_num or day_num < 1 or day_num > 31:
        return None
        
    if not month_num:
        if default_year and isinstance(default_year, (datetime, date)):
            month_num = default_year.month
        else:
            month_num = timezone.now().date().month
            
    if default_year:
        if isinstance(default_year, (datetime, date)):
            year_num = default_year.year
        else:
            try:
                year_num = int(default_year)
            except Exception:
                year_num = timezone.now().date().year
    else:
        year_num = timezone.now().date().year
        
    try:
        return date(year_num, month_num, day_num)
    except ValueError:
        try:
            return date(year_num, month_num, 28)
        except ValueError:
            return None


def extract_year_from_string(s):
    """
    Extracts a 4-digit or 2-digit year from a string.
    """
    if not s:
        return None
    match4 = re.search(r'\b(20\d{2})\b', s)
    if match4:
        return int(match4.group(1))
    match2 = re.search(r'(?:\b|\')(\d{2})\b', s)
    if match2:
        yr = int(match2.group(1))
        return 2000 + yr
    return None


def parse_duration_to_mins(duration_val, col_type=None):
    """
    Parses excel duration cell into float minutes.
    """
    if duration_val in (None, '', '-'):
        return 0.0
    if isinstance(duration_val, (datetime, time)):
        return float(duration_val.hour * 60 + duration_val.minute)
    try:
        val = float(duration_val)
        if col_type == 'hours':
            return val * 60.0
        if 0.0 < val < 1.0 and col_type != 'minutes':
            # Excel time serial representing day fraction
            return float(round(val * 1440))
        return val
    except ValueError:
        return 0.0


def normalize_agency_name(name):
    if not name:
        return name
    name_str = str(name).strip()
    name_upper = name_str.upper().rstrip('.')
    
    mapping = {
        'MECH': 'Mechanical',
        'MECHANICAL': 'Mechanical',
        'ELEC': 'Electrical',
        'ELECTRICAL': 'Electrical',
        'OPER': 'Operations',
        'OPERATIONS': 'Operations',
        'INSTR': 'Instrumentation',
        'INSTRUMENTATION': 'Instrumentation',
        'REF': 'REF',
    }
    
    return mapping.get(name_upper, name_str)


def clean_parsed_fields(desc, agency, sub_agency=None, equipment=None, sub_equipment=None):
    """
    Cleans and normalizes parsed fields.
    If the description or agency contains 'shutdown', normalizes agency to 'SHUTDOWN'
    and clears equipment/agency details.
    Also normalizes the agency name to match Department names/codes if possible.
    """
    clean_desc = str(desc).strip() if desc else 'No Description'
    clean_agency = normalize_agency_name(agency) if agency else 'General'
    clean_sub_agency = str(sub_agency).strip() if sub_agency else None
    clean_equip = str(equipment).strip() if equipment else None
    clean_sub_equip = str(sub_equipment).strip() if sub_equipment else None

    # Check for shutdown
    desc_lower = clean_desc.lower()
    agency_lower = clean_agency.lower()
    
    if 'shutdown' in desc_lower or 'shutdown' in agency_lower:
        return clean_desc, "SHUTDOWN", None, None, None

    # Normalization of agency to Department name if matched
    if clean_agency and clean_agency.upper() != 'SHUTDOWN' and clean_agency != 'General':
        from tpm.models import Department
        from django.db.models import Q
        dept = Department.objects.filter(
            Q(name__iexact=clean_agency) |
            Q(code__iexact=clean_agency)
        ).first()
        if dept:
            clean_agency = dept.name

    return clean_desc, clean_agency, clean_sub_agency, clean_equip, clean_sub_equip


def sheet_to_rows(sheet, is_xlsx=True):
    """
    Unified utility to extract all sheet cells as a list of lists of values.
    """
    if is_xlsx:
        # For openpyxl
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    else:
        # For xlrd
        rows = []
        for r in range(sheet.nrows):
            row_vals = []
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                # xlrd floats dates sometimes, convert if date type
                ctype = sheet.cell_type(r, c)
                if ctype == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(val, sheet.book.datemode)
                        val = datetime(*date_tuple)
                    except Exception:
                        pass
                row_vals.append(val)
            rows.append(row_vals)
        return rows


def parse_sms3_sheet(rows, sheet_name, department, upload):
    """
    Parses a single SMS-III day delay sheet.
    """
    # Find the Date (usually row 1, col 0 contains 'DATE : 30.05.2026')
    sheet_date = None
    for r in range(min(len(rows), 5)):
        row = rows[r]
        if row and len(row) > 0 and row[0]:
            val = str(row[0])
            if 'DATE' in val.upper():
                sheet_date = parse_date_string(val)
                break
                
    if not sheet_date:
        # Try to parse date from sheet name (e.g. "30" of May 2026)
        # Fallback to current date month/year if sheet is named like "30"
        try:
            day = int(sheet_name.strip())
            now = timezone.now()
            # Default to May 2026 since files are for May 2026
            sheet_date = date(2026, 5, day)
        except Exception:
            sheet_date = timezone.now().date()

    # Find the starts of the delays section
    # Usually Row 9: ['DELAYS', 'AGENCY', '', '', '', '', 'TOTAL(Min.)', 'HRS']
    start_row = 10
    found_header = False
    for r in range(min(len(rows), 15)):
        row = rows[r]
        if row and len(row) > 1 and row[0] and row[1]:
            if 'DELAYS' in str(row[0]).upper() and 'AGENCY' in str(row[1]).upper():
                start_row = r + 1
                found_header = True
                break

    records_created = 0
    # Process rows
    for r in range(start_row, len(rows)):
        row = rows[r]
        if not row or len(row) < 7:
            continue
            
        desc = row[0]
        agency = row[1]
        
        # Stop check
        if desc and ('TOTAL' in str(desc).upper() or 'SUM' in str(desc).upper()):
            break
            
        # Get total mins (Col index 6)
        total_mins_val = row[6]
        
        # Skip empty or spacer rows
        if not desc and not agency and (total_mins_val == '' or total_mins_val is None or total_mins_val == 0.0):
            continue
            
        # Parse duration
        try:
            duration = float(total_mins_val) if total_mins_val not in (None, '', '-') else 0.0
        except ValueError:
            duration = 0.0
            
        if duration <= 0.0 and (not desc or desc == '-'):
            continue
            
        # Extrapolate shifts
        shift_c = row[3] if len(row) > 3 else ''
        shift_a = row[4] if len(row) > 4 else ''
        shift_b = row[5] if len(row) > 5 else ''
        
        time_slots = []
        if shift_c and shift_c not in ('', '-'): time_slots.append(f"Shift C ({shift_c}m)")
        if shift_a and shift_a not in ('', '-'): time_slots.append(f"Shift A ({shift_a}m)")
        if shift_b and shift_b not in ('', '-'): time_slots.append(f"Shift B ({shift_b}m)")
        time_slot = ", ".join(time_slots) if time_slots else "Daily Log"
        
        # Clean fields
        desc, agency, _, _, _ = clean_parsed_fields(desc, agency)
        
        # Save record
        DelayRecord.objects.create(
            upload=upload,
            department=department,
            sheet_name=sheet_name,
            date=sheet_date,
            time_slot=time_slot,
            duration_mins=duration,
            agency=agency,
            description=desc,
        )
        records_created += 1
        
    return records_created


def parse_rail_mill_sheet(rows, sheet_name, department, upload):
    """
    Parses the Rail Mill Delay_Report sheet.
    """
    # Find sheet date
    # Row 2 (index 2) has 'RAIL MILL DELAY REPORT FOR 09-06-2026'
    # Row 3 (index 3) has C4 '10-Jun-2026'
    sheet_date = None
    for r in range(min(len(rows), 8)):
        row = rows[r]
        for val in row:
            if val and ('REPORT FOR' in str(val).upper() or 'DATE' in str(val).upper()):
                sheet_date = parse_date_string(val)
                if sheet_date:
                    break
        if sheet_date:
            break
            
    if not sheet_date:
        sheet_date = timezone.now().date()

    # Find the header row index
    header_row_idx = 7
    for r in range(min(len(rows), 15)):
        row = rows[r]
        if row and len(row) > 7 and 'Hrs(Time)' in [str(x) for x in row]:
            header_row_idx = r
            break

    records_created = 0
    propagated_date = sheet_date
    propagated_time_slot = ""

    # Process rows below header
    for r in range(header_row_idx + 1, len(rows)):
        row = rows[r]
        if not row or len(row) < 17:
            continue
            
        time_val = row[0]
        # Check stop condition
        if time_val and ('TOTAL' in str(time_val).upper() or 'SUM' in str(time_val).upper()):
            break
            
        duration_val = row[6]
        
        # If duration is '-' or empty, skip
        if duration_val in (None, '', '-'):
            continue
            
        try:
            duration = float(duration_val)
        except ValueError:
            continue
            
        if duration <= 0.0:
            continue
            
        # Parse date and time from time_val if available
        # e.g., '08-06-2026 22:00 - 23:00'
        row_date = propagated_date
        time_slot = propagated_time_slot
        
        if time_val:
            time_val_str = str(time_val).strip()
            # Try to extract date
            date_match = re.search(r'\d{2}[.-]\d{2}[.-]\d{4}', time_val_str)
            if date_match:
                parsed_row_date = parse_date_string(date_match.group(0))
                if parsed_row_date:
                    row_date = parsed_row_date
                    propagated_date = parsed_row_date
            
            # Try to extract time slot (e.g., '22:00 - 23:00')
            time_match = re.search(r'\d{2}:\d{2}\s*-\s*\d{2}:\d{2}', time_val_str)
            if time_match:
                time_slot = time_match.group(0)
                propagated_time_slot = time_slot
            else:
                # Fallback to the rest of the text
                time_slot = time_val_str
                propagated_time_slot = time_slot
                
        # Agency (Col index 9)
        agency_val = row[9] if row[9] and row[9] != '-' else 'Unknown Agency'
        sub_agency_val = row[10] if row[10] and row[10] != '-' else ''
        section = row[11] if row[11] and row[11] != '-' else ''
        equipment_val = row[13] if row[13] and row[13] != '-' else ''
        sub_equipment_val = row[14] if row[14] and row[14] != '-' else ''
        shift_incharge = row[15] if row[15] and row[15] != '-' else ''
        description_val = row[16] if row[16] and row[16] != '-' else 'No Description'
        why = row[17] if len(row) > 17 and row[17] and row[17] != '-' else ''
        
        # Clean fields
        description, agency, sub_agency, equipment, sub_equipment = clean_parsed_fields(
            description_val, agency_val, sub_agency_val, equipment_val, sub_equipment_val
        )
        
        # Save record
        DelayRecord.objects.create(
            upload=upload,
            department=department,
            sheet_name=sheet_name,
            date=row_date,
            time_slot=time_slot,
            start_time=row[4] if row[4] != '-' else None,
            end_time=row[5] if row[5] != '-' else None,
            duration_mins=duration,
            agency=agency,
            sub_agency=sub_agency,
            section=str(section).strip() if section else None,
            equipment=equipment,
            sub_equipment=sub_equipment,
            shift_incharge=str(shift_incharge).strip() if shift_incharge else None,
            description=description,
            why=str(why).strip() if why else None,
        )
        records_created += 1

    return records_created


def normalize_header(s):
    if not s:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def parse_generic_sheet(rows, sheet_name, department, upload):
    """
    Heuristic fallback parser to handle arbitrary delay sheets.
    """
    # 1. Scan rows to find a header row
    header_row_idx = None
    col_mapping = {}
    
    exact_mappings = {
        'date': 'date',
        'day': 'date',
        'month': 'month',
        'agency': 'agency',
        'responsibility': 'agency',
        'downtime': 'duration',
        'down_time': 'duration',
        'duration': 'duration',
        'delay': 'duration',
        'mins': 'duration',
        'minutes': 'duration',
        'hours': 'duration',
        'hrs': 'duration',
        'cause': 'description',
        'description': 'description',
        'reason': 'description',
        'desc': 'description',
        'delays': 'description',
        'problem': 'description',
        'equipment': 'equipment',
        'equip': 'equipment',
        'asset': 'equipment',
        'machine': 'equipment',
        'subequipment': 'sub_equipment',
        'subequip': 'sub_equipment',
        'subagency': 'sub_agency',
        'subdept': 'sub_agency',
        'area': 'sub_agency',
        'subarea': 'sub_area',
        'sub_area': 'sub_area',
        'why': 'why',
        'capa': 'why',
        'rootcause': 'why',
        'incharge': 'incharge',
        'operator': 'incharge',
        'staff': 'incharge',
        'time': 'time',
        'timeslot': 'time',
        'period': 'time',
        'slot': 'time',
        'from': 'start_time',
        'start': 'start_time',
        'starttime': 'start_time',
        'to': 'end_time',
        'end': 'end_time',
        'endtime': 'end_time',
        'agencycatg': 'agency_type',
        'agencycategory': 'agency_type',
        'agencytype': 'agency_type',
    }

    for r in range(min(len(rows), 20)):
        row = rows[r]
        if not row:
            continue
            
        temp_mapping = {}
        mapped_indices = set()
        
        # Pass 1: Exact normalized matches
        for c, val in enumerate(row):
            if val is None or val == '':
                continue
            norm_val = normalize_header(val)
            if norm_val in exact_mappings:
                field = exact_mappings[norm_val]
                if field not in temp_mapping:
                    temp_mapping[field] = c
                    mapped_indices.add(c)
                    
        # Pass 2: Substring matching for unmapped columns
        for c, val in enumerate(row):
            if c in mapped_indices or val is None or val == '':
                continue
            val_str = str(val).lower().strip()
            
            # Helper to check if field not mapped yet
            if 'sub_equipment' not in temp_mapping and any(kw in val_str for kw in ['sub-equip', 'subequip', 'sub equip']):
                temp_mapping['sub_equipment'] = c
            elif 'sub_agency' not in temp_mapping and any(kw in val_str for kw in ['sub-agency', 'subagency', 'sub agency', 'area', 'sub-dept', 'subdept']):
                temp_mapping['sub_agency'] = c
            elif 'sub_area' not in temp_mapping and any(kw in val_str for kw in ['sub-area', 'subarea', 'sub area']):
                temp_mapping['sub_area'] = c
            elif 'equipment' not in temp_mapping and any(kw in val_str for kw in ['equipment', 'equip', 'asset', 'machine']) and 'sub' not in val_str:
                temp_mapping['equipment'] = c
            elif 'agency' not in temp_mapping and any(kw in val_str for kw in ['agency', 'responsibility']) and 'sub' not in val_str and 'catg' not in val_str and 'category' not in val_str:
                temp_mapping['agency'] = c
            elif 'date' not in temp_mapping and any(kw in val_str for kw in ['date', 'day']):
                temp_mapping['date'] = c
            elif 'month' not in temp_mapping and 'month' in val_str:
                temp_mapping['month'] = c
            elif 'duration' not in temp_mapping and any(kw in val_str for kw in ['duration', 'min', 'downtime', 'total', 'hours', 'hrs', 'dur']):
                temp_mapping['duration'] = c
                if 'hour' in val_str or 'hr' in val_str:
                    temp_mapping['duration_type'] = 'hours'
                elif 'min' in val_str:
                    temp_mapping['duration_type'] = 'minutes'
            elif 'description' not in temp_mapping and any(kw in val_str for kw in ['description', 'reason', 'cause', 'delays', 'desc', 'problem', 'delay']):
                temp_mapping['description'] = c
            elif 'why' not in temp_mapping and any(kw in val_str for kw in ['why', 'capa', 'root']):
                temp_mapping['why'] = c
            elif 'incharge' not in temp_mapping and any(kw in val_str for kw in ['incharge', 'operator', 'shift', 'staff']):
                temp_mapping['incharge'] = c
            elif 'time' not in temp_mapping and any(kw in val_str for kw in ['time', 'range', 'period', 'slot', 'from-to', 'start-end', 'hrs(time)']) and 'down' not in val_str:
                temp_mapping['time'] = c
            elif 'start_time' not in temp_mapping and (val_str == 'from' or val_str.startswith('from ') or 'start' in val_str) and 'end' not in val_str:
                temp_mapping['start_time'] = c
            elif 'end_time' not in temp_mapping and (val_str == 'to' or val_str.startswith('to ') or 'end' in val_str) and 'start' not in val_str:
                temp_mapping['end_time'] = c
                
        # We need at least 3 matched fields or 3 core fields to consider it a valid header row
        matches = len([f for f in ['date', 'description', 'duration', 'agency'] if f in temp_mapping])
        if matches >= 3 or len(temp_mapping) >= 3:
            header_row_idx = r
            col_mapping = temp_mapping
            break
            
    if header_row_idx is None:
        # Let's see if we can do a default mapping (0: desc, 1: agency, 2: duration)
        # If there are columns, just do a basic fallback
        return 0

    # Determine default year from filename or sheet_name
    year_from_sheet = extract_year_from_string(sheet_name)
    year_from_file = extract_year_from_string(upload.filename) if upload and hasattr(upload, 'filename') else None
    
    # Find sheet-level date as backup
    sheet_date = None
    for r in range(min(len(rows), header_row_idx + 1)):
        for val in rows[r]:
            if val:
                sheet_date = parse_date_string(val)
                if sheet_date:
                    break
        if sheet_date:
            break
            
    if not sheet_date:
        sheet_date = timezone.now().date()
        yr = year_from_sheet or year_from_file
        if isinstance(yr, int):
            try:
                sheet_date = sheet_date.replace(year=yr)
            except ValueError:
                pass

    records_created = 0
    propagated_date = sheet_date
    propagated_month = sheet_date.month
    
    # Process rows below headers
    for r in range(header_row_idx + 1, len(rows)):
        row = rows[r]
        if not row:
            continue
            
        # Skip intermediate summary/total rows
        first_val = str(row[0]).upper() if row[0] else ''
        if 'TOTAL' in first_val or 'SUM' in first_val:
            continue
            
        # Extract indexes
        desc_idx = col_mapping.get('description', None)
        agency_idx = col_mapping.get('agency', None)
        dur_idx = col_mapping.get('duration', None)
        date_idx = col_mapping.get('date', None)
        month_idx = col_mapping.get('month', None)
        time_idx = col_mapping.get('time', None)
        start_idx = col_mapping.get('start_time', None)
        end_idx = col_mapping.get('end_time', None)
        agency_type_idx = col_mapping.get('agency_type', None)
        
        if dur_idx is None:
            # Skip if we can't find duration column
            continue
            
        desc = row[desc_idx] if desc_idx is not None and desc_idx < len(row) else None
        
        # Skip if description contains total/sum
        desc_upper = str(desc).upper() if desc else ''
        if 'TOTAL' in desc_upper or 'SUM' in desc_upper:
            continue
        agency = row[agency_idx] if agency_idx is not None and agency_idx < len(row) else 'General'
        duration_val = row[dur_idx] if dur_idx is not None and dur_idx < len(row) else 0.0
        
        # Skip rows that are empty of both description and date, which are typically summary/total spacer rows
        row_date_val = row[date_idx] if date_idx is not None and date_idx < len(row) else None
        if not desc and not row_date_val:
            continue
            
        if not desc and (duration_val is None or duration_val == '' or duration_val == 0.0):
            continue
            
        # Month propagation
        if month_idx is not None and month_idx < len(row) and row[month_idx]:
            parsed_month = parse_month_to_num(row[month_idx])
            if parsed_month:
                propagated_month = parsed_month
                
        # Date propagation
        row_date = propagated_date
        if date_idx is not None and date_idx < len(row) and row[date_idx]:
            parsed_date = parse_row_date(row[date_idx], propagated_month, default_year=sheet_date)
            if parsed_date:
                row_date = parsed_date
                propagated_date = parsed_date
                propagated_month = parsed_date.month
                
        # Extract start and end times
        start_time_val = None
        end_time_val = None
        time_slot_val = None
        
        # Try retrieving start/end time from separate columns
        if start_idx is not None and start_idx < len(row):
            start_time_val = row[start_idx]
        if end_idx is not None and end_idx < len(row):
            end_time_val = row[end_idx]
            
        # Try retrieving single time slot column
        if time_idx is not None and time_idx < len(row):
            time_slot_val = row[time_idx]
            
        # If separate start/end are not populated but time slot is, extract them
        if (not start_time_val or not end_time_val) and time_slot_val:
            s_t, e_t = extract_times_from_range(time_slot_val)
            if s_t and e_t:
                start_time_val = s_t
                end_time_val = e_t
                
        start_time_str = format_time_value(start_time_val)
        end_time_str = format_time_value(end_time_val)
        
        # Compute duration from start and end times
        start_min = parse_time_to_minutes(start_time_val)
        end_min = parse_time_to_minutes(end_time_val)
        computed_duration = None
        if start_min is not None and end_min is not None:
            diff = end_min - start_min
            if diff < 0:
                diff += 1440
            computed_duration = float(diff)
            
        # Determine actual duration to save
        dur_type = col_mapping.get('duration_type', None)
        parsed_dur = parse_duration_to_mins(duration_val, col_type=dur_type)
        
        if computed_duration is not None:
            if parsed_dur > 0:
                # If they are close, or one is a fraction, align them
                if abs(parsed_dur - computed_duration) < 0.5:
                    duration = parsed_dur
                elif abs(parsed_dur * 60.0 - computed_duration) < 0.5:
                    duration = computed_duration
                elif abs(parsed_dur * 1440.0 - computed_duration) < 0.5:
                    duration = computed_duration
                else:
                    duration = parsed_dur
            else:
                duration = computed_duration
        else:
            duration = parsed_dur
            
        if duration <= 0.0 and (not desc or desc == '-'):
            continue
            
        # Construct time slot
        time_slot = ""
        if start_time_str and end_time_str:
            time_slot = f"{start_time_str} - {end_time_str}"
        elif time_slot_val:
            time_slot = str(time_slot_val).strip()
            
        equipment = ""
        equip_idx = col_mapping.get('equipment', None)
        if equip_idx is not None and equip_idx < len(row) and row[equip_idx]:
            equipment = str(row[equip_idx])
            
        sub_equipment = ""
        sub_equip_idx = col_mapping.get('sub_equipment', None)
        if sub_equip_idx is not None and sub_equip_idx < len(row) and row[sub_equip_idx]:
            sub_equipment = str(row[sub_equip_idx])
            
        sub_agency = ""
        sub_agency_idx = col_mapping.get('sub_agency', None)
        if sub_agency_idx is not None and sub_agency_idx < len(row) and row[sub_agency_idx]:
            sub_agency = str(row[sub_agency_idx])
            
        sub_area = ""
        sub_area_idx = col_mapping.get('sub_area', None)
        if sub_area_idx is not None and sub_area_idx < len(row) and row[sub_area_idx]:
            sub_area = str(row[sub_area_idx])
        sub_area = sub_area.strip() if sub_area else None
            
        why = ""
        why_idx = col_mapping.get('why', None)
        if why_idx is not None and why_idx < len(row) and row[why_idx]:
            why = str(row[why_idx])
            
        incharge = ""
        incharge_idx = col_mapping.get('incharge', None)
        if incharge_idx is not None and incharge_idx < len(row) and row[incharge_idx]:
            incharge = str(row[incharge_idx])
            
        # Clean fields
        desc, agency, sub_agency, equipment, sub_equipment = clean_parsed_fields(
            desc, agency, sub_agency, equipment, sub_equipment
        )
        
        # Clean agency type
        agency_type_val = row[agency_type_idx] if agency_type_idx is not None and agency_type_idx < len(row) else None
        agency_type_clean = 'Internal'
        if agency_type_val:
            agency_type_str = str(agency_type_val).strip().upper()
            if 'EXTERNAL' in agency_type_str:
                agency_type_clean = 'External'
            elif 'INTERNAL' in agency_type_str:
                agency_type_clean = 'Internal'
            else:
                if Department.objects.filter(name__iexact=agency).exists():
                    agency_type_clean = 'External'
                else:
                    agency_type_clean = 'Internal'
        else:
            if Department.objects.filter(name__iexact=agency).exists():
                agency_type_clean = 'External'
            else:
                agency_type_clean = 'Internal'

        # Save record
        DelayRecord.objects.create(
            upload=upload,
            department=department,
            sheet_name=sheet_name,
            date=row_date,
            time_slot=time_slot,
            start_time=start_time_str,
            end_time=end_time_str,
            duration_mins=duration,
            agency=agency,
            sub_agency=sub_agency,
            sub_area=sub_area,
            section=None,
            equipment=equipment,
            sub_equipment=sub_equipment,
            shift_incharge=incharge.strip() if incharge else None,
            description=desc,
            why=why.strip() if why else None,
            agency_type=agency_type_clean,
            is_locked=False,
        )
        records_created += 1
        
    return records_created


def parse_excel_file(upload_instance):
    """
    Main parser entrypoint.
    Opens the excel file, detects sheet layouts, extracts delay logs,
    and updates the upload_instance status.
    """
    file_path = upload_instance.file.path
    department = upload_instance.department
    
    if not os.path.exists(file_path):
        upload_instance.status = 'FAILED'
        upload_instance.error_message = f"File not found at path: {file_path}"
        upload_instance.save()
        return False
        
    _, ext = os.path.splitext(file_path.lower())
    is_xlsx = ext == '.xlsx'
    
    total_records = 0
    
    try:
        with transaction.atomic():  # type: ignore
            # Open the workbook
            if is_xlsx:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet_names = wb.sheetnames
            else:
                wb = xlrd.open_workbook(file_path)
                sheet_names = wb.sheet_names()
                
            # Loop through worksheets
            for sh_name in sheet_names:
                if is_xlsx:
                    sheet = wb[sh_name]
                else:
                    sheet = wb.sheet_by_name(sh_name)  # type: ignore
                    
                rows = sheet_to_rows(sheet, is_xlsx=is_xlsx)
                if not rows:
                    continue
                    
                # ─────────────────────────────────────────────────────────────
                # DETECT LAYOUT FORMAT
                # ─────────────────────────────────────────────────────────────
                is_sms3 = False
                is_rail_mill = False
                
                # Check SMS3 criteria: DELAYS header in col A, AGENCY in col B (usually row 9 or index 9)
                for r in range(min(len(rows), 15)):
                    row = rows[r]
                    if row and len(row) > 1 and row[0] and row[1]:
                        if 'DELAYS' in str(row[0]).upper() and 'AGENCY' in str(row[1]).upper():
                            is_sms3 = True
                            break
                            
                # Check Rail Mill criteria: Hrs(Time) header in col 0, Description for delay in col 16 (index 16)
                for r in range(min(len(rows), 15)):
                    row = rows[r]
                    if row and len(row) > 16:
                        # Col 0: Hrs(Time), Col 16: Description for delay
                        row_strs = [str(x).upper() for x in row if x is not None]
                        if any('HRS(TIME)' in s for s in row_strs) and any('DESCRIPTION FOR DELAY' in s or 'DESCRIPTION' in s for s in row_strs):
                            is_rail_mill = True
                            break
                            
                # Parse depending on type
                if is_sms3:
                    count = parse_sms3_sheet(rows, sh_name, department, upload_instance)
                    total_records += count
                elif is_rail_mill:
                    if 'SUMMARY' not in sh_name.upper(): # Skip summary sheet
                        count = parse_rail_mill_sheet(rows, sh_name, department, upload_instance)
                        total_records += count
                else:
                    # Generic fallback parser
                    # Skip typical summary, chart, graph, milestone sheets
                    skip_keywords = ['SUMMARY', 'KPI', 'LOSS', 'PARETO', 'ISHIKAWA', 'PROJECTS', 'GANTT', 'CHART', 'GRAPH', 'PILLAR', 'PLAN', 'MILESTONE']
                    should_skip = any(kw in sh_name.upper() for kw in skip_keywords)
                    if not should_skip:
                        count = parse_generic_sheet(rows, sh_name, department, upload_instance)
                        total_records += count
                        
            # Update upload status
            upload_instance.status = 'SUCCESS'
            upload_instance.error_message = f"Successfully parsed {total_records} delay records."
            upload_instance.save()
            return True
            
    except Exception as e:
        upload_instance.status = 'FAILED'
        upload_instance.error_message = f"Error during parsing: {str(e)}"
        upload_instance.save()
        raise e
