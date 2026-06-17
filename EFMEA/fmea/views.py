import json
import openpyxl
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db.models import Count, Q

from tpm.models import Department, User
from portal.utils.decorators import dept_visibility_required, module_access_required
from .models import FMEARecord, FMEAAuditLog, FMEAExcelUpload, FMEACriticalSpare

@login_required
@dept_visibility_required
@module_access_required('FMEA')
def fmea_dashboard(request, dept_id):
    if int(dept_id) == 0:
        class DummyDept:
            id = 0
            name = "Overall Plant"
            code = "Overall"
        active_dept = DummyDept()
        
        # Retrieve all Excel uploads across all departments
        uploads = FMEAExcelUpload.objects.all().order_by('-sheet_date', '-uploaded_at')
        
        # Retrieve all records across all departments
        records = FMEARecord.objects.all()
    else:
        active_dept = get_object_or_404(Department, id=dept_id)
        
        # Retrieve all Excel uploads for this department
        uploads = FMEAExcelUpload.objects.filter(department=active_dept).order_by('-sheet_date', '-uploaded_at')
        
        # Retrieve all records for this department (combining all files & manual entries)
        records = FMEARecord.objects.filter(department=active_dept)
            
    # Calculate RPN categories (overall)
    total_risks = records.count()
    high_risks = records.filter(rpn__gte=301).count()
    medium_risks = records.filter(rpn__gte=101, rpn__lte=300).count()
    low_risks = records.filter(rpn__gte=1, rpn__lte=100).count()
    
    # Top 5 critical risks across all files
    top_critical = records.order_by('-rpn')[:5]
    
    # Build comparative analysis list
    sheet_comparisons = []
    
    if int(dept_id) == 0:
        # Group by department for overall view
        all_depts = Department.objects.all().order_by('name')
        for d in all_depts:
            d_records = records.filter(department=d)
            if d_records.exists():
                sheet_comparisons.append({
                    'id': d.id,
                    'name': f"{d.name} ({d.code})",
                    'date': '—',
                    'total': d_records.count(),
                    'high': d_records.filter(rpn__gte=301).count(),
                    'medium': d_records.filter(rpn__gte=101, rpn__lte=300).count(),
                    'low': d_records.filter(rpn__gte=1, rpn__lte=100).count(),
                    'dept_id': d.id,
                    'dept_code': d.code,
                })
    else:
        # 1. Check uploads
        for u in uploads:
            recs = FMEARecord.objects.filter(excel_upload=u)
            if recs.exists():
                sheet_comparisons.append({
                    'id': u.id,
                    'name': u.filename.replace('.xlsx', ''),
                    'date': u.sheet_date.strftime('%d.%m.%Y') if u.sheet_date else '—',
                    'total': recs.count(),
                    'high': recs.filter(rpn__gte=301).count(),
                    'medium': recs.filter(rpn__gte=101, rpn__lte=300).count(),
                    'low': recs.filter(rpn__gte=1, rpn__lte=100).count(),
                    'dept_id': u.department.id,
                })
                
        # 2. Check default manual list
        manual_recs = FMEARecord.objects.filter(department=active_dept, excel_upload=None)
            
        if manual_recs.exists():
            sheet_comparisons.append({
                'id': 'new_file',
                'name': 'Default manual list',
                'date': '—',
                'total': manual_recs.count(),
                'high': manual_recs.filter(rpn__gte=301).count(),
                'medium': manual_recs.filter(rpn__gte=101, rpn__lte=300).count(),
                'low': manual_recs.filter(rpn__gte=1, rpn__lte=100).count(),
                'dept_id': active_dept.id,
            })
        
    if int(dept_id) == 0:
        comp_labels = [item.get('dept_code', '') for item in sheet_comparisons]
    else:
        comp_labels = [item['name'] for item in sheet_comparisons]
    comp_ids = [item['id'] for item in sheet_comparisons]
    comp_high = [item['high'] for item in sheet_comparisons]
    comp_medium = [item['medium'] for item in sheet_comparisons]
    comp_low = [item['low'] for item in sheet_comparisons]
    
    # Compile detailed failure mode RPN comparison data for each sheet/file
    detailed_data = {}
    for item in sheet_comparisons:
        sid = item['id']
        if int(dept_id) == 0:
            detailed_data[str(sid)] = []
        else:
            if sid == 'new_file':
                recs = FMEARecord.objects.none() if int(dept_id) == 0 else FMEARecord.objects.filter(department=active_dept, excel_upload=None).order_by('sn', 'id')
            else:
                recs = FMEARecord.objects.filter(excel_upload_id=sid).order_by('sn', 'id')
            detailed_data[str(sid)] = [
                {
                    'failure_mode': r.potential_failure_mode[:20] + ('...' if len(r.potential_failure_mode) > 20 else ''),
                    'full_failure_mode': r.potential_failure_mode,
                    'rpn': r.rpn,
                    'action_rpn': r.action_rpn if r.action_rpn is not None else 0
                }
                for r in recs
            ]
        
    # Mitigation Status Overview for donut chart (all combined)
    status_counts = records.values('status').annotate(count=Count('id'))
    status_map = {'Completed': 0, 'Under Progress': 0, 'Not Started': 0}
    for s in status_counts:
        st = s['status'] or 'Not Started'
        if st in status_map:
            status_map[st] += s['count']
        else:
            status_map['Not Started'] += s['count']
            
    status_chart_data = [status_map['Completed'], status_map['Under Progress'], status_map['Not Started']]
    
    context = {
        'active_dept_id': dept_id,
        'active_dept': active_dept,
        'active_module': 'FMEA',
        'active_tab': 'dashboard',
        
        # Metrics
        'total_risks': total_risks,
        'high_risks': high_risks,
        'medium_risks': medium_risks,
        'low_risks': low_risks,
        
        # Top Table
        'top_critical': top_critical,
        
        # Comparative Chart Data
        'comp_labels': json.dumps(comp_labels),
        'comp_ids': json.dumps(comp_ids),
        'comp_high': json.dumps(comp_high),
        'comp_medium': json.dumps(comp_medium),
        'comp_low': json.dumps(comp_low),
        'detailed_data_json': json.dumps(detailed_data),
        'status_chart_data': json.dumps(status_chart_data),
        
        # Table of comparative sheets
        'sheet_comparisons': sheet_comparisons,
    }
    return render(request, 'fmea/dashboard.html', context)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def risk_identification(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    departments = Department.objects.all().order_by('name')
    
    # Gather potential risk owners
    owners = User.objects.filter(is_active=True).order_by('first_name')
    
    # Retrieve uploads for file association dropdown
    uploads = FMEAExcelUpload.objects.filter(department=active_dept).order_by('-sheet_date', '-uploaded_at')
    selected_upload_id = request.GET.get('upload_id')
    
    # Fetch unique main_equipment and sub_equipment values from database for this department
    db_equipments = FMEARecord.objects.filter(department=active_dept).exclude(main_equipment='').values_list('main_equipment', flat=True).distinct()
    db_sub_equipments = FMEARecord.objects.filter(department=active_dept).exclude(sub_equipment='').values_list('sub_equipment', flat=True).distinct()
    
    # Predefined equipment lists (common categories)
    predefined_sub_equipments = [
        'CABLE', 'EBT', 'PURGING', 'SNORKEL', 'DSL', 'GUIDE', 'GUIDE ROLL',
        'ELECTRODE', 'SEN', 'RAIL TRACK', 'KT BLOCK', 'HOOTER', 'CYLINDER',
        'ROLLER', 'MOTOR', 'CASTING POWDER', 'HYDRAULIC LINE', 'CHOCKING',
        'ROPE DRUM', 'PANEL', 'GUNNING MACHINE', 'REFRACTORY', 'LADLE GATE',
        'LADLE LATE', 'CHEMISTRY', 'WITHDRAWAL', 'GANTRY', 'GANATRY'
    ]
    
    predefined_equipments = [
        'Overhead Crane', 'Ladle Furnace', 'Continuous Caster', 'Electric Arc Furnace',
        'Argon Rinsing Station', 'Vacuum Degasser', 'Tundish', 'Mold Assembly',
        'Withdrawal Straightener', 'Runout Table', 'Reheating Furnace', 'Roughing Mill',
        'Finishing Mill', 'Cooling Bed', 'Straightening Machine', 'Cold Shear'
    ]
    
    equipments = sorted(list(set(list(db_equipments) + predefined_equipments)))
    sub_equipments = sorted(list(set([s.upper() for s in db_sub_equipments] + predefined_sub_equipments)))

    context = {
        'active_dept_id': dept_id,
        'active_dept': active_dept,
        'active_module': 'FMEA',
        'active_tab': 'identification',
        'departments': departments,
        'owners': owners,
        'today': datetime.now().strftime('%Y-%m-%d'),
        'uploads': uploads,
        'selected_upload_id': selected_upload_id,
        'equipments': equipments,
        'sub_equipments': sub_equipments,
    }
    return render(request, 'fmea/identification.html', context)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def risk_register(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    
    # Filter inputs
    dept_filter = request.GET.get('department_id')
    owner_filter = request.GET.get('owner')
    search_desc = request.GET.get('description', '').strip()
    
    records = FMEARecord.objects.filter(department=active_dept)
    if dept_filter:
        records = records.filter(department_id=dept_filter)
    if owner_filter:
        records = records.filter(risk_owner=owner_filter)
    if search_desc:
        records = records.filter(Q(main_equipment__icontains=search_desc) | Q(potential_failure_mode__icontains=search_desc) | Q(potential_effects__icontains=search_desc))
        
    departments = Department.objects.all().order_by('name')
    owners_list = FMEARecord.objects.filter(department=active_dept).values_list('risk_owner', flat=True).distinct()
    owners_list = [o for o in owners_list if o]
    
    # Prepare list for Alpine/Template
    records_data = []
    for r in records:
        # Decode consequences list
        consequences = [c.strip() for c in r.potential_effects.split('\n') if c.strip()]
        
        # Decode action plan JSON list
        try:
            action_plan_items = json.loads(r.action_plan_data or '[]')
        except:
            action_plan_items = []

        if not action_plan_items:
            # On-the-fly migration from legacy quarterly fields
            legacy_items = []
            for q, field_val, target_val, status_val in [
                ('q1', r.mitigation_q1, r.mitigation_q1_target, r.mitigation_q1_status),
                ('q2', r.mitigation_q2, r.mitigation_q2_target, r.mitigation_q2_status),
                ('q3', r.mitigation_q3, r.mitigation_q3_target, r.mitigation_q3_status),
                ('q4', r.mitigation_q4, r.mitigation_q4_target, r.mitigation_q4_status),
            ]:
                try:
                    acts = json.loads(field_val or '[]')
                    for a in acts:
                        if a.strip():
                            legacy_items.append({
                                'action': a.strip(),
                                'responsibility': r.risk_owner or '',
                                'plan_term': 'Short' if q in ['q1', 'q2'] else 'Medium',
                                'target_date': target_val.strftime('%Y-%m-%d') if target_val else '',
                                'status': status_val or 'Not Started'
                            })
                except:
                    pass
            if legacy_items:
                action_plan_items = legacy_items
                # Save to database to persist
                r.action_plan_data = json.dumps(action_plan_items)
                r.save(update_fields=['action_plan_data'])

        # Remarks
        remarks = [rem.strip() for rem in r.status_remarks.split('\n') if rem.strip()]
        
        records_data.append({
            'record': r,
            'consequences': consequences,
            'action_plan_items': action_plan_items,
            'remarks': remarks,
        })
        
    context = {
        'active_dept_id': dept_id,
        'active_dept': active_dept,
        'active_module': 'FMEA',
        'active_tab': 'register',
        'records_data': records_data,
        'departments': departments,
        'owners_list': owners_list,
        'selected_dept': dept_filter,
        'selected_owner': owner_filter,
        'search_desc': search_desc,
    }
    return render(request, 'fmea/register.html', context)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def risk_report(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    
    # Retrieve all Excel uploads for this department
    uploads = FMEAExcelUpload.objects.filter(department=active_dept).order_by('-sheet_date', '-uploaded_at')
    
    upload_id = request.GET.get('upload_id')
    selected_upload_id = 'new_file'
    selected_upload = None
    
    if upload_id and upload_id != 'new_file':
        try:
            selected_upload_id = int(upload_id)
            selected_upload = get_object_or_404(FMEAExcelUpload, id=selected_upload_id, department=active_dept)
            records = FMEARecord.objects.filter(department=active_dept, excel_upload=selected_upload)
        except ValueError:
            records = FMEARecord.objects.filter(department=active_dept, excel_upload=None)
    else:
        records = FMEARecord.objects.filter(department=active_dept, excel_upload=None)
        
    dept_filter = request.GET.get('department_id')
    rpn_sort = request.GET.get('sort_by', 'rpn_desc')
    quarter_filter = request.GET.get('quarter', 'all')
    
    if dept_filter:
        records = records.filter(department_id=dept_filter)
        
    # Sort
    if rpn_sort == 'rpn_desc':
        records = records.order_by('-rpn')
    elif rpn_sort == 'rpn_asc':
        records = records.order_by('rpn')
    elif rpn_sort == 'sn':
        records = records.order_by('sn', 'id')
        
    records_data = []
    for r in records:
        consequences = [c.strip() for c in r.potential_effects.split('\n') if c.strip()]
        
        # Quarter specific details
        q_mitigation = ""
        q_target = ""
        q_status = ""
        
        if quarter_filter == 'q1':
            try:
                actions = json.loads(r.mitigation_q1 or '[]')
                q_mitigation = "\n".join(actions)
            except:
                q_mitigation = r.recommended_actions
            q_target = r.mitigation_q1_target.strftime('%Y-%m-%d') if r.mitigation_q1_target else ""
            q_status = r.mitigation_q1_status or r.status
        elif quarter_filter == 'q2':
            try:
                actions = json.loads(r.mitigation_q2 or '[]')
                q_mitigation = "\n".join(actions)
            except:
                q_mitigation = r.recommended_actions
            q_target = r.mitigation_q2_target.strftime('%Y-%m-%d') if r.mitigation_q2_target else ""
            q_status = r.mitigation_q2_status or r.status
        elif quarter_filter == 'q3':
            try:
                actions = json.loads(r.mitigation_q3 or '[]')
                q_mitigation = "\n".join(actions)
            except:
                q_mitigation = r.recommended_actions
            q_target = r.mitigation_q3_target.strftime('%Y-%m-%d') if r.mitigation_q3_target else ""
            q_status = r.mitigation_q3_status or r.status
        elif quarter_filter == 'q4':
            try:
                actions = json.loads(r.mitigation_q4 or '[]')
                q_mitigation = "\n".join(actions)
            except:
                q_mitigation = r.recommended_actions
            q_target = r.mitigation_q4_target.strftime('%Y-%m-%d') if r.mitigation_q4_target else ""
            q_status = r.mitigation_q4_status or r.status
        else: # All
            q_mitigation = r.recommended_actions
            q_target = r.target_date.strftime('%Y-%m-%d') if r.target_date else ""
            q_status = r.status
            
        records_data.append({
            'record': r,
            'consequences': consequences,
            'q_mitigation': q_mitigation,
            'q_target': q_target,
            'q_status': q_status,
        })
        
    # Calculate rowspans for Equipment/Function/Failure grouping
    n = len(records_data)
    i = 0
    while i < n:
        item = records_data[i]
        r = item['record']
        
        # Find how many subsequent records match
        match_count = 1
        j = i + 1
        while j < n:
            next_item = records_data[j]
            next_r = next_item['record']
            if (next_r.main_equipment == r.main_equipment and 
                next_r.main_equipment_function == r.main_equipment_function and 
                next_r.functional_failure == r.functional_failure):
                match_count += 1
                j += 1
            else:
                break
        
        # Set rowspan values
        item['equip_rowspan'] = match_count
        item['is_equip_lead'] = True
        
        for k in range(i + 1, j):
            records_data[k]['equip_rowspan'] = 0
            records_data[k]['is_equip_lead'] = False
            
        i = j
        
    can_edit = True

    departments = Department.objects.all().order_by('name')
    
    if selected_upload:
        sheet_metadata = {
            'key_contact': selected_upload.key_contact,
            'core_team': selected_upload.core_team,
            'objective': selected_upload.objective,
            'ref_no': selected_upload.ref_no,
            'date_str': selected_upload.sheet_date.strftime('%Y-%m-%d') if selected_upload.sheet_date else '',
        }
    else:
        sheet_metadata = {
            'key_contact': '',
            'core_team': '',
            'objective': '',
            'ref_no': '',
            'date_str': datetime.now().strftime('%Y-%m-%d'),
        }
    
    if selected_upload:
        critical_spares = FMEACriticalSpare.objects.filter(excel_upload=selected_upload, department=active_dept)
    else:
        critical_spares = FMEACriticalSpare.objects.filter(excel_upload=None, department=active_dept)
        
    # Fetch unique main_equipment and sub_equipment values from database for this department
    db_equipments = FMEARecord.objects.filter(department=active_dept).exclude(main_equipment='').values_list('main_equipment', flat=True).distinct()
    db_sub_equipments = FMEARecord.objects.filter(department=active_dept).exclude(sub_equipment='').values_list('sub_equipment', flat=True).distinct()
    
    # Predefined equipment lists (common categories)
    predefined_sub_equipments = [
        'CABLE', 'EBT', 'PURGING', 'SNORKEL', 'DSL', 'GUIDE', 'GUIDE ROLL',
        'ELECTRODE', 'SEN', 'RAIL TRACK', 'KT BLOCK', 'HOOTER', 'CYLINDER',
        'ROLLER', 'MOTOR', 'CASTING POWDER', 'HYDRAULIC LINE', 'CHOCKING',
        'ROPE DRUM', 'PANEL', 'GUNNING MACHINE', 'REFRACTORY', 'LADLE GATE',
        'LADLE LATE', 'CHEMISTRY', 'WITHDRAWAL', 'GANTRY', 'GANATRY'
    ]
    
    predefined_equipments = [
        'Overhead Crane', 'Ladle Furnace', 'Continuous Caster', 'Electric Arc Furnace',
        'Argon Rinsing Station', 'Vacuum Degasser', 'Tundish', 'Mold Assembly',
        'Withdrawal Straightener', 'Runout Table', 'Reheating Furnace', 'Roughing Mill',
        'Finishing Mill', 'Cooling Bed', 'Straightening Machine', 'Cold Shear'
    ]
    
    equipments = sorted(list(set(list(db_equipments) + predefined_equipments)))
    sub_equipments = sorted(list(set([s.upper() for s in db_sub_equipments] + predefined_sub_equipments)))

    context = {
        'active_dept_id': dept_id,
        'active_dept': active_dept,
        'active_module': 'FMEA',
        'active_tab': 'report',
        'records_data': records_data,
        'departments': departments,
        'selected_dept': dept_filter,
        'sort_by': rpn_sort,
        'selected_quarter': quarter_filter,
        'can_edit': can_edit,
        
        'uploads': uploads,
        'selected_upload_id': selected_upload_id,
        'sheet_metadata': sheet_metadata,
        'critical_spares': critical_spares,
        'equipments': equipments,
        'sub_equipments': sub_equipments,
    }
    return render(request, 'fmea/report.html', context)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def risk_history(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    
    # Query uploaded excels annotated with record counts
    uploaded_files = FMEAExcelUpload.objects.filter(department=active_dept, is_manual=False).annotate(
        num_records=Count('records')
    ).select_related('uploaded_by').order_by('-uploaded_at')
    
    # Query manually created sheets
    manual_sheets = FMEAExcelUpload.objects.filter(department=active_dept, is_manual=True).annotate(
        num_records=Count('records')
    ).select_related('uploaded_by').order_by('-uploaded_at')
    
    # Query manual records count for deletion functionality (scratchpad records)
    manual_count = FMEARecord.objects.filter(department=active_dept, excel_upload=None).count()
    
    context = {
        'active_dept_id': dept_id,
        'active_dept': active_dept,
        'active_module': 'FMEA',
        'active_tab': 'history',
        'uploaded_files': uploaded_files,
        'manual_sheets': manual_sheets,
        'manual_count': manual_count,
    }
    return render(request, 'fmea/history.html', context)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def health_checklist(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    context = {
        'active_dept_id': dept_id,
        'active_dept': active_dept,
        'active_module': 'FMEA',
        'active_tab': 'checklist',
    }
    return render(request, 'fmea/checklist.html', context)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def monsoon_report(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    context = {
        'active_dept_id': dept_id,
        'active_dept': active_dept,
        'active_module': 'FMEA',
        'active_tab': 'monsoon',
    }
    return render(request, 'fmea/monsoon.html', context)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def save_risk(request, dept_id, record_id=None):
    active_dept = get_object_or_404(Department, id=dept_id)
    if request.method != 'POST':
        return redirect('fmea:register', dept_id=dept_id)
        
    if record_id:
        record = get_object_or_404(FMEARecord, id=record_id, department=active_dept)
        action_type = 'UPDATED'
    else:
        record = FMEARecord(department=active_dept)
        action_type = 'CREATED'
        
    # Bind form inputs
    record.main_equipment = request.POST.get('main_equipment', '').strip()
    record.risk_owner = request.POST.get('risk_owner', '').strip()
    record.main_equipment_function = request.POST.get('main_equipment_function', '').strip()
    record.functional_failure = request.POST.get('functional_failure', '').strip()
    record.sub_equipment = request.POST.get('sub_equipment', '').strip()
    record.component = request.POST.get('component', '').strip()
    record.component_function = request.POST.get('component_function', '').strip()
    record.potential_failure_mode = request.POST.get('potential_failure_mode', '').strip()
    record.contingency_plan = request.POST.get('contingency_plan', '').strip()
    record.potential_causes = request.POST.get('potential_causes', '').strip()
    record.current_controls = request.POST.get('current_controls', '').strip()
    record.recommended_actions = request.POST.get('recommended_actions', '').strip()
    
    # Consolidate Event Consequences list
    consequences = request.POST.getlist('event_consequences[]')
    record.potential_effects = "\n".join([c.strip() for c in consequences if c.strip()])
    
    # Date formatting
    date_str = request.POST.get('identification_date', '').strip()
    if date_str:
        try:
            record.identification_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
            
    # Ratings
    try:
        record.severity = int(request.POST.get('severity', 1))
    except ValueError:
        record.severity = 1
    try:
        record.occurrence = int(request.POST.get('occurrence', 1))
    except ValueError:
        record.occurrence = 1
    try:
        record.detection = int(request.POST.get('detection', 1))
    except ValueError:
        record.detection = 1
        
    # Handle File Association if it is a new record
    if action_type == 'CREATED':
        association_type = request.POST.get('association_type', 'default')
        if association_type == 'new':
            new_file_name = request.POST.get('new_file_name', '').strip()
            if not new_file_name:
                new_file_name = f"Manual_Entry_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            if not new_file_name.endswith('.xlsx'):
                new_file_name += '.xlsx'
                
            upload = FMEAExcelUpload.objects.create(
                department=active_dept,
                filename=new_file_name,
                sheet_date=record.identification_date or datetime.now().date(),
                uploaded_by=request.user,
                is_manual=True,
                key_contact=record.risk_owner,
                core_team=f"{active_dept.name} Operations Team",
                objective=f"Minimize Production Loss Time due to {active_dept.name} issues"
            )
            record.excel_upload = upload
        elif association_type == 'existing':
            existing_file_id = request.POST.get('existing_file_id', '').strip()
            if existing_file_id:
                try:
                    upload = FMEAExcelUpload.objects.get(id=int(existing_file_id), department=active_dept)
                    record.excel_upload = upload
                except (ValueError, FMEAExcelUpload.DoesNotExist):
                    pass
        else:
            record.excel_upload = None

    # Generate automatic Serial Number if empty
    if not record.sn:
        max_sn = FMEARecord.objects.filter(department=active_dept, excel_upload=record.excel_upload).count() + 1
        record.sn = f"{max_sn:02d}"
        
    record.save()
    
    # Audit log
    FMEAAuditLog.objects.create(
        record=record,
        user=request.user,
        action=action_type,
        details=f"Risk description: {record.potential_failure_mode[:100]}... Initial RPN: {record.rpn}"
    )
    
    messages.success(request, f"Risk {record.risk_id} successfully saved.")
    
    if record.excel_upload:
        from django.urls import reverse
        return redirect(reverse('fmea:report', args=[dept_id]) + f'?upload_id={record.excel_upload.id}')
    else:
        return redirect('fmea:report', dept_id=dept_id)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def save_mitigation(request, dept_id, record_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    record = get_object_or_404(FMEARecord, id=record_id, department=active_dept)
    
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)
        
    # Get action plan data
    action_plan_data_str = request.POST.get('action_plan_data', '[]')
    try:
        # validate json
        json.loads(action_plan_data_str)
        record.action_plan_data = action_plan_data_str
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Invalid action plan data: {str(e)}'}, status=400)
            
    # Status
    status = request.POST.get('status', 'Not Started').strip()
    
    # Remarks
    remarks_list = request.POST.getlist('status_remarks[]')
    remarks = "\n".join([r.strip() for r in remarks_list if r.strip()])
    as_on_str = request.POST.get('as_on_date', '').strip()
    as_on_date = None
    if as_on_str:
        try:
            as_on_date = datetime.strptime(as_on_str, '%Y-%m-%d').date()
        except ValueError:
            pass
            
    # Also save overall status
    record.status = status
    record.status_remarks = remarks
    record.as_on_date = as_on_date
    
    # Revised RPN fields (if user passes revised values)
    rev_sev = request.POST.get('action_severity')
    rev_occ = request.POST.get('action_occurrence')
    rev_det = request.POST.get('action_detection')
    rev_action = request.POST.get('action_taken')
    
    if rev_sev and rev_occ and rev_det:
        try:
            record.action_severity = int(rev_sev)
            record.action_occurrence = int(rev_occ)
            record.action_detection = int(rev_det)
            record.action_taken = rev_action or ""
        except ValueError:
            pass
            
    record.save()
    
    FMEAAuditLog.objects.create(
        record=record,
        user=request.user,
        action='MITIGATED',
        details=f"Updated Action Plan details. Current Status: {status}. Revised RPN: {record.action_rpn or 'N/A'}"
    )
    
    return JsonResponse({'status': 'success', 'message': 'Mitigation updated successfully.'})


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def download_excel(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    
    upload_id = request.GET.get('upload_id')
    if upload_id and upload_id != 'new_file':
        records = FMEARecord.objects.filter(department=active_dept, excel_upload_id=upload_id).order_by('sn', 'id')
    else:
        records = FMEARecord.objects.filter(department=active_dept, excel_upload=None).order_by('sn', 'id')
    
    # Load FMEA Format.xlsx template
    from django.conf import settings
    import os
    template_path = os.path.join(settings.BASE_DIR, 'EFMEA', 'fmea', 'FMEA Format.xlsx')
    try:
        wb = openpyxl.load_workbook(template_path, data_only=False)
    except Exception as e:
        return HttpResponse(f"Error loading Excel template: {str(e)}", status=500)
        
    # Get the FMEA sheet
    if 'FMEA' not in wb.sheetnames:
        return HttpResponse("FMEA sheet not found in the template.", status=500)
        
    sheet = wb['FMEA']
    
    # Read custom metadata fields from GET parameters
    key_contact = request.GET.get('key_contact', '').strip()
    core_team = request.GET.get('core_team', '').strip()
    date_val = request.GET.get('date', '').strip()
    ref_no = request.GET.get('ref_no', '').strip()
    objective = request.GET.get('objective', '').strip()

    # Format date from YYYY-MM-DD to DD.MM.YYYY
    if date_val:
        try:
            parsed_date = datetime.strptime(date_val, '%Y-%m-%d')
            formatted_date = parsed_date.strftime('%d.%m.%Y')
        except ValueError:
            formatted_date = date_val
    else:
        formatted_date = datetime.now().strftime('%d.%m.%Y')

    # Fill department header info into template
    if ref_no:
        sheet['Q1'] = f"Reference: {ref_no}"
    else:
        year = datetime.now().year
        import random
        rand_num = f"{random.randint(1, 99):02d}"
        sheet['Q1'] = f"Reference: JS-Raigarh/{active_dept.code}/{year}/FMEA/{rand_num}"

    sheet['Q2'] = f"Date : {formatted_date}"

    if key_contact:
        sheet['C4'] = key_contact
    else:
        sheet['C4'] = request.user.get_display_name() if hasattr(request.user, 'get_display_name') else request.user.username

    if core_team:
        sheet['H4'] = core_team
    else:
        sheet['H4'] = f"{active_dept.name} Operations Team"

    if objective:
        sheet['R4'] = objective
    else:
        sheet['R4'] = f"Minimize Production Loss Time due to {active_dept.name} issues"
    
    # Clean template rows starting at row 7
    num_records = len(records)
    placeholder_ranges = []
    legend_ranges_offsets = []

    # 1. Identify and unmerge all merged ranges starting at or below row 7.
    # We save the legend ranges (rows 12+) to re-merge them later at their new positions,
    # preventing openpyxl's insert_rows from corrupting merged cell references.
    for r_range in list(sheet.merged_cells.ranges):
        if r_range.min_row >= 7:
            if r_range.min_row < 12:
                placeholder_ranges.append(r_range)
            else:
                legend_ranges_offsets.append({
                    'min_col': r_range.min_col,
                    'max_col': r_range.max_col,
                    'row_offset_min': r_range.min_row - 12,
                    'row_offset_max': r_range.max_row - 12
                })
            sheet.unmerge_cells(
                start_row=r_range.min_row,
                start_column=r_range.min_col,
                end_row=r_range.max_row,
                end_column=r_range.max_col
            )

    # 2. Insert blank rows to shift the legend down if we have more than 5 records
    legend_start_row = 12
    if num_records > 5:
        shift_amount = num_records - 5
        sheet.insert_rows(12, amount=shift_amount)
        legend_start_row += shift_amount

    # 3. Re-merge the legend ranges at their new shifted positions
    for offset in legend_ranges_offsets:
        sheet.merge_cells(
            start_row=legend_start_row + offset['row_offset_min'],
            start_column=offset['min_col'],
            end_row=legend_start_row + offset['row_offset_max'],
            end_column=offset['max_col']
        )

    # Determine the range of rows we need to clear/write to
    max_touch_row = max(11, 6 + num_records)

    # Load openpyxl styles to apply beautiful, clean borders and vertical alignment
    from openpyxl.styles import Border, Side, Alignment, Font
    from copy import copy
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Copy column styles from row 7 of the template (before clearing) to preserve colors & styles
    col_styles = {}
    for c in range(1, 24):
        ref_cell = sheet.cell(row=7, column=c)
        col_styles[c] = {
            'font': copy(ref_cell.font),
            'fill': copy(ref_cell.fill),
            'alignment': copy(ref_cell.alignment),
            'border': copy(ref_cell.border) if ref_cell.border else thin_border
        }

    # Clear columns 1 to 23 for rows 7 to max_touch_row before writing to ensure clean data
    for r in range(7, max_touch_row + 1):
        for c in range(1, 24):
            sheet.cell(row=r, column=c).value = None

    # Style all rows from 7 to max_touch_row to ensure consistent gridlines and fonts
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for r in range(7, max_touch_row + 1):
        for c in range(1, 24):
            cell = sheet.cell(row=r, column=c)
            style = col_styles[c]
            if style['font']:
                cell.font = style['font']
            if style['fill']:
                cell.fill = style['fill']
            if c in [1, 10, 12, 15, 16, 18, 20, 21, 22, 23]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
            cell.border = style['border'] if style['border'] else thin_border

    # Group records by (main_equipment, main_equipment_function, functional_failure) for vertical merging
    groups = []
    current_group = []
    for rec in records:
        if not current_group:
            current_group = [rec]
        elif (current_group[0].main_equipment == rec.main_equipment and 
              current_group[0].main_equipment_function == rec.main_equipment_function and 
              current_group[0].functional_failure == rec.functional_failure):
            current_group.append(rec)
        else:
            groups.append(current_group)
            current_group = [rec]
    if current_group:
        groups.append(current_group)

    # Write records group by group
    row_idx = 7
    for group in groups:
        group_size = len(group)
        for offset, rec in enumerate(group):
            r = row_idx + offset
            
            # Write unique fields
            sheet.cell(row=r, column=5).value = rec.sub_equipment or "—"
            sheet.cell(row=r, column=6).value = rec.component or "—"
            sheet.cell(row=r, column=7).value = rec.component_function or "—"
            sheet.cell(row=r, column=8).value = rec.potential_failure_mode or "—"
            sheet.cell(row=r, column=9).value = rec.potential_effects or "—"
            sheet.cell(row=r, column=10).value = int(rec.severity or 1)
            sheet.cell(row=r, column=11).value = rec.potential_causes or "—"
            sheet.cell(row=r, column=12).value = int(rec.occurrence or 1)
            sheet.cell(row=r, column=13).value = rec.current_controls or "—"
            sheet.cell(row=r, column=14).value = rec.recommended_actions or "—"
            sheet.cell(row=r, column=15).value = int(rec.detection or 1)
            sheet.cell(row=r, column=16).value = f"=J{r}*L{r}*O{r}"
            sheet.cell(row=r, column=17).value = rec.contingency_plan or "—"
            sheet.cell(row=r, column=18).value = rec.status or "Not Started"
            sheet.cell(row=r, column=19).value = rec.action_taken or "—"
            
            if rec.action_severity is not None:
                sheet.cell(row=r, column=20).value = int(rec.action_severity)
            else:
                sheet.cell(row=r, column=20).value = None
            if rec.action_occurrence is not None:
                sheet.cell(row=r, column=21).value = int(rec.action_occurrence)
            else:
                sheet.cell(row=r, column=21).value = None
            if rec.action_detection is not None:
                sheet.cell(row=r, column=22).value = int(rec.action_detection)
            else:
                sheet.cell(row=r, column=22).value = None
            sheet.cell(row=r, column=23).value = f"=T{r}*U{r}*V{r}"
            
        # Write merged fields to the first row of the group, and merge vertically
        lead_rec = group[0]
        sheet.cell(row=row_idx, column=1).value = lead_rec.sn or str(row_idx - 6)
        sheet.cell(row=row_idx, column=2).value = lead_rec.main_equipment
        sheet.cell(row=row_idx, column=3).value = lead_rec.main_equipment_function
        sheet.cell(row=row_idx, column=4).value = lead_rec.functional_failure
        
        if group_size > 1:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + group_size - 1, end_column=1)
            sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx + group_size - 1, end_column=2)
            sheet.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx + group_size - 1, end_column=3)
            sheet.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx + group_size - 1, end_column=4)
        row_idx += group_size
        
    filename = None
    if upload_id and upload_id != 'new_file':
        try:
            upload_obj = FMEAExcelUpload.objects.get(id=int(upload_id), department=active_dept)
            filename = upload_obj.filename
        except (ValueError, FMEAExcelUpload.DoesNotExist):
            pass
            
    if not filename:
        filename = f"FMEA_{active_dept.code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
        
    # Write to Action Plan sheet if present in template
    action_sheet = None
    for name in wb.sheetnames:
        if name.lower().strip() == 'action plan':
            action_sheet = wb[name]
            break
            
    if action_sheet:
        # Clear rows 4 to max row to clean up template placeholders
        for r in range(4, action_sheet.max_row + 1):
            for c in range(1, 9):
                action_sheet.cell(row=r, column=c).value = None
                
        # Copy style from row 4 if exists, or use default thin border
        action_col_styles = {}
        for c in range(2, 8): # columns B (2) to G (7)
            ref_cell = action_sheet.cell(row=4, column=c)
            action_col_styles[c] = {
                'font': copy(ref_cell.font) if ref_cell.font else None,
                'fill': copy(ref_cell.fill) if ref_cell.fill else None,
                'alignment': copy(ref_cell.alignment) if ref_cell.alignment else Alignment(vertical='center', wrap_text=True),
                'border': copy(ref_cell.border) if ref_cell.border else thin_border
            }
            
        action_row_idx = 4
        for rec in records:
            # Decode action_plan_data
            try:
                items = json.loads(rec.action_plan_data or '[]')
            except:
                items = []
                
            # Fallback migration to export legacy quarterly fields if new Action Plan is empty
            if not items:
                legacy_items = []
                for q, field_val, target_val, status_val in [
                    ('q1', rec.mitigation_q1, rec.mitigation_q1_target, rec.mitigation_q1_status),
                    ('q2', rec.mitigation_q2, rec.mitigation_q2_target, rec.mitigation_q2_status),
                    ('q3', rec.mitigation_q3, rec.mitigation_q3_target, rec.mitigation_q3_status),
                    ('q4', rec.mitigation_q4, rec.mitigation_q4_target, rec.mitigation_q4_status),
                ]:
                    try:
                        acts = json.loads(field_val or '[]')
                        for a in acts:
                            if a.strip():
                                legacy_items.append({
                                    'action': a.strip(),
                                    'responsibility': rec.risk_owner or '',
                                    'plan_term': 'Short' if q in ['q1', 'q2'] else 'Medium',
                                    'target_date': target_val.strftime('%Y-%m-%d') if target_val else '',
                                    'status': status_val or 'Not Started'
                                })
                    except:
                        pass
                items = legacy_items
                
            for item in items:
                # Column B (2): SN
                action_sheet.cell(row=action_row_idx, column=2).value = rec.sn or str(rec.id)
                # Column C (3): Actions Required
                action_sheet.cell(row=action_row_idx, column=3).value = item.get('action', '')
                # Column D (4): Responsibility
                action_sheet.cell(row=action_row_idx, column=4).value = item.get('responsibility', rec.risk_owner or '')
                # Column E (5): Plan Term
                action_sheet.cell(row=action_row_idx, column=5).value = item.get('plan_term', 'Short')
                
                # Column F (6): Target Date
                t_date = item.get('target_date', '')
                if t_date:
                    try:
                        parsed_t_date = datetime.strptime(t_date, '%Y-%m-%d')
                        formatted_t_date = parsed_t_date.strftime('%d.%m.%Y')
                        action_sheet.cell(row=action_row_idx, column=6).value = formatted_t_date
                    except ValueError:
                        action_sheet.cell(row=action_row_idx, column=6).value = t_date
                else:
                    action_sheet.cell(row=action_row_idx, column=6).value = ""
                    
                # Column G (7): Status as on Date
                action_sheet.cell(row=action_row_idx, column=7).value = item.get('status', 'Pending')
                
                # Apply styles
                for c in range(2, 8):
                    cell = action_sheet.cell(row=action_row_idx, column=c)
                    style = action_col_styles.get(c, {})
                    if style.get('font'):
                        cell.font = style['font']
                    if style.get('fill'):
                        cell.fill = style['fill']
                    if style.get('alignment'):
                        cell.alignment = style['alignment']
                    cell.border = style.get('border') or thin_border
                    
                action_row_idx += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.force_full_calculation = True
    wb.save(response)
    return response


def parse_sheet_date(date_val):
    if not date_val:
        return None
    date_str = str(date_val).strip()
    if ':' in date_str:
        date_str = date_str.split(':', 1)[1].strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def upload_excel(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    if request.method != 'POST' or 'excel_file' not in request.FILES:
        messages.error(request, "Please select a valid Excel file to upload.")
        return redirect('fmea:report', dept_id=dept_id)
        
    excel_file = request.FILES['excel_file']
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        messages.error(request, f"Failed to load uploaded file: {str(e)}")
        return redirect('fmea:report', dept_id=dept_id)
        
    sheet_name = None
    for name in wb.sheetnames:
        if name.upper() in ['FMEA', 'EFMEA']:
            sheet_name = name
            break
            
    if not sheet_name:
        for name in wb.sheetnames:
            if 'FMEA' in name.upper() or 'EFMEA' in name.upper():
                sheet_name = name
                break
                
    if not sheet_name:
        for name in wb.sheetnames:
            name_upper = name.upper().strip()
            if 'PROCESS FLOW' not in name_upper and 'ACTION PLAN' not in name_upper and 'LEGEND' not in name_upper:
                sheet_name = name
                break
                
    if not sheet_name and wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        
    if not sheet_name:
        messages.error(request, "Could not find a valid FMEA sheet in the uploaded file.")
        return redirect('fmea:report', dept_id=dept_id)
        
    sheet = wb[sheet_name]
    
    # Validation: Check if the uploaded Excel matches the current department
    import re
    
    def split_to_words(text):
        return [w.lower() for w in re.split(r'[^a-zA-Z0-9]+', str(text)) if w]
        
    filename_words = split_to_words(excel_file.name)
    
    # Gather header cells (rows 1-4, cols 1-5)
    excel_header_text = ""
    for r in range(1, 5):
        for c in range(1, 6):
            val = sheet.cell(row=r, column=c).value
            if val:
                excel_header_text += " " + str(val)
    excel_words = split_to_words(excel_header_text)
    
    def is_dept_referenced(dept, words_list, text_raw):
        # Exact match on code
        if dept.code and dept.code.lower() in words_list:
            return True
            
        # Match all words of department name
        dept_name_words = split_to_words(dept.name)
        if dept_name_words and all(w in words_list for w in dept_name_words):
            return True
            
        # Match concatenated normalized name (e.g. "railmill" in "railmillraigar...")
        dept_norm = "".join(dept_name_words)
        text_norm = "".join(words_list)
        if dept_norm and dept_norm in text_norm:
            return True
            
        return False
        
    # Check if active department is referenced
    active_referenced = is_dept_referenced(active_dept, filename_words, excel_file.name) or is_dept_referenced(active_dept, excel_words, excel_header_text)
    
    # Check if any OTHER department is referenced
    other_departments = Department.objects.exclude(id=active_dept.id)
    mismatched_dept = None
    for dept in other_departments:
        if is_dept_referenced(dept, filename_words, excel_file.name) or is_dept_referenced(dept, excel_words, excel_header_text):
            mismatched_dept = dept
            break
            
    # If the file references another department and NOT the active department, block it
    if mismatched_dept and not active_referenced:
        messages.error(
            request,
            f"Validation Error: The uploaded file/header references '{mismatched_dept.name}', "
            f"but you are uploading to '{active_dept.name}'. Please switch to the correct department hub first."
        )
        return redirect('fmea:report', dept_id=dept_id)
    
    # Extract metadata headers dynamically from the first 4 rows
    ref_no = ""
    sheet_date = None
    key_contact = ""
    core_team = ""
    objective = ""
    
    # Helper to find a value after a label or in the next non-empty cell in the same row
    def find_field_value(sheet_obj, keyword, skip_keywords=None):
        if skip_keywords is None:
            skip_keywords = []
        for r in range(1, 5):
            for c in range(1, min(25, sheet_obj.max_column + 1)):
                val = sheet_obj.cell(row=r, column=c).value
                if val is None:
                    continue
                val_str = str(val).strip()
                if keyword.lower() in val_str.lower():
                    # If the cell itself contains a colon, the value might be in this cell after the colon
                    if ":" in val_str:
                        part = val_str.split(":", 1)[1].strip()
                        if part:
                            part_lower = part.lower()
                            if not any(sk.lower() in part_lower for sk in skip_keywords):
                                return part
                    
                    # Otherwise look at subsequent cells in the same row
                    for offset in range(1, 5):
                        if c + offset > sheet_obj.max_column:
                            break
                        candidate = sheet_obj.cell(row=r, column=c+offset).value
                        if candidate is not None:
                            cand_str = str(candidate).strip()
                            if cand_str:
                                cand_lower = cand_str.lower()
                                # Make sure the candidate is not another keyword/header
                                if not any(sk.lower() in cand_lower for sk in [keyword] + skip_keywords):
                                    return cand_str
        return ""

    # Search for Reference
    ref_val = find_field_value(sheet, "Reference", ["Date", "Pages", "Key Contact", "Core Team", "Objective"])
    if ref_val:
        ref_no = ref_val
    else:
        # Fallback to hardcoded template cell Q1 (column 17)
        ref_raw = sheet.cell(row=1, column=17).value
        if ref_raw:
            ref_no = str(ref_raw).replace("Reference:", "").replace("Reference :", "").strip()

    # Search for Date
    date_val = find_field_value(sheet, "Date", ["Reference", "Pages", "Key Contact", "Core Team", "Objective"])
    if date_val:
        sheet_date = parse_sheet_date(date_val)
    else:
        # Fallback to hardcoded template cell Q2 (column 17)
        date_raw = sheet.cell(row=2, column=17).value
        sheet_date = parse_sheet_date(date_raw)

    # Search for Key Contact
    key_contact_val = find_field_value(sheet, "Key Contact", ["Core Team", "Objective", "Reference", "Date", "Pages"])
    if key_contact_val:
        key_contact = key_contact_val
    else:
        key_contact = str(sheet.cell(row=4, column=3).value or "").strip()

    # Search for Core Team
    core_team_val = find_field_value(sheet, "Core Team", ["Key Contact", "Objective", "Reference", "Date", "Pages"])
    if core_team_val:
        core_team = core_team_val
    else:
        core_team = str(sheet.cell(row=4, column=8).value or "").strip()

    # Search for Objective
    objective_val = find_field_value(sheet, "Objective", ["Key Contact", "Core Team", "Reference", "Date", "Pages"])
    if objective_val:
        objective = objective_val
    else:
        objective = str(sheet.cell(row=4, column=18).value or "").strip()
        
    # Create the Excel upload record
    upload = FMEAExcelUpload.objects.create(
        department=active_dept,
        filename=excel_file.name,
        sheet_date=sheet_date,
        uploaded_by=request.user,
        key_contact=key_contact,
        core_team=core_team,
        objective=objective,
        ref_no=ref_no
    )
    
    # Read headers from rows 5 and 6 dynamically to handle merged and shifted columns
    header_row = 5
    mapping = {}
    parent_header = ""
    for c in range(1, sheet.max_column + 1):
        r5_val = sheet.cell(row=5, column=c).value
        r6_val = sheet.cell(row=6, column=c).value
        
        r5_str = str(r5_val).strip().replace('\n', ' ') if r5_val is not None else ""
        r6_str = str(r6_val).strip().replace('\n', ' ') if r6_val is not None else ""
        
        if r5_str:
            parent_header = r5_str
            
        if r6_str:
            header = f"{parent_header} {r6_str}".lower()
        else:
            header = r5_str.lower()
            
        if not header:
            continue
            
        col_num = c
        if header == 'sn' or 's.n' in header or 'serial' in header:
            mapping['sn'] = col_num
        elif 'main equipment' in header or ('equipment' in header and 'function' not in header and 'sub' not in header):
            mapping['main_equipment'] = col_num
        elif 'main equipment function' in header or 'equipment function' in header:
            mapping['main_equipment_function'] = col_num
        elif 'functional failure' in header:
            mapping['functional_failure'] = col_num
        elif 'sub-equipment' in header or 'sub equipment' in header:
            mapping['sub_equipment'] = col_num
        elif 'component function' in header:
            mapping['component_function'] = col_num
        elif 'component' in header:
            mapping['component'] = col_num
        elif 'potential failure mode' in header or 'failure mode' in header or 'risk description' in header:
            mapping['potential_failure_mode'] = col_num
        elif 'potential effect' in header or 'consequences' in header:
            mapping['potential_effects'] = col_num
        elif header == 'sev' or 'severity' in header:
            mapping['severity'] = col_num
        elif 'potential cause' in header or 'cause' in header:
            mapping['potential_causes'] = col_num
        elif 'occur' in header or 'occurrence' in header:
            mapping['occurrence'] = col_num
        elif 'current control' in header or 'prevention' in header:
            mapping['current_controls'] = col_num
        elif 'recommended' in header or 'action plan' in header or 'improvement' in header:
            mapping['recommended_actions'] = col_num
        elif header == 'det' or 'detect' in header or 'detection' in header:
            mapping['detection'] = col_num
        elif 'action taken' in header or 'corrective action' in header:
            mapping['action_taken'] = col_num
        elif 'action sev' in header or 'revised sev' in header or ('action results' in header and 'sev' in header):
            mapping['action_severity'] = col_num
        elif 'action occ' in header or 'revised occ' in header or ('action results' in header and 'occ' in header):
            mapping['action_occurrence'] = col_num
        elif 'action det' in header or 'revised det' in header or ('action results' in header and 'det' in header):
            mapping['action_detection'] = col_num
        elif 'status' in header:
            mapping['status'] = col_num
        elif 'contingency' in header:
            mapping['contingency_plan'] = col_num
            
    if 'main_equipment' not in mapping and 'potential_failure_mode' not in mapping:
        messages.error(request, f"Sheet '{sheet_name}' does not contain recognized column headers on row {header_row} (e.g. 'Equipment', 'Failure Mode').")
        return redirect('fmea:report', dept_id=dept_id)

    # Helper to normalize SN for linking FMEA records with their action plans
    def normalize_sn(val):
        if val is None:
            return ""
        val_str = str(val).strip()
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        if val_str.isdigit():
            return str(int(val_str))
        return val_str.lower()
        
    action_sheet = None
    for name in wb.sheetnames:
        if name.lower().strip() == 'action plan':
            action_sheet = wb[name]
            break
            
    action_plans_by_sn = {}
    if action_sheet:
        for r in range(4, action_sheet.max_row + 1):
            sn_val = action_sheet.cell(row=r, column=2).value
            action_val = action_sheet.cell(row=r, column=3).value
            resp_val = action_sheet.cell(row=r, column=4).value
            term_val = action_sheet.cell(row=r, column=5).value
            target_val = action_sheet.cell(row=r, column=6).value
            status_val = action_sheet.cell(row=r, column=7).value
            
            if sn_val is None and action_val is None:
                continue
                
            normalized_sn = normalize_sn(sn_val)
            if not normalized_sn:
                continue
                
            target_date_str = ""
            if target_val:
                parsed_date = parse_sheet_date(target_val)
                if parsed_date:
                    target_date_str = parsed_date.strftime('%Y-%m-%d')
                else:
                    target_date_str = str(target_val).strip()
                    
            item = {
                'action': str(action_val or '').strip(),
                'responsibility': str(resp_val or '').strip(),
                'plan_term': str(term_val or 'Short').strip(),
                'target_date': target_date_str,
                'status': str(status_val or 'Pending').strip()
            }
            
            if normalized_sn not in action_plans_by_sn:
                action_plans_by_sn[normalized_sn] = []
            action_plans_by_sn[normalized_sn].append(item)

    start_row = 7
    rows_created = 0
    
    last_sn = ""
    last_main_eq = ""
    last_main_eq_function = ""
    last_functional_failure = ""
    
    first_record = None
    for r in range(start_row, sheet.max_row + 1):
        def get_val(field):
            col = mapping.get(field)
            if col is None:
                return ""
            val = sheet.cell(row=r, column=col).value
            return str(val).strip() if val is not None else ""
            
        sn = get_val('sn')
        main_eq = get_val('main_equipment')
        main_eq_function = get_val('main_equipment_function')
        func_fail = get_val('functional_failure')
        
        sn_upper = sn.upper()
        main_eq_upper = main_eq.upper()
        if any(keyword in sn_upper or keyword in main_eq_upper for keyword in ['LEGEND', 'COLOUR CODE', 'CRITICAL SPARES']):
            break

        comp = get_val('component')
        fail_mode = get_val('potential_failure_mode')
        
        if not sn and not main_eq and not comp and not fail_mode:
            continue
            
        if not main_eq and last_main_eq:
            main_eq = last_main_eq
            sn = last_sn
            if not main_eq_function:
                main_eq_function = last_main_eq_function
            if not func_fail:
                func_fail = last_functional_failure
        else:
            if sn:
                last_sn = sn
            if main_eq:
                last_main_eq = main_eq
            if main_eq_function:
                last_main_eq_function = main_eq_function
            if func_fail:
                last_functional_failure = func_fail
            
        def safe_int(field, default=1):
            val_str = main_eq if field == 'main_equipment' else (sn if field == 'sn' else (main_eq_function if field == 'main_equipment_function' else (func_fail if field == 'functional_failure' else get_val(field))))
            if not val_str:
                return default
            try:
                return int(float(val_str))
            except:
                return default
                
        def safe_int_nullable(field):
            val_str = main_eq if field == 'main_equipment' else (sn if field == 'sn' else (main_eq_function if field == 'main_equipment_function' else (func_fail if field == 'functional_failure' else get_val(field))))
            if not val_str:
                return None
            try:
                return int(float(val_str))
            except:
                return None

        # Create new record for this upload
        record = FMEARecord(department=active_dept, excel_upload=upload, sn=sn)
        rows_created += 1
        
        # Link action plan items by matching SN
        normalized_record_sn = normalize_sn(sn)
        if normalized_record_sn in action_plans_by_sn:
            record.action_plan_data = json.dumps(action_plans_by_sn[normalized_record_sn])
            
        record.main_equipment = main_eq
        record.main_equipment_function = main_eq_function
        record.functional_failure = func_fail
        record.sub_equipment = get_val('sub_equipment')
        record.component = get_val('component')
        record.component_function = get_val('component_function')
        record.potential_failure_mode = get_val('potential_failure_mode')
        record.potential_effects = get_val('potential_effects')
        record.severity = safe_int('severity', 1)
        record.potential_causes = get_val('potential_causes')
        record.occurrence = safe_int('occurrence', 1)
        record.current_controls = get_val('current_controls')
        record.recommended_actions = get_val('recommended_actions')
        record.detection = safe_int('detection', 1)
        record.contingency_plan = get_val('contingency_plan')
        record.status = get_val('status') or "Not Started"
        record.action_taken = get_val('action_taken')
        record.action_severity = safe_int_nullable('action_severity')
        record.action_occurrence = safe_int_nullable('action_occurrence')
        record.action_detection = safe_int_nullable('action_detection')
        
        record.save()
        if not first_record:
            first_record = record
            
    # Parse critical spares from Excel sheet
    spares_row = None
    spares_col = None
    for r in range(10, min(150, sheet.max_row + 1)):
        for c in range(1, min(30, sheet.max_column + 1)):
            val = str(sheet.cell(row=r, column=c).value or "").strip().lower()
            if "critical spares" in val:
                spares_row = r
                spares_col = c
                break
        if spares_row:
            break
            
    if spares_row and spares_col:
        for r in range(spares_row + 2, spares_row + 15):
            spare_desc = str(sheet.cell(row=r, column=spares_col).value or "").strip()
            qty = str(sheet.cell(row=r, column=spares_col + 1).value or "").strip()
            remarks_1 = str(sheet.cell(row=r, column=spares_col + 2).value or "").strip()
            lead_time = str(sheet.cell(row=r, column=spares_col + 3).value or "").strip()
            remarks_2 = str(sheet.cell(row=r, column=spares_col + 4).value or "").strip()
            
            if not spare_desc and not qty and not remarks_1 and not lead_time and not remarks_2:
                continue
                
            FMEACriticalSpare.objects.create(
                excel_upload=upload,
                department=active_dept,
                spare_description=spare_desc,
                qty=qty,
                remarks_1=remarks_1,
                lead_time=lead_time,
                remarks_2=remarks_2
            )
            
    if first_record:
        FMEAAuditLog.objects.create(
            record=first_record,
            user=request.user,
            action='EXCEL_UPLOADED',
            details=f"Uploaded Excel '{excel_file.name}' with sheet date {sheet_date or 'N/A'}. Created {rows_created} records."
        )
        
    messages.success(request, f"Excel file parsed successfully. Created new upload with date {sheet_date or 'N/A'}. {rows_created} records imported.")
    return redirect(f'/fmea/department/{dept_id}/report/?upload_id={upload.id}')


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def delete_upload(request, dept_id, upload_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    upload = get_object_or_404(FMEAExcelUpload, id=upload_id, department=active_dept)
    filename = upload.filename
    # Delete the upload. The associated FMEARecords are automatically deleted via CASCADE.
    upload.delete()
    messages.success(request, f"Successfully deleted Excel file '{filename}' and all its associated records.")
    return redirect('fmea:history', dept_id=dept_id)


@login_required
@dept_visibility_required
@module_access_required('FMEA')
def clear_manual_records(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    records = FMEARecord.objects.filter(department=active_dept, excel_upload=None)
    count = records.count()
    records.delete()
    messages.success(request, f"Successfully cleared all {count} manual FMEA records.")
    return redirect('fmea:history', dept_id=dept_id)


@login_required
@dept_visibility_required
@module_access_required('FMEA', require_edit=False)
def save_report_rows(request, dept_id):
    active_dept = get_object_or_404(Department, id=dept_id)
    if request.method != 'POST':
        return redirect('fmea:report', dept_id=dept_id)
        
    upload_id = request.POST.get('upload_id')
    if upload_id and upload_id != 'new_file':
        try:
            upload_obj = FMEAExcelUpload.objects.get(id=int(upload_id), department=active_dept)
            upload_obj.key_contact = request.POST.get('key_contact', '').strip()
            upload_obj.core_team = request.POST.get('core_team', '').strip()
            upload_obj.objective = request.POST.get('objective', '').strip()
            upload_obj.ref_no = request.POST.get('ref_no', '').strip()
            
            filename = request.POST.get('filename', '').strip()
            if filename:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                upload_obj.filename = filename
            
            date_str = request.POST.get('date', '').strip()
            if date_str:
                try:
                    upload_obj.sheet_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
            upload_obj.save()
        except (ValueError, FMEAExcelUpload.DoesNotExist):
            pass
            
        records = FMEARecord.objects.filter(department=active_dept, excel_upload_id=upload_id).order_by('sn', 'id')
    else:
        records = FMEARecord.objects.filter(department=active_dept, excel_upload=None).order_by('sn', 'id')
    
    last_lead_sn = ""
    last_lead_main_eq = ""
    last_lead_main_eq_function = ""
    last_lead_functional_failure = ""
    
    def get_int_or_none(val_str):
        if not val_str:
            return None
        try:
            return int(float(val_str))
        except:
            return None
            
    for r in records:
        rid = r.id
        if f'main_equipment_{rid}' in request.POST:
            r.sn = request.POST.get(f'sn_{rid}', '').strip()
            r.main_equipment = request.POST.get(f'main_equipment_{rid}', '').strip()
            r.main_equipment_function = request.POST.get(f'main_equipment_function_{rid}', '').strip()
            r.functional_failure = request.POST.get(f'functional_failure_{rid}', '').strip()
            
            last_lead_sn = r.sn
            last_lead_main_eq = r.main_equipment
            last_lead_main_eq_function = r.main_equipment_function
            last_lead_functional_failure = r.functional_failure
        else:
            if last_lead_main_eq:
                r.sn = last_lead_sn
                r.main_equipment = last_lead_main_eq
                r.main_equipment_function = last_lead_main_eq_function
                r.functional_failure = last_lead_functional_failure
                
        r.sub_equipment = request.POST.get(f'sub_equipment_{rid}', '').strip()
        r.component = request.POST.get(f'component_{rid}', '').strip()
        r.component_function = request.POST.get(f'component_function_{rid}', '').strip()
        r.potential_failure_mode = request.POST.get(f'potential_failure_mode_{rid}', '').strip()
        r.potential_effects = request.POST.get(f'potential_effects_{rid}', '').strip()
        
        try:
            r.severity = int(request.POST.get(f'severity_{rid}', 1) or 1)
        except ValueError:
            r.severity = 1
            
        r.potential_causes = request.POST.get(f'potential_causes_{rid}', '').strip()
        
        try:
            r.occurrence = int(request.POST.get(f'occurrence_{rid}', 1) or 1)
        except ValueError:
            r.occurrence = 1
            
        r.current_controls = request.POST.get(f'current_controls_{rid}', '').strip()
        r.recommended_actions = request.POST.get(f'recommended_actions_{rid}', '').strip()
        
        try:
            r.detection = int(request.POST.get(f'detection_{rid}', 1) or 1)
        except ValueError:
            r.detection = 1
            
        r.contingency_plan = request.POST.get(f'contingency_plan_{rid}', '').strip()
        r.status = request.POST.get(f'status_{rid}', 'Not Started').strip()
        r.action_taken = request.POST.get(f'action_taken_{rid}', '').strip()
        r.action_severity = get_int_or_none(request.POST.get(f'action_severity_{rid}', '').strip())
        r.action_occurrence = get_int_or_none(request.POST.get(f'action_occurrence_{rid}', '').strip())
        r.action_detection = get_int_or_none(request.POST.get(f'action_detection_{rid}', '').strip())
        
        r.save()
        
        FMEAAuditLog.objects.create(
            record=r,
            user=request.user,
            action='UPDATED',
            details=f"Inline edited bulk save via report. RPN: {r.rpn}."
        )
        
    # Save critical spares
    if upload_id and upload_id != 'new_file':
        FMEACriticalSpare.objects.filter(excel_upload_id=int(upload_id), department=active_dept).delete()
        upload_obj_id = int(upload_id)
    else:
        FMEACriticalSpare.objects.filter(excel_upload=None, department=active_dept).delete()
        upload_obj_id = None
        
    spare_descs = request.POST.getlist('spare_description[]')
    qtys = request.POST.getlist('spare_qty[]')
    remarks_1s = request.POST.getlist('spare_remarks_1[]')
    lead_times = request.POST.getlist('spare_lead_time[]')
    remarks_2s = request.POST.getlist('spare_remarks_2[]')
    
    for idx in range(len(spare_descs)):
        desc = spare_descs[idx].strip()
        qty = qtys[idx].strip() if idx < len(qtys) else ""
        rem1 = remarks_1s[idx].strip() if idx < len(remarks_1s) else ""
        lt = lead_times[idx].strip() if idx < len(lead_times) else ""
        rem2 = remarks_2s[idx].strip() if idx < len(remarks_2s) else ""
        
        if desc or qty or rem1 or lt or rem2:
            FMEACriticalSpare.objects.create(
                excel_upload_id=upload_obj_id,
                department=active_dept,
                spare_description=desc,
                qty=qty,
                remarks_1=rem1,
                lead_time=lt,
                remarks_2=rem2
            )
        
    messages.success(request, "FMEA report updated successfully.")
    from django.urls import reverse
    return redirect(reverse('fmea:report', args=[dept_id]) + f'?upload_id={upload_id or "new_file"}')

