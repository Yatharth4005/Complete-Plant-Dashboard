import os
import openpyxl
from django.core.management.base import BaseCommand
from tpm.models import Department
from cmc.models import Equipment, EquipmentBearingPoint

class Command(BaseCommand):
    help = 'Seed CMC Equipment master from Excel PM schedule'

    def handle(self, *args, **kwargs):
        excel_path = 'CMC Requirements.xlsx'
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f"Could not find '{excel_path}' in the current directory."))
            return

        self.stdout.write("Loading workbook...")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Check sheet names to find the schedule sheet
        self.stdout.write(f"Sheets found: {wb.sheetnames}")
        
        # Try to find a sheet with 'schedule' or 'cmc' in it, otherwise default to first sheet
        sheet_name = None
        for name in wb.sheetnames:
            if 'schedule' in name.lower() or 'cmc' in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
            
        self.stdout.write(f"Reading from sheet: '{sheet_name}'")
        ws = wb[sheet_name]

        # Skip header rows (first 2 rows are headers)
        count = 0
        rows = list(ws.iter_rows(values_only=True))
        
        # Find the row that contains 'S. No.' or similar headers
        start_row = 1
        for idx, r in enumerate(rows[:5]):
            if r and any(str(x).strip().lower() in ['s. no.', 's.no.'] for x in r if x is not None):
                start_row = idx + 1
                break
                
        self.stdout.write(f"Header starts at row {start_row}")

        for r_idx, row in enumerate(rows[start_row:]):
            if not row or len(row) < 4:
                continue
                
            scheduled_days = row[0]
            dept_name = row[1]
            equip_name = row[2]
            equip_class = row[3]
            
            if not equip_name or not dept_name:
                continue

            sap_mech = str(row[4] or '').strip() if len(row) > 4 else ''
            sap_elec = str(row[5] or '').strip() if len(row) > 5 else ''
            asset_cost = str(row[6] or '').strip() if len(row) > 6 else ''
            prod_loss = str(row[7] or '').strip() if len(row) > 7 else ''
            
            rating = None
            if len(row) > 8 and row[8] is not None:
                try:
                    rating = float(row[8])
                except (ValueError, TypeError):
                    pass
                    
            frequency = 'MONTHLY'
            if len(row) > 10 and row[10]:
                frequency = self._map_frequency(str(row[10]))

            # Map department name to Department object
            dept = self._get_or_create_dept(dept_name)
            if not dept:
                # Log mapping issue but don't crash
                continue

            equip, created = Equipment.objects.get_or_create(
                department=dept,
                name=str(equip_name).strip(),
                defaults={
                    'equipment_class': str(equip_class or 'B').strip(),
                    'sap_code_mech':   sap_mech,
                    'sap_code_elec':   sap_elec,
                    'asset_cost':      asset_cost,
                    'production_loss': prod_loss,
                    'rating_kw':       rating,
                    'frequency':       frequency,
                    'scheduled_days':  str(scheduled_days or '').strip(),
                }
            )
            
            # Default generate standard 4 bearing points for each seeded equipment
            if created:
                EquipmentBearingPoint.objects.create(equipment=equip, label='DE', sort_order=1)
                EquipmentBearingPoint.objects.create(equipment=equip, label='NDE', sort_order=2)
                EquipmentBearingPoint.objects.create(equipment=equip, label='Pump DE', sort_order=3)
                EquipmentBearingPoint.objects.create(equipment=equip, label='Pump NDE', sort_order=4)
                count += 1

        self.stdout.write(self.style.SUCCESS(f'Seeded {count} equipment records with default bearing points.'))

    def _get_or_create_dept(self, name):
        # Map CMC dept names to portal Department codes
        CMC_TO_PORTAL = {
            'PP-3': 'PP3', 'PP-2': 'PP2', 'PP-1': 'PP1', 'PP-2 Ph-3': 'PPP3',
            'BF-2': 'BF2', 'BF-1': 'BF1', 'DRI-2': 'DRI2', 'DRI-1': 'DRI1',
            'SMS-2': 'SMS2', 'SMS-3': 'SMS3', 'Plate Mill': 'PM', 'SPM': 'SPM',
            'Rail Mill': 'RM', 'SAF': 'SAF1', 'LDP': 'LDP', 'Sinter Plant': 'SINT',
            'Coke Oven': 'CO', 'Cement Plant': 'CP', 'Oxygen Plant': 'OP',
            'RMH-3': 'RMHS3', 'RMH-1': 'RMHS1', 'PGP-2': 'PGP2', 'PGP-3': 'PGP3',
            'CTL-3': 'PM',  # CTL is under Plate Mill
            'Coal Washery': 'CO',
        }
        code = CMC_TO_PORTAL.get(str(name).strip())
        if code:
            try:
                return Department.objects.get(code=code)
            except Department.DoesNotExist:
                pass
        return None

    def _map_frequency(self, freq_str):
        freq_lower = freq_str.lower()
        if 'weekly' in freq_lower:
            return 'WEEKLY'
        if 'fortnightly' in freq_lower or 'fortnight' in freq_lower:
            return 'FORTNIGHTLY'
        if 'quarterly' in freq_lower or 'quaterly' in freq_lower:
            return 'QUARTERLY'
        return 'MONTHLY'
