import io
from datetime import date
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from cmc.models import Equipment, PMScheduleEntry, VibrationLog, OilTestLog, WDALog, SAPNotification

def generate_pdf_report(department, month, year):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#003478'),
        spaceAfter=15
    )
    section_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#7C3AED'),
        spaceBefore=12,
        spaceAfter=8
    )
    body_style = styles['BodyText']

    # Title
    story.append(Paragraph(f"Jindal Steels Operations Portal Condition Monitoring Cell (CMC) Report", title_style))
    story.append(Paragraph(f"Department: {department.name} | Period: {calendar_month_name(month)} {year}", body_style))
    story.append(Spacer(1, 15))

    # 1. PM Schedule Summary
    story.append(Paragraph("1. Predictive Maintenance Schedule Compliance", section_style))
    schedule_entries = PMScheduleEntry.objects.filter(
        equipment__department=department,
        scheduled_date__month=month,
        scheduled_date__year=year
    )
    total = schedule_entries.count()
    done = schedule_entries.filter(status=PMScheduleEntry.VisitStatus.DONE).count()
    compliance = (done / total * 100) if total > 0 else 100.0
    story.append(Paragraph(f"Total Scheduled Inspections: {total} | Completed: {done} | Compliance Rate: {compliance:.1f}%", body_style))
    story.append(Spacer(1, 10))

    # 2. Vibration logs summary
    story.append(Paragraph("2. Recent Vibration Inspections", section_style))
    vibs = VibrationLog.objects.filter(equipment__department=department, date__month=month, date__year=year)[:10]
    
    vib_data = [["Date", "Equipment", "Instrument", "Status", "Remarks"]]
    for v in vibs:
        vib_data.append([
            v.date.strftime('%Y-%m-%d'),
            v.equipment.name[:30],
            v.instrument,
            v.status,
            v.remarks[:30]
        ])
    
    if len(vib_data) > 1:
        t = Table(vib_data, colWidths=[70, 150, 80, 80, 160])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F5F3FF')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#7C3AED')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDD6FE')),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No vibration inspections logged in this period.", body_style))
    story.append(Spacer(1, 15))

    # 3. Oil Tests Summary
    story.append(Paragraph("3. Recent Oil Test Analyses", section_style))
    oils = OilTestLog.objects.filter(equipment__department=department, date__month=month, date__year=year)[:10]
    
    oil_data = [["Date", "Equipment", "Viscosity (cSt)", "Moisture", "Status"]]
    for o in oils:
        oil_data.append([
            o.date.strftime('%Y-%m-%d'),
            o.equipment.name[:35],
            str(o.viscosity or ''),
            o.moisture,
            o.status
        ])
        
    if len(oil_data) > 1:
        t = Table(oil_data, colWidths=[70, 200, 100, 100, 70])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F5F3FF')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#7C3AED')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDD6FE')),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No oil tests logged in this period.", body_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(department, month, year):
    wb = openpyxl.Workbook()
    
    # Setup Sheet 1: PM Schedule
    ws1 = wb.active
    ws1.title = "PM Schedule"
    
    # Headers
    ws1.merge_cells("A1:G1")
    ws1["A1"] = f"Jindal Steels Operations Portal CMC Predictive Maintenance Schedule - {department.name} - {calendar_month_name(month)} {year}"
    ws1["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="003478")
    
    headers = ["S.No.", "Equipment", "Class", "Frequency", "Scheduled Days", "Status", "Done Date"]
    ws1.append([])
    ws1.append(headers)
    
    entries = PMScheduleEntry.objects.filter(
        equipment__department=department,
        scheduled_date__month=month,
        scheduled_date__year=year
    ).order_by('equipment__name')
    
    for idx, e in enumerate(entries, 1):
        ws1.append([
            idx,
            e.equipment.name,
            e.equipment.equipment_class,
            e.equipment.frequency,
            e.equipment.scheduled_days,
            e.status,
            e.actual_date.strftime('%Y-%m-%d') if e.actual_date else '—'
        ])
        
    # Styling Excel Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Setup Sheet 2: Oil Test
    ws2 = wb.create_sheet(title="Oil Test Register")
    ws2.append(["Date", "Equipment", "Viscosity (cSt)", "Moisture", "NAS Class", "Test No", "Status", "Notification No", "Remarks"])
    oils = OilTestLog.objects.filter(equipment__department=department, date__month=month, date__year=year)
    for o in oils:
        ws2.append([
            o.date.strftime('%Y-%m-%d'),
            o.equipment.name,
            o.viscosity,
            o.moisture,
            o.nas_class,
            o.test_no,
            o.status,
            o.notification_no,
            o.remarks
        ])
        
    # Setup Sheet 3: WDA Log
    ws3 = wb.create_sheet(title="WDA Register")
    ws3.append(["Date", "Equipment", "Ratio", "DL", "DS", "WPC", "Slide", "Checked By", "Status", "Notification No", "Remarks"])
    wdas = WDALog.objects.filter(equipment__department=department, date__month=month, date__year=year)
    for w in wdas:
        ws3.append([
            w.date.strftime('%Y-%m-%d'),
            w.equipment.name,
            w.ratio,
            w.dl,
            w.ds,
            w.wpc,
            w.slide,
            w.checked_by,
            w.final_status,
            w.notification_no,
            w.remarks
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def calendar_month_name(month_idx):
    return [
        "", "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ][month_idx]
