import os
import sys
import django
import openpyxl
import xlrd

# Setup Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'main_portal.settings')
django.setup()

from tpm.models import Department
from delays.models import DelayDropdownOption

def seed_checklists():
    folder = r"E:\Jindal Steel Projects\DEPARTMENTS DASHBOARD\Checklists\SMS3"
    
    # 1. Find SMS-3 department
    target_dept = Department.objects.filter(code__iexact="SMS-3").first() or \
                  Department.objects.filter(code__iexact="SMS3").first() or \
                  Department.objects.filter(name__icontains="SMS-3").first() or \
                  Department.objects.filter(name__icontains="SMS3").first()
                  
    if not target_dept:
        print("Department SMS-3/SMS3 not found in database!")
        return
        
    print(f"Found Department: {target_dept.name} ({target_dept.code})")

    # Clear old options and checklists for target_dept to prevent duplicates/scrambling
    from delays.models import MaintenanceChecklist, DelayDropdownOption
    DelayDropdownOption.objects.filter(department=target_dept, category__in=['Equipment', 'Action']).delete()
    MaintenanceChecklist.objects.filter(department=target_dept).delete()
    print("Cleared existing dropdown options and checklist entries for SMS-3 to ensure a clean slate.")

    # Define helper to seed Option
    def seed_option(category, value, parent_value=None, is_header=False):
        opt, created = DelayDropdownOption.objects.update_or_create(
            department=target_dept,
            category=category,
            value=value[:255],  # max_length safety
            parent_value=parent_value[:255] if parent_value else None,
            defaults={'is_header': is_header}
        )
        if created:
            print(f"  [+] Created {category}: '{value}' (Parent: {parent_value}, Header: {is_header})")

    # Ensure 'Maintenance' is registered as Area (Sub-Agency) for checklists
    seed_option('Sub-Agency', 'Maintenance')

    # Checklist definitions
    checklists = [
        {
            "name": "Conveyors",
            "file": "Check list _Conveyors.xlsx",
            "type": "xlsx"
        },
        {
            "name": "DG Engines",
            "file": "Checklist (DG) (2).xls",
            "type": "xls"
        },
        {
            "name": "Pump Vibration Checking",
            "file": "Checklist for pump vibration checking.xlsx",
            "type": "xlsx"
        },
        {
            "name": "Ladle Repairing",
            "file": "Checklist_Laddle repairing.xlsx",
            "type": "xlsx"
        },
        {
            "name": "100T VD",
            "file": "Daily checklist _VD (2).xlsx",
            "type": "xlsx"
        },
        {
            "name": "EAF",
            "file": "Daily Checklist_ EAF.xls",
            "type": "xls"
        },
        {
            "name": "FES & Bag House",
            "file": "Daily Checklist_ RMH & FES-001.xls",
            "type": "xls"
        },
        {
            "name": "Ladle Refining Furnace (LRF)",
            "file": "Daily Checklist_LRF.xls",
            "type": "xls"
        },
        {
            "name": "Hydraulic",
            "file": "Hydraulic_checklist.xlsx",
            "type": "xlsx"
        }
    ]

    for cl in checklists:
        path = os.path.join(folder, cl["file"])
        print("-" * 50)
        print(f"Seeding checklist: {cl['name']} from {cl['file']}")
        if not os.path.exists(path):
            print(f"  Error: File not found at {path}")
            continue

        # Register the Checklist Name as an Equipment dropdown option under Area 'Maintenance'
        seed_option('Equipment', cl['name'], 'Maintenance')

        if cl["type"] == "xlsx":
            wb = openpyxl.load_workbook(path, data_only=True)
            
            if cl["name"] == "Conveyors":
                sheet = wb["Sheet1"]
                # Seed initial section header
                seed_option('Action', "CONVEYORS", cl['name'], is_header=True)
                
                checkpoints = [
                    "Pull cord", "Belt sway", "Coupling guard", "Tail pulley guard", 
                    "Take up pulley guard", "Emergency switch", "Walk way", "Hand railings", 
                    "Acess to conveyors", "Illumination Level", "ZSS", "Hooter"
                ]
                
                for row in list(sheet.iter_rows(min_row=3, values_only=True)):
                    col0 = row[0]
                    col1 = row[1]
                    
                    if col0 is not None:
                        col0_str = str(col0).strip()
                        if col0_str and (col1 is None or str(col1).strip() == ""):
                            if "Shift" in col0_str or "Mech" in col0_str or "Remarks" in col0_str:
                                continue
                            seed_option('Action', col0_str, cl['name'], is_header=True)
                            
                    if col1 is not None:
                        col1_str = str(col1).strip()
                        if col1_str:
                            if "Shift" in col1_str or "Mech" in col1_str or "Remarks" in col1_str:
                                continue
                            # Seed conveyor name as a sub-header
                            seed_option('Action', col1_str, cl['name'], is_header=True)
                            # Seed all 12 checkpoints prefixed with conveyor name for uniqueness
                            for cp in checkpoints:
                                seed_option('Action', f"{col1_str} - {cp}", cl['name'], is_header=False)
                        
            elif cl["name"] == "Pump Vibration Checking":
                for side in ["D.S", "N.D.S", "ROTTED", "NORMAL"]:
                    seed_option('Action', side, cl['name'], is_header=True)
                    for point in ["Horizontal", "Axial", "Vertical", "Temperature", "Pressure"]:
                        seed_option('Action', f"{side} - {point}", cl['name'], is_header=False)
                        
            elif cl["name"] == "Ladle Repairing":
                sheet = wb["Table 2"]
                for row in list(sheet.iter_rows(min_row=2, values_only=True)):
                    job = row[1]
                    if job and str(job).strip():
                        seed_option('Action', str(job).strip(), cl['name'], is_header=False)
                        
            elif cl["name"] == "100T VD":
                sheet = wb["Table 1"]
                current_eq = ""
                for row in list(sheet.iter_rows(min_row=3, values_only=True)):
                    eq_col = row[0]
                    check_item = row[1]
                    if eq_col:
                        current_eq = str(eq_col).strip()
                        seed_option('Action', current_eq, cl['name'], is_header=True)
                    if check_item:
                        seed_option('Action', f"{current_eq}: {str(check_item).strip()}", cl['name'], is_header=False)
                        
            elif cl["name"] == "Hydraulic":
                # Table 2
                if "Table 2" in wb.sheetnames:
                    sheet = wb["Table 2"]
                    for row in list(sheet.iter_rows(min_row=2, values_only=True)):
                        sr_no = row[0]
                        part = row[1]
                        if part:
                            part_str = str(part).strip()
                            if sr_no is None or str(sr_no).strip() == "":
                                seed_option('Action', part_str, cl['name'], is_header=True)
                            else:
                                seed_option('Action', part_str, cl['name'], is_header=False)
                # Table 3
                if "Table 3" in wb.sheetnames:
                    sheet = wb["Table 3"]
                    for row in list(sheet.iter_rows(min_row=1, values_only=True)):
                        part = row[0] or row[1]
                        if part:
                            part_str = str(part).strip()
                            if "Motor Cover" in part_str or "Process" in part_str:
                                seed_option('Action', part_str, cl['name'], is_header=True)
                            else:
                                seed_option('Action', part_str, cl['name'], is_header=False)

        else: # xls files
            wb = xlrd.open_workbook(path)
            
            if cl["name"] == "DG Engines":
                sheet = wb.sheet_by_index(0)
                checkpoints = []
                for r in range(3, 9):
                    val = str(sheet.cell_value(r, 1)).strip()
                    if val:
                        checkpoints.append(val)
                for engine in ["Fire Fighting DG Engine", "EAF DG Engine", "CCM DG Engine"]:
                    seed_option('Action', engine, cl['name'], is_header=True)
                    for cp in checkpoints:
                        seed_option('Action', f"{engine} - {cp}", cl['name'], is_header=False)
                
                seed_option('Action', "Jockey Pumps", cl['name'], is_header=True)
                for r in range(9, sheet.nrows):
                    val = str(sheet.cell_value(r, 1)).strip()
                    if val and val != "Check Points" and not val.startswith("Jockey"):
                        seed_option('Action', f"Jockey Pumps - {val}", cl['name'], is_header=False)
                        
            elif cl["name"] == "EAF":
                sheet = wb.sheet_by_index(0)
                current_eq = ""
                for r in range(5, sheet.nrows):
                    sno = str(sheet.cell_value(r, 0)).strip()
                    eq = str(sheet.cell_value(r, 1)).strip()
                    detail = str(sheet.cell_value(r, 2)).strip()
                    if eq:
                        is_hdr = False
                        if not sno or sno == "" or sno == "None":
                            is_hdr = True
                        elif not detail or detail == "" or detail == "None":
                            eq_lower = eq.lower()
                            if "flow" in eq_lower or "water" in eq_lower or "m3/hr" in eq_lower:
                                is_hdr = True
                            else:
                                is_hdr = False
                        
                        if is_hdr:
                            current_eq = eq
                            seed_option('Action', eq, cl['name'], is_header=True)
                        else:
                            val_to_seed = f"{current_eq} - {eq}" if detail == "" else f"{eq} - {detail}"
                            seed_option('Action', val_to_seed, cl['name'], is_header=False)
                        
            elif cl["name"] == "FES & Bag House":
                sheet = wb.sheet_by_index(0)
                current_hdr = "FES"
                for r in range(2, sheet.nrows):
                    sno = str(sheet.cell_value(r, 0)).strip()
                    part = str(sheet.cell_value(r, 1)).strip()
                    if part:
                        is_hdr = False
                        if not sno or sno == "" or sno == "None":
                            is_hdr = True
                        elif sno:
                            try:
                                float(sno)
                                is_hdr = False
                            except ValueError:
                                is_hdr = True
                        
                        # Skip signature fields/remarks lines
                        if "ENGINEER" in part or "OPERATOR" in part or "Remark" in part:
                            continue
                            
                        if is_hdr:
                            current_hdr = part
                            seed_option('Action', part, cl['name'], is_header=True)
                        else:
                            seed_option('Action', f"{current_hdr} - {part}", cl['name'], is_header=False)
                        
            elif cl["name"] == "Ladle Refining Furnace (LRF)":
                sheet = wb.sheet_by_index(0)
                current_eq = ""
                for r in range(4, sheet.nrows):
                    sno = str(sheet.cell_value(r, 0)).strip()
                    eq = str(sheet.cell_value(r, 1)).strip()
                    detail = str(sheet.cell_value(r, 2)).strip()
                    if eq:
                        is_hdr = False
                        if not sno or sno == "" or sno == "None":
                            is_hdr = True
                        elif not detail or detail == "" or detail == "None":
                            eq_lower = eq.lower()
                            if "flow" in eq_lower or "water" in eq_lower or "m3/hr" in eq_lower:
                                is_hdr = True
                            else:
                                is_hdr = False
                        
                        if is_hdr:
                            current_eq = eq
                            seed_option('Action', eq, cl['name'], is_header=True)
                        else:
                            val_to_seed = f"{current_eq} - {eq}" if detail == "" else f"{eq} - {detail}"
                            seed_option('Action', val_to_seed, cl['name'], is_header=False)

    print("Seeding SMS-3 daily checklists completed successfully!")

if __name__ == "__main__":
    seed_checklists()
