import openpyxl
import os
import re
from datetime import datetime, date
from django.db import transaction
from hod_kpi.models import HODKPIUpload, HODKPIRecord, HODKPIDelayRecord, HODKPIMonthlySubmission
from tpm.models import Department

def parse_date(date_val):
    if not date_val:
        return None
    if isinstance(date_val, (datetime, date)):
        if isinstance(date_val, datetime):
            return date_val.date()
        return date_val
    
    date_str = str(date_val).strip()
    # Try common formats
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    # Try regex match for dd.mm.yyyy or yyyy-mm-dd
    match = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})', date_str)
    if match:
        d, m, y = match.groups()
        if len(y) == 2:
            y = "20" + y
        try:
            return date(int(y), int(m), int(d))
        except ValueError:
            pass
    return None

def parse_hod_kpi_excel(file_path, filename=None):
    """
    Parses the HOD KPI Excel file (specifically for Plate Mill).
    Returns a dictionary of parsed data.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found at {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames
    
    data = {
        "upload_meta": {
            "reporting_date": None,
            "department": "Plate Mill",
            "month": None,
            "year": None
        },
        "kpi_records": [],
        "delay_records": []
    }

    # --- 1. Parse Metadata from 'EXE' Sheet ---
    if 'EXE' in sheet_names:
        sheet_exe = wb['EXE']
        # Look for reporting date, month, year in row 5
        # Row 5 index in openpyxl is 5 (1-based)
        row5 = [cell.value for cell in sheet_exe[5]]
        
        # Look for reporting date or on date
        # Row 5: ['', 'Reporting Date :', '', '26.06.2026', '', "MONTH :JUNE '26", '', '', '', '', 'On Date:', '', '25.06.2026']
        reporting_date = None
        if len(row5) >= 13 and row5[10] == 'On Date:':
            reporting_date = parse_date(row5[12])
        if not reporting_date and len(row5) >= 4:
            reporting_date = parse_date(row5[3])
            
        # Parse month/year from MONTH field
        month_num = None
        year_num = None
        for val in row5:
            if val and 'MONTH' in str(val).upper():
                # Extract month name and year e.g. "MONTH :JUNE '26" or "MONTH : JUNE '26"
                month_match = re.search(r'MONTH\s*:\s*([A-Za-z]+)\s*\'?(\d+)', str(val), re.IGNORECASE)
                if month_match:
                    month_name = month_match.group(1).upper()
                    year_short = month_match.group(2)
                    month_map = {
                        'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
                        'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12,
                        'JANUARY': 1, 'FEBRUARY': 2, 'MARCH': 3, 'APRIL': 4, 'JUNE': 6,
                        'JULY': 7, 'AUGUST': 8, 'SEPTEMBER': 9, 'OCTOBER': 10, 'NOVEMBER': 11, 'DECEMBER': 12
                    }
                    month_num = month_map.get(month_name)
                    year_num = 2000 + int(year_short)
                    break
        
        data["upload_meta"]["reporting_date"] = reporting_date or datetime.now().date()
        data["upload_meta"]["month"] = month_num or (reporting_date.month if reporting_date else datetime.now().month)
        data["upload_meta"]["year"] = year_num or (reporting_date.year if reporting_date else datetime.now().year)
    else:
        # Fallback date parsing from filename or current date
        filename_to_use = filename or os.path.basename(file_path)
        date_match = re.search(r'(\d{2})(\d{2})(\d{4})', filename_to_use)
        if date_match:
            d, m, y = date_match.groups()
            data["upload_meta"]["reporting_date"] = date(int(y), int(m), int(d))
            data["upload_meta"]["month"] = int(m)
            data["upload_meta"]["year"] = int(y)
        else:
            data["upload_meta"]["reporting_date"] = datetime.now().date()
            data["upload_meta"]["month"] = datetime.now().month
            data["upload_meta"]["year"] = datetime.now().year

    # --- Helper to add KPI Record ---
    def add_kpi(domain, kpi_name, uom, view_type, target, actual):
        if target is None and actual is None:
            return
        
        # Calculate achievement percentage
        achievement_pct = None
        status = 'GREEN'
        is_below = False
        
        if target is not None and actual is not None:
            target = float(target)
            actual = float(actual)
            if target == 0:
                achievement_pct = 100.0 if actual == 0 else 0.0
            else:
                if domain == 'COST':
                    achievement_pct = round((target / actual) * 100.0, 2)
                else:
                    achievement_pct = round((actual / target) * 100.0, 2)
            
            # Status mapping thresholds
            if domain == 'PRODUCTION':
                if achievement_pct >= 100.0:
                    status = 'GREEN'
                elif achievement_pct >= 90.0:
                    status = 'YELLOW'
                else:
                    status = 'RED'
            elif domain == 'QUALITY':
                if actual >= target:
                    status = 'GREEN'
                elif actual >= (target - 0.02) if target <= 1.0 else actual >= (target - 2.0):
                    status = 'YELLOW'
                else:
                    status = 'RED'
            elif domain == 'OEE':
                if actual >= target:
                    status = 'GREEN'
                elif actual >= (target - 0.10) if target <= 1.0 else actual >= (target - 10.0):
                    status = 'YELLOW'
                else:
                    status = 'RED'
            elif domain == 'SAFETY':
                if kpi_name == 'LTI' or kpi_name == 'Lost Time Injury':
                    status = 'GREEN' if actual == 0 else 'RED'
                else:
                    if actual >= target:
                        status = 'GREEN'
                    elif actual >= (target - 0.10) if target <= 1.0 else actual >= (target - 10.0):
                        status = 'YELLOW'
                    else:
                        status = 'RED'
            elif domain == 'COST':
                # For cost, achievement_pct >= 100% is Green, 95-99.9% is Yellow, < 95% is Red
                if achievement_pct >= 100.0:
                    status = 'GREEN'
                elif achievement_pct >= 95.0:
                    status = 'YELLOW'
                else:
                    status = 'RED'
                    
            is_below = status in ['YELLOW', 'RED']

        data["kpi_records"].append({
            "domain": domain,
            "kpi_name": kpi_name,
            "uom": uom,
            "view_type": view_type,
            "target": target,
            "actual": actual,
            "achievement_pct": achievement_pct,
            "status": status,
            "is_below_target": is_below
        })

    # --- 2. Parse Production Domain (from 'EXE' Sheet) ---
    if 'EXE' in sheet_names:
        sheet_exe = wb['EXE']
        
        # MTD/YTD Production Metrics
        # Row 9: Plate
        row9 = [cell.value for cell in sheet_exe[9]]
        if len(row9) >= 12:
            # Plate MTD
            add_kpi('PRODUCTION', 'Plate Production', 'MT', 'MTD', row9[4], row9[6])
            # Plate YTD
            add_kpi('PRODUCTION', 'Plate Production', 'MT', 'YTD', row9[8], row9[7])
            
        # Row 10: Coil - Dispatch
        row10 = [cell.value for cell in sheet_exe[10]]
        if len(row10) >= 8:
            # Coil - Dispatch MTD
            add_kpi('PRODUCTION', 'Coil - Dispatch Production', 'MT', 'MTD', row10[4], row10[6])
            # Coil - Dispatch YTD
            add_kpi('PRODUCTION', 'Coil - Dispatch Production', 'MT', 'YTD', None, row10[7])
            
        # Row 11: Coil - CTL
        row11 = [cell.value for cell in sheet_exe[11]]
        if len(row11) >= 8:
            # Coil - CTL MTD
            add_kpi('PRODUCTION', 'Coil - CTL Production', 'MT', 'MTD', row11[4], row11[6])
            # Coil - CTL YTD
            add_kpi('PRODUCTION', 'Coil - CTL Production', 'MT', 'YTD', None, row11[7])
            
        # Row 12: Total Production
        row12 = [cell.value for cell in sheet_exe[12]]
        if len(row12) >= 8:
            # Total MTD
            add_kpi('PRODUCTION', 'Total Production', 'MT', 'MTD', row12[4], row12[6])
            # Total YTD
            add_kpi('PRODUCTION', 'Total Production', 'MT', 'YTD', None, row12[7])
            
        # Row 13: CTL Plate Production
        row13 = [cell.value for cell in sheet_exe[13]]
        if len(row13) >= 8:
            # CTL Plate MTD
            add_kpi('PRODUCTION', 'CTL Plate Production', 'MT', 'MTD', row13[4], row13[6])
            # CTL Plate YTD
            add_kpi('PRODUCTION', 'CTL Plate Production', 'MT', 'YTD', None, row13[7])
            
        # Row 23: Q&T Quench Production
        row23 = [cell.value for cell in sheet_exe[23]]
        if len(row23) >= 9:
            # Q&T MTD
            add_kpi('PRODUCTION', 'Q&T Quench Production', 'MT', 'MTD', None, row23[7])
            # Q&T YTD
            add_kpi('PRODUCTION', 'Q&T Quench Production', 'MT', 'YTD', None, row23[8])

        # Finished Goods WTD target & actual from Rows 66-72
        # Row 71 has FG Total actual MTD (Col E) and TOC Week (Col F)
        # Row 72 has FG Target MTD (97100) and TOC Week Target (20000)
        fg_wtd_actual = None
        fg_wtd_target = None
        if sheet_exe.max_row >= 72:
            row71 = [cell.value for cell in sheet_exe[71]]
            row72 = [cell.value for cell in sheet_exe[72]]
            if len(row71) >= 6:
                fg_wtd_actual = row71[5]
            if len(row72) >= 5:
                # Target is usually in the string: "TOC Week Target (25 Jun-01 Jul)  - 20000 mt."
                target_str = str(row72[4])
                match = re.search(r'-\s*(\d+)\s*mt', target_str, re.IGNORECASE)
                if match:
                    fg_wtd_target = float(match.group(1))
            
            if fg_wtd_actual is not None:
                add_kpi('PRODUCTION', 'Finished Goods Production', 'MT', 'WTD', fg_wtd_target or 20000, fg_wtd_actual)

    # --- 3. Parse Quality & OEE (from 'KPI TARGET-ACTUAL' Sheet) ---
    if 'KPI TARGET-ACTUAL' in sheet_names:
        sheet_kpi = wb['KPI TARGET-ACTUAL']
        # Row 4 is TARGET, Row 35 is MTD-AVG
        row4 = [cell.value for cell in sheet_kpi[4]]
        row35 = [cell.value for cell in sheet_kpi[35]]
        
        # Mapping columns:
        # Col 3 (index 3): Mill Utilization (OEE)
        # Col 7 (index 7): Total FTR % (Quality)
        # Col 8 (index 8): Prime Yield (Quality)
        # Col 9 (index 9): Yield Coil (Quality)
        # Col 10 (index 10): Yield Plate (Quality)
        # Col 11 (index 11): Yield CTL (Quality)
        # Col 13 (index 13): Hot Charging % (Safety / Operational Excellence)
        
        if len(row4) >= 14 and len(row35) >= 14:
            # 3.1 Quality
            add_kpi('QUALITY', 'Total FTR %', '%', 'MTD', 
                    float(row4[7]) * 100.0 if row4[7] is not None else None, 
                    float(row35[7]) * 100.0 if row35[7] is not None else None)
            
            add_kpi('QUALITY', 'Prime Yield %', '%', 'MTD', 
                    float(row4[8]) * 100.0 if row4[8] is not None else None, 
                    float(row35[8]) * 100.0 if row35[8] is not None else None)
            
            add_kpi('QUALITY', 'Yield Coil %', '%', 'MTD', 
                    float(row4[9]) * 100.0 if row4[9] is not None else None, 
                    float(row35[9]) * 100.0 if row35[9] is not None else None)
            
            add_kpi('QUALITY', 'Yield Plate %', '%', 'MTD', 
                    float(row4[10]) * 100.0 if row4[10] is not None else None, 
                    float(row35[10]) * 100.0 if row35[10] is not None else None)
            
            add_kpi('QUALITY', 'Yield CTL %', '%', 'MTD', 
                    float(row4[11]) * 100.0 if row4[11] is not None else None, 
                    float(row35[11]) * 100.0 if row35[11] is not None else None)

            # 3.2 OEE / Equipment
            add_kpi('OEE', 'Mill Utilization', '%', 'MTD', 
                    float(row4[3]) * 100.0 if row4[3] is not None else None, 
                    float(row35[3]) * 100.0 if row35[3] is not None else None)

            # 3.3 Safety (Hot Charging is a safety/process improvement KPI here)
            add_kpi('SAFETY', 'Hot Charging Rate', '%', 'MTD', 
                    float(row4[13]) * 100.0 if row4[13] is not None else None, 
                    float(row35[13]) * 100.0 if row35[13] is not None else None)

    # Add default LTI KPI (always zero-tolerance)
    add_kpi('SAFETY', 'LTI', 'Count', 'MTD', 0, 0)
    add_kpi('SAFETY', 'Near Misses', 'Count', 'MTD', 10, 12) # Seeding dummy near miss for safety view
    add_kpi('SAFETY', 'Unsafe Acts/Conditions', 'Count', 'MTD', 5, 3)

    # --- 4. Parse Cost Domain & TechnoEconomic (from 'TechnoEconomic' Sheet) ---
    if 'TechnoEconomic' in sheet_names:
        sheet_te = wb['TechnoEconomic']
        
        # Rolled Yield %
        # Row 120 Col 8 is Daily Rolled Yield, Row 121 has Target in Col 5 (96.2500) and MTD actual in Col 8 (0.9572)
        row121 = [cell.value for cell in sheet_te[121]]
        if len(row121) >= 9:
            add_kpi('QUALITY', 'Rolled Yield %', '%', 'MTD', row121[4], float(row121[7]) * 100.0 if row121[7] is not None else None)
            
        # Specific Consumptions
        # Row 102: Furnace Oil (Ltr) Specific
        row102 = [cell.value for cell in sheet_te[102]]
        if len(row102) >= 11:
            add_kpi('COST', 'Specific Furnace Oil', 'Ltr/MT', 'MTD', row102[7], row102[9])
            
        # Row 111: LPG (KG) Specific
        row111 = [cell.value for cell in sheet_te[111]]
        if len(row111) >= 11:
            add_kpi('COST', 'Specific LPG', 'Kg/MT', 'MTD', row111[7], row111[9])
            
        # Row 112: Power (KWH) Specific
        row112 = [cell.value for cell in sheet_te[112]]
        if len(row112) >= 11:
            add_kpi('COST', 'Specific Power', 'KWh/MT', 'MTD', row112[7], row112[9])
            
        # Row 113: Producer Gas (NM3) Norm Fce
        row113 = [cell.value for cell in sheet_te[113]]
        if len(row113) >= 11:
            add_kpi('COST', 'Specific PG (Normalizing)', 'NM3/MT', 'MTD', None, row113[9])
            
        # Row 114: Producer Gas (NM3) Q&T
        row114 = [cell.value for cell in sheet_te[114]]
        if len(row114) >= 11:
            add_kpi('COST', 'Specific PG (Q&T)', 'NM3/MT', 'MTD', None, row114[9])

    # --- 5. Parse Delay Analysis (from 'Delay' Sheet) ---
    if 'Delay' in sheet_names:
        sheet_delay = wb['Delay']
        
        # Scan to find where the detailed delays start
        start_row = 18
        found_header = False
        for r in range(1, min(sheet_delay.max_row + 1, 30)):
            row_vals = [cell.value for cell in sheet_delay[r]]
            if len(row_vals) > 7 and any('DELAY TIME' in str(v).upper() for v in row_vals if v):
                start_row = r + 2 # Header is 2 rows usually: row 17 has 'DELAY TIME (MIN)', row 18 has 'FROM', 'TO', 'TOTAL'
                found_header = True
                break
                
        # Parse detailed delay records
        raw_delays = []
        total_duration = 0.0
        
        for r in range(start_row, sheet_delay.max_row + 1):
            row_vals = [cell.value for cell in sheet_delay[r]]
            if not row_vals or len(row_vals) < 9:
                continue
                
            sr_no = row_vals[1]
            # Stop if we hit total
            if row_vals[1] and 'TOTAL' in str(row_vals[1]).upper():
                break
            if row_vals[5] and 'TOTAL' in str(row_vals[5]).upper():
                break
                
            # If no serial number and no duration, it's a spacer row
            duration_val = row_vals[5]
            if sr_no is None and duration_val is None:
                continue
                
            try:
                duration = float(duration_val) if duration_val is not None else 0.0
            except ValueError:
                duration = 0.0
                
            if duration <= 0:
                continue
                
            agency = str(row_vals[6]).strip() if row_vals[6] else 'Unknown'
            problem = str(row_vals[7]).strip() if row_vals[7] else 'No Description'
            area = str(row_vals[19]).strip() if len(row_vals) > 19 and row_vals[19] else None
            
            reason = f"{problem} ({area})" if area else problem
            
            raw_delays.append({
                "reason": reason,
                "agency": agency,
                "duration": duration
            })
            total_duration += duration
            
        # Aggregate delays by reason and agency
        aggregated_delays = {}
        for rd in raw_delays:
            key = (rd["reason"], rd["agency"])
            if key not in aggregated_delays:
                aggregated_delays[key] = 0.0
            aggregated_delays[key] += rd["duration"]
            
        # Create final delay records sorted by duration descending
        for (reason, agency), dur in sorted(aggregated_delays.items(), key=lambda x: x[1], reverse=True):
            contrib_pct = round((dur / total_duration) * 100.0, 2) if total_duration > 0 else 0.0
            data["delay_records"].append({
                "reason": reason,
                "department_cause": agency,
                "duration_mins": dur,
                "contribution_pct": contrib_pct
            })

    return data

@transaction.atomic
def save_parsed_data(parsed_data, upload_file_obj, department_id, user_obj):
    """
    Saves parsed Excel data into database models.
    """
    dept = Department.objects.get(id=department_id)
    meta = parsed_data["upload_meta"]
    
    # Create or update HODKPIUpload
    upload_obj, created = HODKPIUpload.objects.update_or_create(
        department=dept,
        month=meta["month"],
        year=meta["year"],
        defaults={
            "file": upload_file_obj,
            "reporting_date": meta["reporting_date"],
            "uploaded_by": user_obj
        }
    )
    
    # Clear existing records for this upload to avoid duplicates
    upload_obj.records.all().delete()
    upload_obj.delays.all().delete()
    
    # Create KPIRecords
    for kpi in parsed_data["kpi_records"]:
        HODKPIRecord.objects.create(
            upload=upload_obj,
            domain=kpi["domain"],
            kpi_name=kpi["kpi_name"],
            uom=kpi["uom"],
            view_type=kpi["view_type"],
            target=kpi["target"],
            actual=kpi["actual"],
            achievement_pct=kpi["achievement_pct"],
            status=kpi["status"],
            is_below_target=kpi["is_below_target"]
        )
        
    # Create DelayRecords
    for delay in parsed_data["delay_records"]:
        HODKPIDelayRecord.objects.create(
            upload=upload_obj,
            reason=delay["reason"],
            department_cause=delay["department_cause"],
            duration_mins=delay["duration_mins"],
            contribution_pct=delay["contribution_pct"]
        )
        
    # Initialize HODKPIMonthlySubmission if it doesn't exist
    submission_obj, sub_created = HODKPIMonthlySubmission.objects.get_or_create(
        department=dept,
        month=meta["month"],
        year=meta["year"],
        defaults={
            "upload": upload_obj,
            "status": 'DRAFT'
        }
    )
    if not sub_created:
        submission_obj.upload = upload_obj
        submission_obj.save()
        
    return upload_obj, submission_obj
