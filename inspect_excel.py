import openpyxl

wb = openpyxl.load_workbook(r"e:\Jindal Steel Projects\DEPARTMENTS DASHBOARD\Delays Portal\3 month Dec24-Feb 2025 SMS2.xlsx", data_only=True)
for name in wb.sheetnames:
    sheet = wb[name]
    for r in range(1, sheet.max_row + 1):
        row_vals = [cell.value for cell in sheet[r]]
        if any(row_vals):
            # Check if any cell has duration 2055 or 190 or 20.0
            row_str = str(row_vals).upper()
            if '2055' in row_str or 'SEGMENT 1' in row_str or 'CRANE (180T)' in row_str or 'COPPER PAD' in row_str:
                print(f"Sheet: {name}, Row {r}: {row_vals[:15]}")
