import openpyxl

file_path = r"e:\Jindal Steel Projects\DEPARTMENTS DASHBOARD\CMC Portal\CMC Requirements.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['CMC Schedule']

print("=== CMC SCHEDULE GRID INSPECTION ===")

# Row 3 contains month labels, Row 4 contains A, B, C, D
month_headers = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
sub_headers = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

# Let's see some equipment rows
rows = list(sheet.iter_rows(min_row=5, values_only=True))

print(f"Total equipment rows in Excel: {len(rows)}")

# Collect unique departments, frequencies, and classes
depts = set()
classes = set()
frequencies = set()
status_values = set()

for r in rows:
    if r[1] is not None:
        depts.add(r[1])
    if r[3] is not None:
        classes.add(r[3])
    if r[10] is not None:
        frequencies.add(r[10])
    
    # Check schedule grid values (cols 12 to 59)
    for val in r[12:60]:
        if val is not None:
            status_values.add(str(val))

print("Unique Departments in Excel:", sorted(list(depts)))
print("Unique Classes in Excel:", classes)
print("Unique Frequencies in Excel:", frequencies)
print("Sample of status/date values in grid:", list(status_values)[:30])

# Show a few specific rows to see what is written in the grid
print("\nDetailed grid example for first 3 rows:")
for i in range(3):
    r = rows[i]
    print(f"\nEquipment: {r[2]} ({r[1]})")
    print(f"  Class: {r[3]} | SAP Mech: {r[4]} | SAP Elec: {r[5]} | Freq: {r[10]}")
    # Print non-None schedule grid cells
    sched_cells = []
    for col_idx in range(12, 60):
        val = r[col_idx]
        if val is not None:
            # Determine month and sub-column (A, B, C, D)
            month_name = None
            # Find the header column index
            for m_idx in range(col_idx, 11, -1):
                if month_headers[m_idx] is not None:
                    month_name = month_headers[m_idx]
                    break
            sub_col = sub_headers[col_idx]
            sched_cells.append(f"{month_name}({sub_col})={val}")
    print("  Scheduled:", ", ".join(sched_cells))
