import openpyxl

path = r"e:\Jindal Steel Projects\DEPARTMENTS DASHBOARD\EFMEA\fmea\Rail Mill Raigarh - EFMEA and Action Plan.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb['Vertical Roll Balancing Cylinde']

print("Row 5 (Headers) vs Row 10 (Values):")
for col_idx in range(1, sheet.max_column + 1):
    h_val = sheet.cell(row=5, column=col_idx).value
    d_val = sheet.cell(row=10, column=col_idx).value
    print(f"Col {col_idx:2d} ({openpyxl.utils.get_column_letter(col_idx)}): Header={repr(h_val)} | Value={repr(d_val)}")
