import openpyxl
import os

path = r"e:\Jindal Steel Projects\DEPARTMENTS DASHBOARD\EFMEA\fmea\FMEA Format.xlsx"
if os.path.exists(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb['FMEA']
    print("FMEA Format.xlsx Headers (Row 5):")
    for col_idx in range(1, 25):
        h_val = sheet.cell(row=5, column=col_idx).value
        print(f"Col {col_idx:2d} ({openpyxl.utils.get_column_letter(col_idx)}): {repr(h_val)}")
else:
    print(f"Template not found at {path}")
