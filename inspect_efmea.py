import openpyxl

files = {
    "uploaded_past_data": r"e:\Jindal Steel Projects\DEPARTMENTS DASHBOARD\EFMEA\fmea\Rail Mill Raigarh - EFMEA and Action Plan.xlsx",
    "fmea_format": r"e:\Jindal Steel Projects\DEPARTMENTS DASHBOARD\EFMEA\fmea\FMEA Format.xlsx"
}

for name, path in files.items():
    print(f"\n========================================\nFile: {name}\nPath: {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet: {sheet_name} (Dimensions: {sheet.max_row} rows x {sheet.max_column} columns)")
            # Print first 12 rows
            for r in range(1, min(15, sheet.max_row + 1)):
                row_vals = [cell.value for cell in sheet[r]]
                if any(cell is not None for cell in row_vals):
                    print(f"Row {r:2d}: {row_vals[:15]}")
    except Exception as e:
        print(f"Error loading file: {e}")
