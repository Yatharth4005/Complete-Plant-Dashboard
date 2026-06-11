import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def convert_color(xlrd_color_index, book):
    # Map xlrd color index to hex color string
    if xlrd_color_index is None:
        return None
    try:
        rgb = book.colour_map.get(xlrd_color_index)
        if rgb:
            return f"{rgb[0]:02d}{rgb[1]:02d}{rgb[2]:02d}"
    except:
        pass
    return None

def main():
    xls_path = "KAIZEN - Blank Format.xls"
    xlsx_path = "KAIZEN - Blank Format.xlsx"
    
    try:
        # formatting_info=True works for xls files in xlrd
        book = xlrd.open_workbook(xls_path, formatting_info=True)
        xls_sheet = book.sheet_by_index(0)
        
        wb = openpyxl.Workbook()
        xlsx_sheet = wb.active
        xlsx_sheet.title = "KAIZEN IDEA SHEET"
        xlsx_sheet.views.sheetView[0].showGridLines = True
        
        # 1. Copy column widths
        for col_idx in range(xls_sheet.ncols):
            col_width = xls_sheet.computed_column_width(col_idx)
            # xlrd column width is in 1/256 of character width. Openpyxl is in character width.
            xlsx_col_letter = get_column_letter(col_idx + 1)
            xlsx_sheet.column_dimensions[xlsx_col_letter].width = max(col_width / 256.0, 10.0)
            
        # 2. Copy row heights
        for row_idx in range(xls_sheet.nrows):
            rowinfo = xls_sheet.rowinfo_map.get(row_idx)
            if rowinfo and rowinfo.height > 0:
                # height is in twips (1/20 of a point)
                xlsx_sheet.row_dimensions[row_idx + 1].height = rowinfo.height / 20.0
                
        # 3. Copy cells
        for r in range(xls_sheet.nrows):
            for c in range(xls_sheet.ncols):
                cell = xls_sheet.cell(r, c)
                xlsx_cell = xls_sheet.cell(r, c)
                
                # Write value
                xlsx_sheet.cell(row=r+1, column=c+1, value=cell.value)
                
                # Check formatting
                xf = book.xf_list[cell.xf_index]
                
                # Font
                font = book.font_list[xf.font_index]
                font_name = font.name
                font_size = font.height / 20.0 # height in twips
                font_bold = font.bold
                font_italic = font.italic
                
                # Colors and Fills
                bg_color = None
                if xf.background:
                    # xlrd fill pattern background color
                    bg_color_idx = xf.background.pattern_colour_index
                    if bg_color_idx != 64: # 64 is default auto/no fill
                        rgb = book.colour_map.get(bg_color_idx)
                        if rgb:
                            bg_color = f"{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"
                
                # Alignment
                alignment = xf.alignment
                horiz_align = 'left'
                if alignment.hor_align == 2: # Center
                    horiz_align = 'center'
                elif alignment.hor_align == 3: # Right
                    horiz_align = 'right'
                    
                vert_align = 'center'
                if alignment.vert_align == 0: # Top
                    vert_align = 'top'
                elif alignment.vert_align == 2: # Bottom
                    vert_align = 'bottom'
                
                # Apply style to cell
                cell_to_style = xlsx_sheet.cell(row=r+1, column=c+1)
                cell_to_style.font = Font(name=font_name, size=font_size, bold=font_bold, italic=font_italic)
                cell_to_style.alignment = Alignment(horizontal=horiz_align, vertical=vert_align, wrap_text=True)
                
                if bg_color:
                    cell_to_style.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                
                # Borders
                border = xf.border
                thin_side = Side(border_style="thin", color="000000")
                medium_side = Side(border_style="medium", color="000000")
                
                # Check borders in xlrd and apply
                left_border = thin_side if border.left_line_style > 0 else None
                right_border = thin_side if border.right_line_style > 0 else None
                top_border = thin_side if border.top_line_style > 0 else None
                bottom_border = thin_side if border.bottom_line_style > 0 else None
                
                cell_to_style.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
                
        # 4. Copy Merged Cells
        for crange in xls_sheet.merged_cells:
            rlo, rhi, clo, chi = crange
            # openpyxl uses 1-based indexing, inclusive range
            xlsx_sheet.merge_cells(
                start_row=rlo + 1,
                start_column=clo + 1,
                end_row=rhi,
                end_column=chi
            )
            
        wb.save(xlsx_path)
        print("CONVERSION_SUCCESSFUL")
        
    except Exception as e:
        import traceback
        print(f"CONVERSION_ERROR: {e}")
        traceback.print_exc()

if __name__ == '__main__':
    main()
