import os
import io
import json
import random
import datetime
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.conf import settings

from tpm.models import Department, KaizenSheet
from tpm.utils.decorators import dept_access_required
from tpm.utils.calculations import update_jh_kaizen_kpi_value

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as ReportLabImage, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def format_date_to_form(date_str):
    if not date_str:
        return ''
    try:
        dt = datetime.datetime.strptime(date_str.strip(), '%d-%m-%Y')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass
    try:
        dt = datetime.datetime.strptime(date_str.strip(), '%Y-%m-%d')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass
    return date_str

def format_date_to_db(date_str):
    if not date_str:
        return ''
    try:
        dt = datetime.datetime.strptime(date_str.strip(), '%Y-%m-%d')
        return dt.strftime('%d-%m-%Y')
    except ValueError:
        pass
    try:
        dt = datetime.datetime.strptime(date_str.strip(), '%d-%m-%Y')
        return dt.strftime('%d-%m-%Y')
    except ValueError:
        pass
    return date_str

def generate_random_kaizen_no():
    existing_nos = set(KaizenSheet.objects.values_list('kaizen_no', flat=True))
    for _ in range(1000):
        num = random.randint(1, 999)
        candidate = f"K-{num:03d}"
        if candidate not in existing_nos:
            return candidate
    return f"K-{KaizenSheet.objects.count() + 1:03d}"

def update_kaizen_kpis_for_sheet(sheet):
    if sheet.pillar != 'JH':
        return
    if sheet.finish_date:
        for fmt in ('%d-%m-%Y', '%Y-%m-%d'):
            try:
                dt = datetime.datetime.strptime(sheet.finish_date.strip(), fmt)
                update_jh_kaizen_kpi_value(sheet.department, dt.month, dt.year)
                break
            except ValueError:
                pass

@login_required
@dept_access_required
def kaizen_list_partial(request, dept_id, pillar_id):
    dept = get_object_or_404(Department, id=dept_id)
    sheets = KaizenSheet.objects.filter(department=dept, pillar=pillar_id).order_by('-created_at')
    
    context = {
        'dept': dept,
        'pillar_id': pillar_id,
        'sheets': sheets,
    }
    return render(request, 'partials/_kaizen_list.html', context)

@login_required
@dept_access_required
def kaizen_edit_partial(request, dept_id, pillar_id, kaizen_id=None):
    dept = get_object_or_404(Department, id=dept_id)
    kaizen = None
    if kaizen_id:
        kaizen = get_object_or_404(KaizenSheet, id=kaizen_id, department=dept, pillar=pillar_id)
        if kaizen.start_date:
            kaizen.start_date_form = format_date_to_form(kaizen.start_date)
        if kaizen.finish_date:
            kaizen.finish_date_form = format_date_to_form(kaizen.finish_date)
            
        formatted_deployment = []
        for item in kaizen.horizontal_deployment:
            new_item = item.copy()
            if item.get('target_date'):
                new_item['target_date_form'] = format_date_to_form(item['target_date'])
            else:
                new_item['target_date_form'] = ''
            formatted_deployment.append(new_item)
        kaizen.horizontal_deployment = formatted_deployment
        
    # Auto-prefill variables for new sheets
    prefills = {
        'kaizen_no': generate_random_kaizen_no(),
        'activities': [pillar_id],
        'result_areas': [],
        'team_members': ['', '', '', ''],
        'tangible_benefits': ['', '', '', '', ''],
        'intangible_benefits': ['', '', '', '', ''],
        'horizontal_deployment': [
            {'sl_no': '1', 'area_equip': '', 'target_date': '', 'responsibility': '', 'status': 'Pending'},
            {'sl_no': '2', 'area_equip': '', 'target_date': '', 'responsibility': '', 'status': 'Pending'},
            {'sl_no': '3', 'area_equip': '', 'target_date': '', 'responsibility': '', 'status': 'Pending'},
        ]
    }
    
    context = {
        'dept': dept,
        'pillar_id': pillar_id,
        'kaizen': kaizen,
        'prefills': prefills,
    }
    return render(request, 'partials/_kaizen_form.html', context)

@login_required
@dept_access_required
@require_POST
def kaizen_save(request, dept_id, pillar_id, kaizen_id=None):
    dept = get_object_or_404(Department, id=dept_id)
    
    old_finish_date = None
    if kaizen_id:
        kaizen = get_object_or_404(KaizenSheet, id=kaizen_id, department=dept, pillar=pillar_id)
        old_finish_date = kaizen.finish_date
    else:
        kaizen = KaizenSheet(department=dept, pillar=pillar_id, created_by=request.user)
        
    # Standard text fields
    kaizen.kaizen_no = request.POST.get('kaizen_no', '').strip()
    kaizen.loss_name = request.POST.get('loss_name', '').strip()
    kaizen.area_equipment = request.POST.get('area_equipment', '').strip()
    kaizen.circle_name = request.POST.get('circle_name', '').strip()
    kaizen.theme = request.POST.get('theme', '').strip()
    kaizen.idea = request.POST.get('idea', '').strip()
    kaizen.benchmark = request.POST.get('benchmark', '').strip()
    kaizen.target = request.POST.get('target', '').strip()
    kaizen.start_date = format_date_to_db(request.POST.get('start_date', '').strip())
    kaizen.finish_date = format_date_to_db(request.POST.get('finish_date', '').strip())
    kaizen.team_leader = request.POST.get('team_leader', '').strip()
    kaizen.analysis = request.POST.get('analysis', '').strip()
    kaizen.result_text = request.POST.get('result_text', '').strip()
    
    # Checkboxes: Activities & Result Areas
    kaizen.activities = request.POST.getlist('activities')
    kaizen.result_areas = request.POST.getlist('result_areas')
    
    # JSON lists: Team Members
    members = [
        request.POST.get('member_1', '').strip(),
        request.POST.get('member_2', '').strip(),
        request.POST.get('member_3', '').strip(),
        request.POST.get('member_4', '').strip(),
    ]
    kaizen.team_members = [m for m in members if m]
    
    # Tangible & Intangible benefits lists
    tangibles = [
        request.POST.get('tangible_1', '').strip(),
        request.POST.get('tangible_2', '').strip(),
        request.POST.get('tangible_3', '').strip(),
        request.POST.get('tangible_4', '').strip(),
        request.POST.get('tangible_5', '').strip(),
    ]
    kaizen.tangible_benefits = [t for t in tangibles if t]
    
    intangibles = [
        request.POST.get('intangible_1', '').strip(),
        request.POST.get('intangible_2', '').strip(),
        request.POST.get('intangible_3', '').strip(),
        request.POST.get('intangible_4', '').strip(),
        request.POST.get('intangible_5', '').strip(),
    ]
    kaizen.intangible_benefits = [i for i in intangibles if i]
    
    # Horizontal deployment list from form dynamic inputs
    h_sl = request.POST.getlist('h_sl_no')
    h_area = request.POST.getlist('h_area_equip')
    h_date = request.POST.getlist('h_target_date')
    h_resp = request.POST.getlist('h_responsibility')
    h_stat = request.POST.getlist('h_status')
    
    deployment_data = []
    for idx in range(len(h_sl)):
        # Make sure there is at least some text filled
        if h_area[idx].strip() or h_resp[idx].strip():
            deployment_data.append({
                'sl_no': h_sl[idx].strip() or str(idx + 1),
                'area_equip': h_area[idx].strip(),
                'target_date': format_date_to_db(h_date[idx].strip()),
                'responsibility': h_resp[idx].strip(),
                'status': h_stat[idx].strip() or 'Pending',
            })
    kaizen.horizontal_deployment = deployment_data
    
    # Image uploads
    if 'before_image' in request.FILES:
        kaizen.before_image = request.FILES['before_image']
    if 'after_image' in request.FILES:
        kaizen.after_image = request.FILES['after_image']
    if 'result_image' in request.FILES:
        kaizen.result_image = request.FILES['result_image']
        
    kaizen.save()
    
    # Sync Jishu Hozen Kaizen Completed Count KPI actuals
    update_kaizen_kpis_for_sheet(kaizen)
    if old_finish_date and old_finish_date != kaizen.finish_date:
        for fmt in ('%d-%m-%Y', '%Y-%m-%d'):
            try:
                dt = datetime.datetime.strptime(old_finish_date.strip(), fmt)
                update_jh_kaizen_kpi_value(dept, dt.month, dt.year)
                break
            except ValueError:
                pass
                
    return kaizen_list_partial(request, dept.id, pillar_id)

@login_required
@dept_access_required
@require_POST
def kaizen_delete(request, dept_id, pillar_id, kaizen_id):
    dept = get_object_or_404(Department, id=dept_id)
    kaizen = get_object_or_404(KaizenSheet, id=kaizen_id, department=dept, pillar=pillar_id)
    
    finish_date = kaizen.finish_date
    sheet_pillar = kaizen.pillar
    
    kaizen.delete()
    
    # Sync Jishu Hozen Kaizen Completed Count KPI actuals
    if sheet_pillar == 'JH' and finish_date:
        for fmt in ('%d-%m-%Y', '%Y-%m-%d'):
            try:
                dt = datetime.datetime.strptime(finish_date.strip(), fmt)
                update_jh_kaizen_kpi_value(dept, dt.month, dt.year)
                break
            except ValueError:
                pass
                
    return kaizen_list_partial(request, dept_id, pillar_id)

@login_required
def download_excel(request, kaizen_id):
    kaizen = get_object_or_404(KaizenSheet, id=kaizen_id)
    
    wb_content = generate_kaizen_excel(kaizen)
    
    filename = f"KAIZEN_{kaizen.kaizen_no or 'Draft'}_{kaizen.pillar}.xlsx"
    response = HttpResponse(
        wb_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def download_pdf(request, kaizen_id):
    kaizen = get_object_or_404(KaizenSheet, id=kaizen_id)
    
    pdf_content = generate_kaizen_pdf(kaizen)
    
    filename = f"KAIZEN_{kaizen.kaizen_no or 'Draft'}_{kaizen.pillar}.pdf"
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def write_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ == 'MergedCell':
        for r in ws.merged_cells.ranges:
            if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
                ws.cell(row=r.min_row, column=r.min_col, value=value)
                return
    else:
        cell.value = value


def generate_kaizen_excel(kaizen):
    template_path = os.path.join(settings.BASE_DIR, 'KAIZEN - Blank Format.xlsx')
    if not os.path.exists(template_path):
        template_path = os.path.join(settings.BASE_DIR, 'TPM Portal', 'KAIZEN - Blank Format.xlsx')
        
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Fill Kaizen No
    write_cell(ws, 3, 2, f"KAIZEN NO.  {kaizen.kaizen_no}")
    
    # Fill Activity checklist
    activity_cols = {
        'KK': 7, 'JH': 8, 'QM': 9, 'PM': 10, 'SHE': 11, 'OTPM': 12, 'DM': 13, 'ET': 14
    }
    for act, col in activity_cols.items():
        val = act
        if act in kaizen.activities:
            val = f"✓ {act}"
        write_cell(ws, 3, col, val)
        
    # Loss Name
    write_cell(ws, 4, 7, kaizen.loss_name)
    
    # Result Area checklist
    result_cols = {
        'S': 7, 'Q': 9, 'P': 10, 'C': 11, 'D': 12, 'M': 13
    }
    for res, col in result_cols.items():
        val = res
        if res in kaizen.result_areas:
            val = f"✓ {res}"
        write_cell(ws, 5, col, val)
        
    # Department, Area, Circle Name
    write_cell(ws, 6, 2, f"DEPARTMENT :  {kaizen.department.name}")
    write_cell(ws, 6, 6, f"AREA/EQUIPMENT :  {kaizen.area_equipment}")
    write_cell(ws, 6, 10, f"TPM CIRCLE NAME:  {kaizen.circle_name}")
    
    # Theme & Idea
    write_cell(ws, 10, 2, kaizen.theme)
    write_cell(ws, 10, 6, kaizen.idea)
    
    # Benchmark, Target, Start, Finish
    write_cell(ws, 8, 12, kaizen.benchmark)
    write_cell(ws, 9, 12, kaizen.target)
    write_cell(ws, 10, 12, kaizen.start_date)
    write_cell(ws, 11, 12, kaizen.finish_date)
    
    # Team Members
    write_cell(ws, 13, 12, kaizen.team_leader)
    members = kaizen.team_members
    for idx in range(4):
        val = members[idx] if idx < len(members) else ""
        write_cell(ws, 14 + idx, 12, val)
        
    # Benefits: Tangible
    tangible = kaizen.tangible_benefits
    for idx in range(5):
        val = tangible[idx] if idx < len(tangible) else ""
        write_cell(ws, 20 + idx, 11, val)
        
    # Benefits: Intangible
    intangible = kaizen.intangible_benefits
    for idx in range(5):
        val = intangible[idx] if idx < len(intangible) else ""
        write_cell(ws, 26 + idx, 11, val)
        
    # Analysis & Result Text
    write_cell(ws, 32, 2, kaizen.analysis)
    write_cell(ws, 32, 6, kaizen.result_text)
    
    # Horizontal Deployment
    deployment = kaizen.horizontal_deployment
    for idx in range(3):
        row_idx = 35 + idx
        if idx < len(deployment):
            item = deployment[idx]
            write_cell(ws, row_idx, 10, item.get('sl_no', str(idx+1)))
            write_cell(ws, row_idx, 11, item.get('area_equip', ''))
            write_cell(ws, row_idx, 12, item.get('target_date', ''))
            write_cell(ws, row_idx, 14, item.get('responsibility', ''))
            write_cell(ws, row_idx, 16, item.get('status', ''))
            
    # Add Images (Before, After, Result Graph)
    if kaizen.before_image:
        try:
            img_path = kaizen.before_image.path
            img = OpenpyxlImage(img_path)
            img.width = 240
            img.height = 185
            ws.add_image(img, 'B19')
        except Exception as e:
            print(f"Error adding before image: {e}")
            
    if kaizen.after_image:
        try:
            img_path = kaizen.after_image.path
            img = OpenpyxlImage(img_path)
            img.width = 240
            img.height = 185
            ws.add_image(img, 'F19')
        except Exception as e:
            print(f"Error adding after image: {e}")
            
    if kaizen.result_image:
        try:
            img_path = kaizen.result_image.path
            img = OpenpyxlImage(img_path)
            img.width = 240
            img.height = 115
            ws.add_image(img, 'F32')
        except Exception as e:
            print(f"Error adding result image: {e}")
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_kaizen_pdf(kaizen):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
    )
    
    # col widths to sum to 555 (A4 is 595 width, 20 margin each side)
    col_widths = [15, 60, 45, 45, 45, 80, 20, 20, 20, 40, 30, 30, 30, 30, 45, 45]
    
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('Cell', fontName='Helvetica', fontSize=7, leading=8, textColor=colors.HexColor('#1A2640'))
    cell_style_bold = ParagraphStyle('CellBold', fontName='Helvetica-Bold', fontSize=7, leading=8, textColor=colors.HexColor('#1A2640'))
    cell_style_center = ParagraphStyle('CellCenter', fontName='Helvetica', fontSize=7, leading=8, textColor=colors.HexColor('#1A2640'), alignment=1)
    cell_style_center_bold = ParagraphStyle('CellCenterBold', fontName='Helvetica-Bold', fontSize=7, leading=8, textColor=colors.HexColor('#1A2640'), alignment=1)
    
    grid = [[Paragraph("", cell_style) for _ in range(16)] for _ in range(39)]
    
    # Jindal Steel Logo at the top
    logo_path = os.path.join(settings.BASE_DIR, 'portal', 'static', 'portal', 'img', 'jindal_logo_dark.png')
    if os.path.exists(logo_path):
        try:
            logo_img = ReportLabImage(logo_path, width=110, height=30)
            grid[0][1] = logo_img
        except Exception as e:
            grid[0][1] = Paragraph(f"[Logo Error: {e}]", cell_style)
            
    # 1. KAIZEN NO.
    grid[2][1] = Paragraph(f"<b>KAIZEN NO.</b><br/>{kaizen.kaizen_no}", cell_style_center)
    
    # 2. Title
    grid[2][2] = Paragraph("<b>KAIZEN IDEA SHEET</b>", ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=12, leading=14, alignment=1, textColor=colors.HexColor('#003478')))
    
    # 3. Activity list
    grid[2][5] = Paragraph("<b>ACTIVITY</b>", cell_style_center_bold)
    activities_list = ['KK', 'JH', 'QM', 'PM', 'SHE', 'OTPM', 'DM', 'ET']
    for idx, act in enumerate(activities_list):
        checked = "✓" if act in kaizen.activities else ""
        grid[2][6 + idx] = Paragraph(f"<b>{act}</b><br/>{checked}", cell_style_center)
        
    # 4. Loss Name
    grid[3][5] = Paragraph("<b>Loss Name/JH STEP (if Any)</b>", cell_style_bold)
    grid[3][6] = Paragraph(kaizen.loss_name, cell_style)
    
    # 5. Result Area
    grid[4][5] = Paragraph("<b>RESULT AREA</b>", cell_style_bold)
    result_areas_list = ['S', 'Q', 'P', 'C', 'D', 'M']
    res_indices = {'S': 6, 'Q': 8, 'P': 9, 'C': 10, 'D': 11, 'M': 12}
    for res in result_areas_list:
        checked = "✓" if res in kaizen.result_areas else ""
        col_idx = res_indices[res]
        grid[4][col_idx] = Paragraph(f"<b>{res}</b><br/>{checked}", cell_style_center)
        
    # 6. Dept, Area, Circle Name
    grid[5][1] = Paragraph(f"<b>DEPARTMENT :</b> {kaizen.department.name}", cell_style)
    grid[5][5] = Paragraph(f"<b>AREA/EQUIPMENT :</b> {kaizen.area_equipment}", cell_style)
    grid[5][9] = Paragraph(f"<b>TPM CIRCLE NAME:</b> {kaizen.circle_name}", cell_style)
    
    # 7. Theme & Idea
    grid[7][1] = Paragraph("<b>KAIZEN THEME</b>", cell_style_bold)
    grid[9][1] = Paragraph(kaizen.theme, cell_style)
    
    grid[7][5] = Paragraph("<b>IDEA</b>", cell_style_bold)
    grid[9][5] = Paragraph(kaizen.idea, cell_style)
    
    # 8. Target parameters
    grid[7][9] = Paragraph("<b>BENCHMARK</b>", cell_style_bold)
    grid[7][11] = Paragraph(kaizen.benchmark, cell_style)
    
    grid[8][9] = Paragraph("<b>TARGET</b>", cell_style_bold)
    grid[8][11] = Paragraph(kaizen.target, cell_style)
    
    grid[9][9] = Paragraph("<b>KAIZEN START</b>", cell_style_bold)
    grid[9][11] = Paragraph(kaizen.start_date, cell_style)
    
    grid[10][9] = Paragraph("<b>KAIZEN FINISH</b>", cell_style_bold)
    grid[10][11] = Paragraph(kaizen.finish_date, cell_style)
    
    # Team members
    grid[11][9] = Paragraph("<b>TEAM MEMBERS</b>", cell_style_center_bold)
    
    grid[12][9] = Paragraph("<b>LEADER</b>", cell_style_bold)
    grid[12][11] = Paragraph(kaizen.team_leader, cell_style)
    
    members = kaizen.team_members
    for idx in range(4):
        grid[13 + idx][9] = Paragraph("<b>MEMBER</b>", cell_style_bold)
        val = members[idx] if idx < len(members) else ""
        grid[13 + idx][11] = Paragraph(val, cell_style)
        
    # Before/After headers
    grid[17][1] = Paragraph("<b>BEFORE</b>", cell_style_center_bold)
    grid[17][5] = Paragraph("<b>AFTER</b>", cell_style_center_bold)
    
    # Before/After images
    if kaizen.before_image:
        try:
            grid[18][1] = ReportLabImage(kaizen.before_image.path, width=140, height=120)
        except:
            grid[18][1] = Paragraph("[Image Error]", cell_style_center)
    else:
        grid[18][1] = Paragraph("(ILLUSTRATE WITH PHOTO/SKETCH)", cell_style_center)
        
    if kaizen.after_image:
        try:
            grid[18][5] = ReportLabImage(kaizen.after_image.path, width=140, height=120)
        except:
            grid[18][5] = Paragraph("[Image Error]", cell_style_center)
    else:
        grid[18][5] = Paragraph("(ILLUSTRATE WITH PHOTO/SKETCH)", cell_style_center)
        
    grid[28][1] = Paragraph("(ILLUSTRATE WITH PHOTO/SKETCH)", cell_style_center)
    grid[28][5] = Paragraph("(ILLUSTRATE WITH PHOTO/SKETCH)", cell_style_center)
    
    # Benefits
    grid[17][9] = Paragraph("<b>BENEFITS</b>", cell_style_center_bold)
    grid[18][9] = Paragraph("<b>TANGIBLE</b>", cell_style_bold)
    
    tangible = kaizen.tangible_benefits
    for idx in range(5):
        val = tangible[idx] if idx < len(tangible) else ""
        grid[19 + idx][9] = Paragraph(f"{idx+1}.0", cell_style_bold)
        grid[19 + idx][10] = Paragraph(val, cell_style)
        
    grid[24][9] = Paragraph("<b>INTANGIBLE</b>", cell_style_bold)
    intangible = kaizen.intangible_benefits
    for idx in range(5):
        val = intangible[idx] if idx < len(intangible) else ""
        grid[25 + idx][9] = Paragraph(f"{idx+1}.0", cell_style_bold)
        grid[25 + idx][10] = Paragraph(val, cell_style)
        
    # Analysis & Result Label
    grid[30][1] = Paragraph("<b>ANAYLSIS - </b>", cell_style_bold)
    grid[31][1] = Paragraph(kaizen.analysis, cell_style)
    
    grid[30][5] = Paragraph("<b>RESULT</b>", cell_style_bold)
    if kaizen.result_image:
        try:
            grid[31][5] = ReportLabImage(kaizen.result_image.path, width=140, height=75)
        except:
            grid[31][5] = Paragraph(kaizen.result_text, cell_style)
    else:
        grid[31][5] = Paragraph(kaizen.result_text, cell_style)
        
    grid[37][5] = Paragraph("(ILLUSTRATE WITH BAR/LINE GRAPH)", cell_style_center)
    
    # Horizontal Deployment
    grid[30][9] = Paragraph("<b>SCOPE & PLAN OF HORIZONTAL DEPLOYMENT</b>", cell_style_center_bold)
    grid[32][9] = Paragraph("<b>SL. NO.</b>", cell_style_center_bold)
    grid[32][10] = Paragraph("<b>AREA/ EQUIP NO.</b>", cell_style_center_bold)
    grid[32][11] = Paragraph("<b>TARGET DATE</b>", cell_style_center_bold)
    grid[32][13] = Paragraph("<b>RESPONSIBILITY</b>", cell_style_center_bold)
    grid[32][15] = Paragraph("<b>STATUS</b>", cell_style_center_bold)
    
    deployment = kaizen.horizontal_deployment
    for idx in range(3):
        row_idx = 35 + idx
        if idx < len(deployment):
            item = deployment[idx]
            grid[row_idx][9] = Paragraph(item.get('sl_no', str(idx+1)), cell_style_center)
            grid[row_idx][10] = Paragraph(item.get('area_equip', ''), cell_style)
            grid[row_idx][11] = Paragraph(item.get('target_date', ''), cell_style_center)
            grid[row_idx][13] = Paragraph(item.get('responsibility', ''), cell_style)
            grid[row_idx][15] = Paragraph(item.get('status', ''), cell_style_center)
            
    # Compile TableStyle spans
    table_styles = [
        ('GRID', (1,0), (-1,37), 0.5, colors.HexColor('#000000')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('TOPPADDING', (0,0), (-1,-1), 1),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        
        ('SPAN', (1,0), (15,1)),  # Jindal Logo header span
        ('SPAN', (1,2), (1,4)),   # Kaizen No
        ('SPAN', (2,2), (4,4)),   # Title
        
        ('SPAN', (7,3), (14,3)),  # Loss Name merge
        
        ('SPAN', (1,5), (4,6)),   # Dept
        ('SPAN', (5,5), (8,6)),   # Area
        ('SPAN', (9,5), (13,6)),  # Circle
        
        ('SPAN', (1,7), (4,8)),   # Theme title
        ('SPAN', (1,9), (4,16)),  # Theme body
        
        ('SPAN', (5,7), (8,8)),   # Idea title
        ('SPAN', (5,9), (8,16)),  # Idea body
        
        ('SPAN', (9,7), (10,7)),  # Bench label
        ('SPAN', (11,7), (15,7)), # Bench val
        
        ('SPAN', (9,8), (10,8)),  # Target label
        ('SPAN', (11,8), (15,8)), # Target val
        
        ('SPAN', (9,9), (10,9)),  # Start label
        ('SPAN', (11,9), (15,9)), # Start val
        
        ('SPAN', (9,10), (10,10)),# Finish label
        ('SPAN', (11,10), (15,10)),# Finish val
        
        ('SPAN', (9,11), (15,11)),# Team member header
        
        ('SPAN', (9,12), (10,12)),# Leader label
        ('SPAN', (11,12), (15,12)),# Leader val
        
        ('SPAN', (9,13), (10,13)),# Member 1 label
        ('SPAN', (11,13), (15,13)),# Member 1 val
        
        ('SPAN', (9,14), (10,14)),# Member 2 label
        ('SPAN', (11,14), (15,14)),# Member 2 val
        
        ('SPAN', (9,15), (10,15)),# Member 3 label
        ('SPAN', (11,15), (15,15)),# Member 3 val
        
        ('SPAN', (9,16), (10,16)),# Member 4 label
        ('SPAN', (11,16), (15,16)),# Member 4 val
        
        ('SPAN', (1,17), (4,17)), # Before title
        ('SPAN', (1,18), (4,27)), # Before image region
        ('SPAN', (1,28), (4,29)), # Before subtext
        
        ('SPAN', (5,17), (8,17)), # After title
        ('SPAN', (5,18), (8,27)), # After image region
        ('SPAN', (5,28), (8,29)), # After subtext
        
        ('SPAN', (9,17), (15,17)),# Benefits title
        ('SPAN', (9,18), (15,18)),# Tangible header
        
        ('SPAN', (10,19), (15,19)),# Tangible 1 val
        ('SPAN', (10,20), (15,20)),# Tangible 2 val
        ('SPAN', (10,21), (15,21)),# Tangible 3 val
        ('SPAN', (10,22), (15,22)),# Tangible 4 val
        ('SPAN', (10,23), (15,23)),# Tangible 5 val
        
        ('SPAN', (9,24), (15,24)),# Intangible header
        
        ('SPAN', (10,25), (15,25)),# Intangible 1 val
        ('SPAN', (10,26), (15,26)),# Intangible 2 val
        ('SPAN', (10,27), (15,27)),# Intangible 3 val
        ('SPAN', (10,28), (15,28)),# Intangible 4 val
        ('SPAN', (10,29), (15,29)),# Intangible 5 val
        
        ('SPAN', (1,30), (4,30)), # Analysis header
        ('SPAN', (1,31), (4,36)), # Analysis body
        
        ('SPAN', (5,30), (8,30)), # Result header
        ('SPAN', (5,31), (8,36)), # Result body
        ('SPAN', (5,37), (8,37)), # Result subtext
        
        ('SPAN', (9,30), (15,31)),# Deployment header
        
        ('SPAN', (11,32), (12,34)),# Deployment col target date span
        ('SPAN', (13,32), (14,34)),# Deployment col responsibility span
        ('SPAN', (9,32), (9,34)),  # SL no span
        ('SPAN', (10,32), (10,34)),# Area span
        ('SPAN', (15,32), (15,34)),# Status span
        
        ('SPAN', (11,35), (12,35)),# Deployment row 1 date
        ('SPAN', (13,35), (14,35)),# Deployment row 1 resp
        
        ('SPAN', (11,36), (12,36)),# Deployment row 2 date
        ('SPAN', (13,36), (14,36)),# Deployment row 2 resp
        
        ('SPAN', (11,37), (12,37)),# Deployment row 3 date
        ('SPAN', (13,37), (14,37)),# Deployment row 3 resp
    ]
    
    pdf_table = Table(grid, colWidths=col_widths, rowHeights=19)
    pdf_table.setStyle(TableStyle(table_styles))
    
    doc.build([pdf_table])
    buffer.seek(0)
    return buffer.getvalue()
