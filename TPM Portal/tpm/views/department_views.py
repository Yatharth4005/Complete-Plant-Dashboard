import json
import datetime
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.contrib import messages
from tpm.models import Department, PillarEntry, KPIValue, WorkstationValue, Workstation, CustomKPIDefinition
from tpm.utils.decorators import dept_access_required
from tpm.utils.calculations import compute_achievement, parse_period, get_date_range_q
from portal.utils.access import user_can_edit_module
from django.views.decorators.http import require_POST

def get_months_list():
    return [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]

def get_months_in_range(from_month, from_year, to_month, to_year):
    months_labels_short = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    result = []
    m = from_month
    y = from_year
    while y < to_year or (y == to_year and m <= to_month):
        result.append({
            'month': m,
            'year': y,
            'label': f"{months_labels_short[m-1]} '{str(y)[-2:]}"
        })
        m += 1
        if m > 12:
            m = 1
            y += 1
    return result

@dept_access_required
def dept_overview(request, dept_id):
    if int(dept_id) == 0:
        return redirect('tpm_dashboard')
        
    dept = get_object_or_404(Department, id=dept_id)
    period = parse_period(request)
    filter_type = period['filter_type']
    selected_month = period['month']
    selected_year = period['year']
    from_month = period['from_month']
    from_year = period['from_year']
    to_month = period['to_month']
    to_year = period['to_year']
    period_label = period['label']
    
    today = datetime.date.today()
    
    pillars_meta = [
        {'id': 'KK', 'label': 'KK (Kobetsu Kaizen)', 'icon': '🎯'},
        {'id': 'JH', 'label': 'JH (Jishu Hozen)', 'icon': '⚙️'},
        {'id': 'PM', 'label': 'PM (Planned Maintenance)', 'icon': '🔧'},
        {'id': 'QM', 'label': 'QM (Quality Maintenance)', 'icon': '💎'},
        {'id': 'ET', 'label': 'ET (Education & Training)', 'icon': '📚'},
        {'id': 'DM', 'label': 'DM (Initial Flow/Design)', 'icon': '📐'},
        {'id': 'SHE', 'label': 'SHE (Safety & Environment)', 'icon': '🛡️'},
        {'id': 'OTPM', 'label': 'OTPM (Office TPM)', 'icon': '🏢'},
    ]
    
    pillar_cards = []
    
    # Standard 8 Pillars scores calculation
    for pm in pillars_meta:
        # Check status
        entries = PillarEntry.objects.filter(
            get_date_range_q(from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
            department=dept,
            pillar=pm['id']
        )
        
        status_label = 'Pending'
        status_class = 'badge-red'
        if entries.exists():
            all_locked = all(e.is_locked() for e in entries)
            status_label = 'Locked' if all_locked else 'Draft'
            status_class = 'badge-green' if all_locked else 'badge-amber'
            
        kpis = KPIValue.objects.filter(
            get_date_range_q(prefix='pillar_entry__', from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
            pillar_entry__department=dept,
            pillar_entry__pillar=pm['id']
        )
        ach_list = []
        for k in kpis:
            if k.actual is not None and k.target is not None:
                ach_list.append(compute_achievement(k.actual, k.target, k.kpi_name))
        
        avg_ach = sum(ach_list) / len(ach_list) if ach_list else 0.0
        
        pillar_cards.append({
            'id': pm['id'],
            'label': pm['label'],
            'icon': pm['icon'],
            'achievement': round(avg_ach, 1),
            'status': status_label,
            'status_class': status_class,
        })
        
    # Workstation KPI (9th card)
    ws_vals = WorkstationValue.objects.filter(
        get_date_range_q(from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
        workstation_kpi__workstation__department=dept
    )
    ws_ach_list = []
    for val in ws_vals:
        if val.actual is not None and val.workstation_kpi.commitment is not None:
            indicator = val.workstation_kpi.goodness_indicator
            target = val.workstation_kpi.commitment
            if indicator == 'LOWER':
                ach = min(100.0, (target / val.actual) * 100.0) if val.actual != 0 else 100.0
            else:
                ach = (val.actual / target) * 100.0 if target != 0 else 100.0
            ws_ach_list.append(ach)
            
    ws_avg_ach = sum(ws_ach_list) / len(ws_ach_list) if ws_ach_list else 0.0
    has_workstations = Workstation.objects.filter(department=dept).exists()
    ws_status = 'N/A' if not has_workstations else ('Locked' if ws_vals.exists() else 'Pending')
    ws_class = 'badge-muted' if ws_status == 'N/A' else ('badge-green' if ws_status == 'Locked' else 'badge-red')
    
    pillar_cards.append({
        'id': 'ws-kpi',
        'label': 'Workstation KPI',
        'icon': '🏭',
        'achievement': round(ws_avg_ach, 1),
        'status': ws_status,
        'status_class': ws_class,
    })

    # PRODUCTION monthly actual vs target line chart trend
    PRODUCTION_trend_actuals = []
    PRODUCTION_trend_targets = []
    months_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    if filter_type == 'range':
        range_months = get_months_in_range(from_month, from_year, to_month, to_year)
    else:
        range_months = [{
            'month': m,
            'year': selected_year,
            'label': months_labels[m-1]
        } for m in range(1, selected_month + 1)]

    chart_labels = []
    for rm in range_months:
        val = KPIValue.objects.filter(
            pillar_entry__department=dept,
            pillar_entry__pillar='KK',
            pillar_entry__month=rm['month'],
            pillar_entry__year=rm['year'],
            sl_no='1'
        ).first()
        PRODUCTION_trend_actuals.append(val.actual if val else None)
        PRODUCTION_trend_targets.append(val.target if val and val.target else 90.0)
        chart_labels.append(rm['label'])

    # 9-axis radar data for this department specifically
    radar_labels = ['KK', 'JH', 'PM', 'QM', 'ET', 'DM', 'SHE', 'OTPM', 'WS KPI']
    radar_data = [card['achievement'] for card in pillar_cards]

    # Recent submissions table (last 6 months relative to range end date)
    recent_months = []
    for i in range(5, -1, -1):
        target_month = to_month - i
        target_year = to_year
        if target_month <= 0:
            target_month += 12
            target_year -= 1
        recent_months.append({
            'month': target_month,
            'year': target_year,
            'label': f"{months_labels[target_month-1]} '{str(target_year)[-2:]}"
        })

    submission_rows = []
    for pm in pillars_meta:
        row = {'pillar_label': pm['label'], 'pillar_id': pm['id'], 'months': []}
        for rm in recent_months:
            entry = PillarEntry.objects.filter(
                department=dept,
                pillar=pm['id'],
                month=rm['month'],
                year=rm['year']
            ).first()
            status_text = 'Pending'
            status_cls = 'text-danger'
            if entry:
                status_text = 'Locked' if entry.is_locked() else 'Draft'
                status_cls = 'text-success font-semibold' if entry.is_locked() else 'text-warning'
            row['months'].append({'status': status_text, 'class': status_cls})
        submission_rows.append(row)

    # Workstation KPI row for recent months table
    ws_row = {'pillar_label': 'Workstation KPI', 'pillar_id': 'ws-kpi', 'months': []}
    for rm in recent_months:
        vals = WorkstationValue.objects.filter(
            workstation_kpi__workstation__department=dept,
            month=rm['month'],
            year=rm['year']
        )
        status_text = 'Locked' if vals.exists() else 'Pending'
        status_cls = 'text-success font-semibold' if vals.exists() else 'text-danger'
        ws_row['months'].append({'status': status_text, 'class': status_cls})
    submission_rows.append(ws_row)

    can_edit = user_can_edit_module(request.user, dept, 'TPM')

    if filter_type == 'range':
        query_params = f"filter_type=range&from_month={from_month}&from_year={from_year}&to_month={to_month}&to_year={to_year}"
    else:
        query_params = f"filter_type=single&month={selected_month}&year={selected_year}"

    context = {
        'dept': dept,
        'filter_type': filter_type,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month': selected_month,
        'year': selected_year,
        'from_month': from_month,
        'from_year': from_year,
        'to_month': to_month,
        'to_year': to_year,
        'period_label': period_label,
        'query_params': query_params,
        'months': get_months_list(),
        'years': range(2025, today.year + 2),
        'pillar_cards': pillar_cards,
        'recent_months': recent_months,
        'submission_rows': submission_rows,
        'PRODUCTION_trend_actuals_json': json.dumps(PRODUCTION_trend_actuals),
        'PRODUCTION_trend_targets_json': json.dumps(PRODUCTION_trend_targets),
        'radar_labels_json': json.dumps(radar_labels),
        'radar_data_json': json.dumps(radar_data),
        'months_labels_json': json.dumps(chart_labels),
        'can_edit': can_edit,
    }
    return render(request, 'department/overview.html', context)


def parse_tpm_filename(filename):
    import re
    from tpm.models import Department
    
    name_part = filename.rsplit('.', 1)[0].upper()
    
    # 1. Match Month
    MONTHS_MAP = {
        'JANUARY': 1, 'FEB': 2, 'FEBRUARY': 2, 'MAR': 3, 'MARCH': 3,
        'APR': 4, 'APRIL': 4, 'MAY': 5, 'JUN': 6, 'JUNE': 6,
        'JUL': 7, 'JULY': 7, 'AUG': 8, 'AUGUST': 8, 'SEP': 9,
        'SEPT': 9, 'SEPTEMBER': 9, 'OCT': 10, 'OCTOBER': 10,
        'NOV': 11, 'NOVEMBER': 11, 'DEC': 12, 'DECEMBER': 12,
        'JAN': 1
    }
    
    month = None
    sorted_months = sorted(MONTHS_MAP.keys(), key=len, reverse=True)
    for m_str in sorted_months:
        if m_str in name_part:
            month = MONTHS_MAP[m_str]
            break
            
    # 2. Match Year
    year = None
    year_match = re.search(r'(?<!\d)(20\d{2})(?!\d)', name_part)
    if year_match:
        year = int(year_match.group(1))
    else:
        year_match = re.search(r'(?<!\d)(\d{2})(?!\d)', name_part)
        if year_match:
            year = 2000 + int(year_match.group(1))
            
    # 3. Match Department Code
    name_no_tpm = re.sub(r'TPM', '', name_part, flags=re.IGNORECASE)
    name_clean = re.sub(r'[^A-Z0-9]', '', name_no_tpm.upper())
    all_depts = Department.objects.all()
    sorted_depts = sorted(all_depts, key=lambda d: len(d.code), reverse=True)
    
    target_dept = None
    for d in sorted_depts:
        code_clean = re.sub(r'[^A-Z0-9]', '', d.code.upper())
        if code_clean and code_clean in name_clean:
            target_dept = d
            break
            
    return target_dept, month, year


def import_sheet_data(ws, dept, pillar_id, month, year):
    import re
    from tpm.models import PillarEntry, KPIValue, CustomKPIDefinition
    from tpm.utils.kpi_definitions import KPI_DEFINITIONS
    from tpm.utils.calculations import compute_PRODUCTION
    
    # 1. Find value column index
    value_col_idx = 3 # Default to Column D
    for c in range(3, ws.max_column):
        val = ws.cell(row=1, column=c+1).value
        if val is not None and str(val).strip():
            value_col_idx = c
            break
            
    # 2. Parse rows
    current_sl_no = None
    kpi_data = {}
    
    for r in range(2, ws.max_row + 1):
        sl_no_val = ws.cell(row=r, column=1).value
        
        # Skip header rows
        if sl_no_val is not None:
            sl_no_clean = str(sl_no_val).strip().upper().replace(' ', '').replace('.', '')
            if sl_no_clean in ('SNO', 'SLNO', 'SERIALNO', 'SERIALNUMBER'):
                continue
                
        particulars_val = ws.cell(row=r, column=3).value
        value_val = ws.cell(row=r, column=value_col_idx+1).value
        kpi_name_raw = ws.cell(row=r, column=2).value
        
        if sl_no_val is not None and str(sl_no_val).strip():
            current_sl_no = str(sl_no_val).strip()
            
        if not current_sl_no:
            continue
            
        if current_sl_no not in kpi_data:
            kpi_data[current_sl_no] = {
                'benchmark': None,
                'target': None,
                'actual': None,
                'availability': None,
                'performance': None,
                'quality': None,
                'remarks': '',
                'name_raw': kpi_name_raw or ''
            }
        else:
            if kpi_name_raw and not kpi_data[current_sl_no]['name_raw']:
                kpi_data[current_sl_no]['name_raw'] = kpi_name_raw
                
        if particulars_val is not None:
            part_lower = str(particulars_val).strip().lower()
            if 'benchmark' in part_lower:
                kpi_data[current_sl_no]['benchmark'] = value_val
            elif 'target' in part_lower:
                kpi_data[current_sl_no]['target'] = value_val
            elif 'actual' in part_lower:
                kpi_data[current_sl_no]['actual'] = value_val
            elif 'availability' in part_lower:
                kpi_data[current_sl_no]['availability'] = value_val
            elif 'performance' in part_lower or 'performace' in part_lower:
                kpi_data[current_sl_no]['performance'] = value_val
            elif 'quality' in part_lower:
                kpi_data[current_sl_no]['quality'] = value_val
            elif 'remark' in part_lower:
                kpi_data[current_sl_no]['remarks'] = str(value_val) if value_val is not None else ''

    if not kpi_data:
        return False, "No data found in sheet"

    # Get or create PillarEntry
    entry, created = PillarEntry.objects.get_or_create(
        department=dept, pillar=pillar_id, month=month, year=year
    )
    
    if entry.is_locked():
        return False, "Locked"
        
    definitions = list(KPI_DEFINITIONS.get(pillar_id, []))
    custom_defs = CustomKPIDefinition.objects.filter(department=dept, pillar=pillar_id)
    for cd in custom_defs:
        definitions.append({
            'sl_no': cd.sl_no,
            'name': cd.name,
            'uom': cd.uom,
            'benchmark': cd.benchmark,
            'target': cd.target,
            'is_custom': True,
        })
        
    def parse_name_uom(raw_name):
        if not raw_name:
            return "", ""
        raw_name = str(raw_name).strip()
        match = re.search(r'\(([^)]+)\)$', raw_name)
        if match:
            uom = match.group(1).strip()
            name = raw_name[:match.start()].strip().rstrip('-').strip()
            return name, uom
        return raw_name, ""
        
    def safe_float(val):
        if val is None:
            return None
        try:
            cleaned = str(val).replace('%', '').replace(',', '').strip()
            if cleaned == '' or cleaned.lower() in ('none', '-', 'n/a', 'null'):
                return None
            return float(cleaned)
        except ValueError:
            return None

    # Save rows
    for sl_no, info in kpi_data.items():
        kpi_meta = next((d for d in definitions if d['sl_no'] == sl_no), None)
        
        if not kpi_meta:
            raw_name = info['name_raw']
            if not raw_name:
                raw_name = f"Custom KPI {sl_no}"
            parsed_name, parsed_uom = parse_name_uom(raw_name)
            
            cd = CustomKPIDefinition.objects.create(
                department=dept,
                pillar=pillar_id,
                sl_no=sl_no,
                name=parsed_name or raw_name,
                uom=parsed_uom,
                benchmark=safe_float(info['benchmark']),
                target=safe_float(info['target'])
            )
            kpi_meta = {
                'sl_no': cd.sl_no,
                'name': cd.name,
                'uom': cd.uom,
                'benchmark': cd.benchmark,
                'target': cd.target,
                'is_custom': True,
            }
            definitions.append(kpi_meta)
            
        db_val, created_val = KPIValue.objects.get_or_create(
            pillar_entry=entry, sl_no=sl_no,
            defaults={
                'kpi_name': kpi_meta['name'],
                'uom': kpi_meta['uom'],
                'benchmark': kpi_meta['benchmark'],
                'target': kpi_meta['target'],
            }
        )
        
        # Update fields if present
        if info['benchmark'] is not None:
            db_val.benchmark = safe_float(info['benchmark'])
        if info['target'] is not None:
            db_val.target = safe_float(info['target'])
            
        is_PRODUCTION_row = kpi_meta.get('is_PRODUCTION_row', False) or 'PRODUCTION' in kpi_meta['name'].upper() or 'OEE' in kpi_meta['name'].upper()
        if is_PRODUCTION_row:
            db_val.availability = safe_float(info['availability'])
            db_val.performance = safe_float(info['performance'])
            db_val.quality = safe_float(info['quality'])
            if db_val.availability is not None and db_val.performance is not None and db_val.quality is not None:
                db_val.actual = round(compute_PRODUCTION(db_val.availability, db_val.performance, db_val.quality), 2)
            else:
                db_val.actual = safe_float(info['actual'])
        else:
            db_val.actual = safe_float(info['actual'])
            
        db_val.remarks = info['remarks'] or ''
        db_val.save()
        
    return True, "Success"


@dept_access_required
@require_POST
def upload_tpm_excel(request, dept_id):
    import openpyxl
    from django.views.decorators.http import require_POST
    from portal.utils.access import user_can_edit_module
    
    current_dept = get_object_or_404(Department, id=dept_id)
    if not user_can_edit_module(request.user, current_dept, 'TPM'):
        messages.error(request, "You do not have permission to edit this department's TPM.")
        return redirect('dept_overview', dept_id=dept_id)
        
    excel_file = request.FILES.get('file')
    if not excel_file:
        messages.error(request, "No Excel file was uploaded.")
        return redirect('dept_overview', dept_id=dept_id)
        
    filename = excel_file.name
    target_dept, month, year = parse_tpm_filename(filename)
    
    if not target_dept:
        messages.error(request, f"Could not determine department from filename: {filename}")
        return redirect('dept_overview', dept_id=dept_id)
        
    if target_dept != current_dept:
        messages.error(request, f"The uploaded file is for department '{target_dept.code}', but you are on '{current_dept.code}' page.")
        return redirect('dept_overview', dept_id=dept_id)
        
    if not month or not year:
        messages.error(request, f"Could not determine month and year from filename: {filename}")
        return redirect('dept_overview', dept_id=dept_id)
        
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        messages.error(request, f"Failed to read Excel file: {str(e)}")
        return redirect('dept_overview', dept_id=dept_id)
        
    SHEET_TO_PILLAR = {
        'KK': 'KK', 'KOBETSU': 'KK', 'KOBETSU KAIZEN': 'KK',
        'JH': 'JH', 'JISHU': 'JH', 'JISHU HOZEN': 'JH',
        'PM': 'PM', 'PLANNED': 'PM', 'PLANNED MAINTENANCE': 'PM',
        'QM': 'QM', 'QUALITY': 'QM', 'QUALITY MAINTENANCE': 'QM',
        'ET': 'ET', 'E & T': 'ET', 'EANDT': 'ET', 'EDUCATION': 'ET', 'EDUCATION & TRAINING': 'ET',
        'DM': 'DM', 'DESIGN': 'DM', 'DESIGN & MANAGEMENT': 'DM',
        'SHE': 'SHE', 'SAFETY': 'SHE', 'SAFETY & ENVIRONMENT': 'SHE', 'SAFETY HEALTH ENVIRONMENT': 'SHE',
        'OTPM': 'OTPM', 'OFFICE': 'OTPM', 'OFFICE TPM': 'OTPM'
    }
    
    updated_pillars = []
    skipped_pillars = []
    errors = []
    
    for sheet_name in wb.sheetnames:
        norm_name = sheet_name.strip().upper()
        pillar_id = SHEET_TO_PILLAR.get(norm_name)
        if not pillar_id:
            continue
            
        ws = wb[sheet_name]
        try:
            success, msg = import_sheet_data(ws, target_dept, pillar_id, month, year)
            if success:
                updated_pillars.append(pillar_id)
            else:
                skipped_pillars.append(f"{pillar_id} ({msg})")
        except Exception as e:
            errors.append(f"{pillar_id}: {str(e)}")
            
    if errors:
        messages.error(request, f"Import errors occurred: {'; '.join(errors)}")
    if skipped_pillars:
        messages.warning(request, f"Pillars skipped: {', '.join(skipped_pillars)}")
        
    if updated_pillars:
        messages.success(request, f"Successfully imported TPM data for: {', '.join(updated_pillars)}")
        
    return redirect(f"/tpm/department/{dept_id}/?filter_type=single&month={month}&year={year}")
