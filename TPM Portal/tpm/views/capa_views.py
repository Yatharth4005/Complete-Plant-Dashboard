import io
import json
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.conf import settings

from tpm.models import Department, CAPAReport
from tpm.utils.decorators import admin_required

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

@login_required
def capa_list_partial(request):
    reports = CAPAReport.objects.all().order_by('-created_at')
    context = {
        'reports': reports,
    }
    return render(request, 'partials/_capa_list.html', context)

@login_required
def capa_edit_partial(request, capa_id=None):
    report = None
    if capa_id:
        report = get_object_or_404(CAPAReport, id=capa_id)
        
    depts = Department.objects.all().order_by('name')
    
    # Prefills
    prefills = {
        'capa_no': f"CAPA-{(report.id if report else CAPAReport.objects.count() + 1):03d}",
        'responsible_team': [
            {'name': '', 'members': '', 'role': '', 'contact': ''},
            {'name': '', 'members': '', 'role': '', 'contact': ''},
            {'name': '', 'members': '', 'role': '', 'contact': ''},
        ],
        'corrective_actions': [
            {'action': '', 'responsibility': '', 'target_date': '', 'impl_date': ''},
            {'action': '', 'responsibility': '', 'target_date': '', 'impl_date': ''},
        ],
        'preventive_actions': [
            {'action': '', 'responsibility': '', 'target_date': '', 'impl_date': ''},
            {'action': '', 'responsibility': '', 'target_date': '', 'impl_date': ''},
        ]
    }
    
    context = {
        'report': report,
        'depts': depts,
        'prefills': prefills,
    }
    return render(request, 'partials/_capa_form.html', context)

@login_required
@require_POST
def capa_save(request, capa_id=None):
    if capa_id:
        report = get_object_or_404(CAPAReport, id=capa_id)
    else:
        report = CAPAReport(created_by=request.user)
        
    dept_id = request.POST.get('department_id')
    report.department = get_object_or_404(Department, id=dept_id)
    
    report.area_section = request.POST.get('area_section', '').strip()
    report.date_incident = request.POST.get('date_incident', '').strip()
    report.capa_no = request.POST.get('capa_no', '').strip()
    
    report.problem_what = request.POST.get('problem_what', '').strip()
    report.problem_where = request.POST.get('problem_where', '').strip()
    report.problem_when = request.POST.get('problem_when', '').strip()
    report.problem_extent = request.POST.get('problem_extent', '').strip()
    
    report.breakdown_applicable = request.POST.get('breakdown_applicable', '').strip()
    report.breakdown_hrs = request.POST.get('breakdown_hrs', '').strip()
    report.breakdown_from = request.POST.get('breakdown_from', '').strip()
    report.breakdown_to = request.POST.get('breakdown_to', '').strip()
    
    report.immediate_action = request.POST.get('immediate_action', '').strip()
    report.action_timeframe = request.POST.get('action_timeframe', '').strip()
    report.action_responsibility = request.POST.get('action_responsibility', '').strip()
    
    report.why_1 = request.POST.get('why_1', '').strip()
    report.why_2 = request.POST.get('why_2', '').strip()
    report.why_3 = request.POST.get('why_3', '').strip()
    report.why_4 = request.POST.get('why_4', '').strip()
    report.why_5 = request.POST.get('why_5', '').strip()
    
    report.conclusion = request.POST.get('conclusion', '').strip()
    report.modified_documents_other = request.POST.get('modified_documents_other', '').strip()
    report.training_details = request.POST.get('training_details', '').strip()
    report.date_implementation = request.POST.get('date_implementation', '').strip()
    report.effectiveness_evaluation = request.POST.get('effectiveness_evaluation', '').strip()
    
    report.prepared_by = request.POST.get('prepared_by', '').strip()
    report.reviewed_by = request.POST.get('reviewed_by', '').strip()
    report.approved_by = request.POST.get('approved_by', '').strip()
    
    # Checkboxes / lists
    report.five_m_applicable = request.POST.getlist('five_m_applicable')
    report.modified_documents = request.POST.getlist('modified_documents')
    
    # Dynamic Lists: Responsible Team
    t_names = request.POST.getlist('t_name')
    t_members = request.POST.getlist('t_members')
    t_roles = request.POST.getlist('t_role')
    t_contacts = request.POST.getlist('t_contact')
    
    team_data = []
    for idx in range(len(t_names)):
        if t_names[idx].strip() or t_members[idx].strip():
            team_data.append({
                'name': t_names[idx].strip(),
                'members': t_members[idx].strip(),
                'role': t_roles[idx].strip(),
                'contact': t_contacts[idx].strip(),
            })
    report.responsible_team = team_data
    
    # Dynamic Lists: Corrective Actions
    c_actions = request.POST.getlist('c_action')
    c_resps = request.POST.getlist('c_responsibility')
    c_targets = request.POST.getlist('c_target_date')
    c_impls = request.POST.getlist('c_impl_date')
    
    corr_data = []
    for idx in range(len(c_actions)):
        if c_actions[idx].strip():
            corr_data.append({
                'action': c_actions[idx].strip(),
                'responsibility': c_resps[idx].strip(),
                'target_date': c_targets[idx].strip(),
                'impl_date': c_impls[idx].strip(),
            })
    report.corrective_actions = corr_data
    
    # Dynamic Lists: Preventive Actions
    p_actions = request.POST.getlist('p_action')
    p_resps = request.POST.getlist('p_responsibility')
    p_targets = request.POST.getlist('p_target_date')
    p_impls = request.POST.getlist('p_impl_date')
    
    prev_data = []
    for idx in range(len(p_actions)):
        if p_actions[idx].strip():
            prev_data.append({
                'action': p_actions[idx].strip(),
                'responsibility': p_resps[idx].strip(),
                'target_date': p_targets[idx].strip(),
                'impl_date': p_impls[idx].strip(),
            })
    report.preventive_actions = prev_data
    
    # Detailed Plan
    report.detailed_plan = request.POST.get('detailed_plan', '').strip()
    
    report.save()
    
    return capa_list_partial(request)

@login_required
@require_POST
def capa_delete(request, capa_id):
    report = get_object_or_404(CAPAReport, id=capa_id)
    report.delete()
    return capa_list_partial(request)

@login_required
def download_excel(request, capa_id):
    report = get_object_or_404(CAPAReport, id=capa_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CAPA Report"
    ws.views.sheetView[0].showGridLines = True
    
    # Let's style Excel beautifully using openpyxl
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    navy_fill = PatternFill(start_color='003478', end_color='003478', fill_type='solid')
    light_fill = PatternFill(start_color='E8F0FA', end_color='E8F0FA', fill_type='solid')
    
    # Title Block
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "JINDAL STEEL LIMITED, RAIGARH — CAPA REPORT"
    title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = navy_fill
    
    ws['A3'] = f"CAPA No: {report.capa_no}"
    ws['A3'].font = Font(name='Segoe UI', bold=True)
    ws['D3'] = f"Date of Incident: {report.date_incident}"
    ws['D3'].font = Font(name='Segoe UI', bold=True)
    
    ws['A4'] = f"Department: {report.department.name}"
    ws['A4'].font = Font(name='Segoe UI')
    ws['D4'] = f"Area / Section: {report.area_section}"
    ws['D4'].font = Font(name='Segoe UI')
    
    # 1. Problem Description
    ws['A6'] = "1. Problem Description"
    ws['A6'].font = Font(name='Segoe UI', size=11, bold=True, color='003478')
    ws['A7'] = "What:"
    ws['B7'] = report.problem_what
    ws['A8'] = "Where:"
    ws['B8'] = report.problem_where
    ws['A9'] = "When:"
    ws['B9'] = report.problem_when
    ws['A10'] = "Extent:"
    ws['B10'] = report.problem_extent
    
    # 2. Responsible Team
    ws['A12'] = "2. Responsible Team"
    ws['A12'].font = Font(name='Segoe UI', size=11, bold=True, color='003478')
    
    headers = ["Team Leader / Member", "Role/Function", "Contact No"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col_idx)
        cell.value = h
        cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0057A8', end_color='0057A8', fill_type='solid')
        
    row_idx = 14
    for member in report.responsible_team:
        ws.cell(row=row_idx, column=1, value=member.get('name', ''))
        ws.cell(row=row_idx, column=2, value=member.get('role', ''))
        ws.cell(row=row_idx, column=3, value=member.get('contact', ''))
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1
        
    # 3. ROOT CAUSE 5 WHY
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="3. Root Cause Analysis (5 Why)").font = Font(name='Segoe UI', size=11, bold=True, color='003478')
    row_idx += 1
    
    whys = [report.why_1, report.why_2, report.why_3, report.why_4, report.why_5]
    for idx, why in enumerate(whys, 1):
        ws.cell(row=row_idx, column=1, value=f"{idx} Why:").font = Font(name='Segoe UI', bold=True)
        ws.cell(row=row_idx, column=2, value=why)
        row_idx += 1
        
    # Corrective actions
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="4. Recommended Corrective Action(s)").font = Font(name='Segoe UI', size=11, bold=True, color='003478')
    row_idx += 1
    
    headers = ["Action Details", "Responsibility", "Target Date", "Implementation Date"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = h
        cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0057A8', end_color='0057A8', fill_type='solid')
    row_idx += 1
    
    for action in report.corrective_actions:
        ws.cell(row=row_idx, column=1, value=action.get('action', ''))
        ws.cell(row=row_idx, column=2, value=action.get('responsibility', ''))
        ws.cell(row=row_idx, column=3, value=action.get('target_date', ''))
        ws.cell(row=row_idx, column=4, value=action.get('impl_date', ''))
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1
        
    # Preventive Actions
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="5. Recommended Preventive Action(s)").font = Font(name='Segoe UI', size=11, bold=True, color='003478')
    row_idx += 1
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = h
        cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0057A8', end_color='0057A8', fill_type='solid')
    row_idx += 1
    
    for action in report.preventive_actions:
        ws.cell(row=row_idx, column=1, value=action.get('action', ''))
        ws.cell(row=row_idx, column=2, value=action.get('responsibility', ''))
        ws.cell(row=row_idx, column=3, value=action.get('target_date', ''))
        ws.cell(row=row_idx, column=4, value=action.get('impl_date', ''))
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1
        
    # Effectiveness
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="6. Effectiveness Evaluation").font = Font(name='Segoe UI', size=11, bold=True, color='003478')
    row_idx += 1
    ws.cell(row=row_idx, column=1, value=report.effectiveness_evaluation)
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"CAPA_{report.capa_no or 'Draft'}_{report.department.code}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response@login_required
def download_pdf(request, capa_id):
    report = get_object_or_404(CAPAReport, id=capa_id)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='CAPATitle',
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
        alignment=1,
        textColor=colors.black
    )
    hdr_meta_style = ParagraphStyle(
        name='CAPAHdrMeta',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        alignment=0
    )
    section_title_style = ParagraphStyle(
        name='CAPASectionTitle',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.black,
        spaceBefore=4,
        spaceAfter=4
    )
    body_style = ParagraphStyle(
        name='CAPABody',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.black
    )
    body_bold_style = ParagraphStyle(
        name='CAPABodyBold',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.black
    )
    
    def check_box(label, checked):
        tick = "Y" if checked else " "
        return f"[{tick}] {label}"
        
    story = []
    
    # ------------------ PAGE 1 ------------------
    header_data = [
        [
            Paragraph("Document No. F-01(10.2.0-01)<br/>Issue No. 8<br/>Issue Date 19.11.2025", hdr_meta_style),
            Paragraph("JINDAL STEEL LIMITED, RAIGARH<br/><br/>Corrective and Preventive Action Report", title_style)
        ]
    ]
    header_table = Table(header_data, colWidths=[150, 373])
    header_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('LINEBEFORE', (1, 0), (1, 0), 1, colors.black),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 8))
    
    meta_data = [
        [
            Paragraph("<b>Department:</b> " + (report.department.name if report.department else ""), body_style),
            Paragraph("<b>Area / Section:</b> " + report.area_section, body_style)
        ],
        [
            Paragraph("<b>Date of Incident:</b> " + report.date_incident, body_style),
            Paragraph("<b>CAPA No:</b> " + report.capa_no, body_style)
        ]
    ]
    meta_table = Table(meta_data, colWidths=[261, 262])
    meta_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 8))
    
    b_cat_1 = check_box(">= 4 Hrs.", report.breakdown_applicable == ">= 4 Hrs." or report.breakdown_applicable == "≥ 4 Hrs.")
    b_cat_2 = check_box("2 - 4 Hrs.", report.breakdown_applicable == "2 - 4 Hrs." or report.breakdown_applicable == "2 – 4 Hrs.")
    b_cat_3 = check_box("1 - 2 Hrs.", report.breakdown_applicable == "1 - 2 Hrs." or report.breakdown_applicable == "1 – 2 Hrs.")
    b_cat_4 = check_box("<= 1 Hrs.", report.breakdown_applicable == "<= 1 Hrs." or report.breakdown_applicable == "≤ 1 Hrs.")
    checkboxes_str = f"{b_cat_1}     {b_cat_2}     {b_cat_3}     {b_cat_4}"
    
    prob_data = [
        [Paragraph("<b>1. Problem description (what/where/when/how extensive?)</b>", section_title_style), ""],
        [Paragraph("<b>What:</b>", body_bold_style), Paragraph(report.problem_what, body_style)],
        [Paragraph("<b>Where:</b>", body_bold_style), Paragraph(report.problem_where, body_style)],
        [Paragraph("<b>When:</b>", body_bold_style), Paragraph(report.problem_when, body_style)],
        [Paragraph("<b>Extent:</b>", body_bold_style), Paragraph(report.problem_extent, body_style)],
        [
            Paragraph("<b>Request to Please Tick (Y / N)<br/>the duration of breakdown,<br/>which is applicable:</b>", body_style),
            Paragraph(checkboxes_str, body_bold_style)
        ],
        [
            Paragraph("<b>In Case of Breakdown:</b>", body_bold_style),
            Paragraph(f"Duration of Breakdown (Hrs.): <b>{report.breakdown_hrs}</b>   From (Time): <b>{report.breakdown_from}</b>   To (time): <b>{report.breakdown_to}</b>", body_style)
        ]
    ]
    prob_table = Table(prob_data, colWidths=[150, 373])
    prob_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('SPAN', (0, 0), (1, 0)),
        ('LINEBELOW', (0, 0), (1, 0), 1, colors.black),
        ('INNERGRID', (0, 1), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(prob_table)
    story.append(Spacer(1, 8))
    
    team_rows = [
        [
            Paragraph("<b>Role Designation</b>", body_bold_style),
            Paragraph("<b>Team Member Name</b>", body_bold_style),
            Paragraph("<b>Role / Function</b>", body_bold_style),
            Paragraph("<b>Contact No</b>", body_bold_style)
        ]
    ]
    for idx in range(3):
        member = {}
        if idx < len(report.responsible_team):
            member = report.responsible_team[idx]
        role_lbl = "Team Leader" if idx == 0 else f"Team Member {idx}"
        team_rows.append([
            Paragraph(role_lbl, body_bold_style),
            Paragraph(member.get('name', ''), body_style),
            Paragraph(member.get('role', ''), body_style),
            Paragraph(member.get('contact', ''), body_style)
        ])
    team_table = Table(team_rows, colWidths=[130, 150, 140, 103])
    team_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F0FA')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(Paragraph("<b>2. Responsible Team for Corrective/Preventive Actions:</b>", section_title_style))
    story.append(team_table)
    story.append(Spacer(1, 8))
    
    corr_imm_data = [
        [Paragraph("<b>3. Correction/ Immediate Actions taken:</b>", section_title_style), ""],
        [Paragraph(report.immediate_action, body_style), ""],
        [
            Paragraph("<b>Time Frame:</b> " + report.action_timeframe, body_style),
            Paragraph("<b>Responsibility:</b> " + report.action_responsibility, body_style)
        ]
    ]
    corr_imm_table = Table(corr_imm_data, colWidths=[261, 262])
    corr_imm_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('SPAN', (0, 0), (1, 0)),
        ('SPAN', (0, 1), (1, 1)),
        ('LINEBELOW', (0, 0), (1, 0), 1, colors.black),
        ('LINEBELOW', (0, 1), (1, 1), 0.5, colors.black),
        ('LINEBEFORE', (1, 2), (1, 2), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(corr_imm_table)
    story.append(Spacer(1, 8))
    
    why_data_p1 = [
        [Paragraph("<b>4. Root cause analysis - Analysis finding</b>", section_title_style), ""],
        [Paragraph("<b>1st Why:</b>", body_bold_style), Paragraph(report.why_1, body_style)],
        [Paragraph("<b>2nd Why:</b>", body_bold_style), Paragraph(report.why_2, body_style)]
    ]
    why_table_p1 = Table(why_data_p1, colWidths=[100, 423])
    why_table_p1.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('SPAN', (0, 0), (1, 0)),
        ('LINEBELOW', (0, 0), (1, 0), 1, colors.black),
        ('INNERGRID', (0, 1), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(why_table_p1)
    
    # ------------------ PAGE 2 ------------------
    story.append(PageBreak())
    
    p2_header = Table([[
        Paragraph("Document No. F-01(10.2.0-01)   Issue No. 8   Issue Date 19.11.2025", body_style),
        Paragraph("<b>Page 2 of 3</b>", ParagraphStyle('PageR', parent=body_style, alignment=2))
    ]], colWidths=[350, 173])
    p2_header.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(p2_header)
    story.append(Spacer(1, 8))
    
    m_list = ["1M Material", "2M Man", "3M Machine", "4M Measure", "5M Method"]
    m_checkboxes = []
    for m in m_list:
        checked = m in report.five_m_applicable
        m_checkboxes.append(Paragraph(check_box(m, checked), body_bold_style))
        
    m_checklist_cell = Table([[c] for c in m_checkboxes], colWidths=[130])
    m_checklist_cell.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    
    why_continued_inner = Table([
        [Paragraph("<b>3rd Why:</b>", body_bold_style), Paragraph(report.why_3, body_style)],
        [Paragraph("<b>4th Why:</b>", body_bold_style), Paragraph(report.why_4, body_style)],
        [Paragraph("<b>5th Why:</b>", body_bold_style), Paragraph(report.why_5, body_style)],
    ], colWidths=[80, 291])
    why_continued_inner.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    why_continued_table = Table([[why_continued_inner, m_checklist_cell]], colWidths=[373, 150])
    why_continued_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('LINEBEFORE', (1, 0), (1, 0), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(Paragraph("<b>4. Root cause analysis - Continued</b>", section_title_style))
    story.append(why_continued_table)
    story.append(Spacer(1, 4))
    
    conclusion_table = Table([[
        Paragraph("<b>Conclusion (s):</b>", body_bold_style),
        Paragraph(report.conclusion, body_style)
    ]], colWidths=[100, 423])
    conclusion_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(conclusion_table)
    story.append(Spacer(1, 8))
    
    ca_rows = [
        [
            Paragraph("<b>Corrective Action(s)</b>", body_bold_style),
            Paragraph("<b>Responsibility</b>", body_bold_style),
            Paragraph("<b>Target Date</b>", body_bold_style),
            Paragraph("<b>Date of implementation</b>", body_bold_style)
        ]
    ]
    for idx in range(2):
        action = {}
        if idx < len(report.corrective_actions):
            action = report.corrective_actions[idx]
        ca_rows.append([
            Paragraph(action.get('action', ''), body_style),
            Paragraph(action.get('responsibility', ''), body_style),
            Paragraph(action.get('target_date', ''), body_style),
            Paragraph(action.get('impl_date', ''), body_style)
        ])
    ca_table = Table(ca_rows, colWidths=[200, 100, 100, 123])
    ca_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F0FA')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(Paragraph("<b>5. Recommended Corrective action(s):</b>", section_title_style))
    story.append(ca_table)
    story.append(Spacer(1, 8))
    
    pa_rows = [
        [
            Paragraph("<b>Preventive Action(s)</b>", body_bold_style),
            Paragraph("<b>Responsibility</b>", body_bold_style),
            Paragraph("<b>Target Date</b>", body_bold_style),
            Paragraph("<b>Date of implementation</b>", body_bold_style)
        ]
    ]
    for idx in range(2):
        action = {}
        if idx < len(report.preventive_actions):
            action = report.preventive_actions[idx]
        pa_rows.append([
            Paragraph(action.get('action', ''), body_style),
            Paragraph(action.get('responsibility', ''), body_style),
            Paragraph(action.get('target_date', ''), body_style),
            Paragraph(action.get('impl_date', ''), body_style)
        ])
    pa_table = Table(pa_rows, colWidths=[200, 100, 100, 123])
    pa_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F0FA')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(Paragraph("<b>6. Recommended Preventive action(s):</b>", section_title_style))
    story.append(pa_table)
    story.append(Spacer(1, 8))
    
    plan_table = Table([[
        Paragraph("<b>7. Detailed Implementation Plan:</b>", body_bold_style),
        Paragraph(report.detailed_plan, body_style)
    ]], colWidths=[150, 373])
    plan_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(plan_table)
    story.append(Spacer(1, 8))
    
    doc_options = [
        "MOC",
        "SOP / SMP",
        "Risk and Opportunity Register",
        "Register of Environmental Aspect Impact and OH & S Risks",
        "Training Need Identification"
    ]
    doc_checkboxes = []
    for d_opt in doc_options:
        checked = d_opt in report.modified_documents
        doc_checkboxes.append(Paragraph(check_box(d_opt, checked), body_bold_style))
        
    doc_checkboxes.append(Paragraph(f"<b>Others (Please mention) :</b> {report.modified_documents_other}", body_style))
    
    doc_table = Table([
        [doc_checkboxes[0], doc_checkboxes[1]],
        [doc_checkboxes[2], doc_checkboxes[3]],
        [doc_checkboxes[4], doc_checkboxes[5]],
    ], colWidths=[261, 262])
    doc_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(Paragraph("<b>8. Modified documents (Please Tick in the applicable document)</b>", section_title_style))
    story.append(doc_table)
    story.append(Spacer(1, 8))
    
    train_table = Table([[
        Paragraph("<b>9. Training Details (If any) :</b>", body_bold_style),
        Paragraph(report.training_details, body_style)
    ]], colWidths=[150, 373])
    train_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(train_table)
    story.append(Spacer(1, 8))
    
    impl_date_table = Table([[
        Paragraph("<b>10. Date of Implementation :</b>", body_bold_style),
        Paragraph(report.date_implementation, body_style)
    ]], colWidths=[150, 373])
    impl_date_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(impl_date_table)
    story.append(Spacer(1, 8))
    
    eval_table = Table([[
        Paragraph("<b>11. Effectiveness evaluation:</b><br/><font size='7' color='gray' face='Helvetica'>(What is the activity to verify that the corrective actions have been effective (measurement, audit, study, assessment)?</font>", body_bold_style),
        Paragraph(report.effectiveness_evaluation, body_style)
    ]], colWidths=[150, 373])
    eval_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(eval_table)
    
    # ------------------ PAGE 3 ------------------
    story.append(PageBreak())
    
    p3_header = Table([[
        Paragraph("Document No. F-01(10.2.0-01)   Issue No. 8   Issue Date 19.11.2025", body_style),
        Paragraph("<b>Page 3 of 3</b>", ParagraphStyle('PageR', parent=body_style, alignment=2))
    ]], colWidths=[350, 173])
    p3_header.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(p3_header)
    story.append(Spacer(1, 40))
    
    sig_data = [
        [
            Paragraph("<b>Prepared By</b><br/>(Name and Signature of Initiator)", ParagraphStyle('SigC', parent=body_bold_style, alignment=1)),
            Paragraph("<b>Reviewed By</b><br/>(Name and Signature of Reviewer)", ParagraphStyle('SigC', parent=body_bold_style, alignment=1)),
            Paragraph("<b>Approved by</b><br/>(Name and Signature of HOD)", ParagraphStyle('SigC', parent=body_bold_style, alignment=1))
        ],
        [
            Paragraph(f"<br/><br/><br/><br/><b>{report.prepared_by}</b>", ParagraphStyle('SigC', parent=body_style, alignment=1)),
            Paragraph(f"<br/><br/><br/><br/><b>{report.reviewed_by}</b>", ParagraphStyle('SigC', parent=body_style, alignment=1)),
            Paragraph(f"<br/><br/><br/><br/><b>{report.approved_by}</b>", ParagraphStyle('SigC', parent=body_style, alignment=1))
        ]
    ]
    sig_table = Table(sig_data, colWidths=[174, 174, 175])
    sig_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(sig_table)
    
    def draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.drawString(36, 18, "F-01(10.2.0-01)/Issue 07/ Issue Date: 19.11.2025")
        canvas.restoreState()
        
    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    buffer.seek(0)
    
    filename = f"CAPA_{report.capa_no or 'Draft'}_{report.department.code}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
