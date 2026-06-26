# tpm/utils/export.py

import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Circle
from reportlab.graphics.charts.piecharts import Pie

from tpm.models import PillarEntry, KPIValue, WorkstationValue, Workstation
from tpm.utils.kpi_definitions import KPI_DEFINITIONS
from tpm.utils.calculations import compute_achievement, get_date_range_q, aggregate_kpi_actual

def generate_pillar_excel(dept, from_month, from_year, to_month, to_year, filter_type):
    """Generates an Excel workbook with sheets for Summary and each active Pillar"""
    wb = Workbook()
    
    months_map = dict([
        (1, 'Jan'), (2, 'Feb'), (3, 'Mar'), (4, 'Apr'),
        (5, 'May'), (6, 'Jun'), (7, 'Jul'), (8, 'Aug'),
        (9, 'Sep'), (10, 'Oct'), (11, 'Nov'), (12, 'Dec')
    ])
    if filter_type == 'range':
        period_label = f"{months_map.get(from_month)} {from_year} - {months_map.get(to_month)} {to_year}"
    else:
        months_full = [
            (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
            (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
            (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
        ]
        period_label = f"{dict(months_full).get(from_month)} {from_year}"

    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary Overview"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells('A1:E2')
    title_cell = ws_summary['A1']
    title_cell.value = f"JINDAL STEEL LTD — TPM PORTAL"
    title_cell.font = Font(name='Segoe UI', size=16, bold=True, color='003478')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_summary['A3'] = f"Department: {dept.name} ({dept.code})"
    ws_summary['A3'].font = Font(name='Segoe UI', size=11, bold=True)
    ws_summary['A4'] = f"Report Period: {period_label}"
    ws_summary['A4'].font = Font(name='Segoe UI', size=11, italic=True)
    
    headers = ["Pillar Code", "Pillar Name", "Total KPIs", "On Track (>=90%)", "At Risk (75-89%)", "Behind (<75%)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=6, column=col_idx)
        cell.value = h
        cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003478', end_color='003478', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        
    thin_border = Border(
        left=Side(style='thin', color='D1DCF0'),
        right=Side(style='thin', color='D1DCF0'),
        top=Side(style='thin', color='D1DCF0'),
        bottom=Side(style='thin', color='D1DCF0')
    )
    
    pillars_meta = [
        ('KK', 'Kobetsu Kaizen'),
        ('JH', 'Jishu Hozen'),
        ('PM', 'Planned Maintenance'),
        ('QM', 'Quality Maintenance'),
        ('ET', 'Education & Training'),
        ('DM', 'Initial Flow Control / Design & Management'),
        ('SHE', 'Safety, Health & Environment'),
        ('OTPM', 'Office TPM')
    ]
    
    row_idx = 7
    for code, name in pillars_meta:
        definitions = KPI_DEFINITIONS.get(code, [])
        total = len(definitions)
        on_track = 0
        at_risk = 0
        behind = 0
        has_any_data = False
        
        for d in definitions:
            kpi_values = KPIValue.objects.filter(
                get_date_range_q(prefix='pillar_entry__', from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
                pillar_entry__department=dept,
                pillar_entry__pillar=code,
                sl_no=d['sl_no']
            )
            if kpi_values.exists():
                has_any_data = True
                actual, target, benchmark = aggregate_kpi_actual(kpi_values, d['uom'], d['name'])
                if actual is not None and target is not None:
                    ach = compute_achievement(actual, target, d['name'])
                    if ach >= 90:
                        on_track += 1
                    elif ach >= 75:
                        at_risk += 1
                    else:
                        behind += 1
                        
        ws_summary.cell(row=row_idx, column=1, value=code)
        ws_summary.cell(row=row_idx, column=2, value=name)
        ws_summary.cell(row=row_idx, column=3, value=total if has_any_data else "N/A")
        ws_summary.cell(row=row_idx, column=4, value=on_track if has_any_data else "N/A")
        ws_summary.cell(row=row_idx, column=5, value=at_risk if has_any_data else "N/A")
        ws_summary.cell(row=row_idx, column=6, value=behind if has_any_data else "N/A")
        
        for col in range(1, 7):
            c = ws_summary.cell(row=row_idx, column=col)
            c.border = thin_border
            c.font = Font(name='Segoe UI')
            if col > 2:
                c.alignment = Alignment(horizontal='center')
        
        row_idx += 1
        
    # Autofit columns
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 2. Add individual Pillar Sheets
    for code, name in pillars_meta:
        ws = wb.create_sheet(title=code)
        ws.views.sheetView[0].showGridLines = True
        
        ws.cell(row=1, column=1, value=f"{code} — {name} Report ({period_label})").font = Font(name='Segoe UI', size=14, bold=True, color='003478')
        
        headers = ["Sl No", "KPI Name", "UOM", "Benchmark", "Target", "Actual", "Achievement %", "Remarks"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = h
            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0057A8', end_color='0057A8', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        definitions = KPI_DEFINITIONS.get(code, [])
        
        p_row = 4
        for d in definitions:
            kpi_values = KPIValue.objects.filter(
                get_date_range_q(prefix='pillar_entry__', from_month=from_month, from_year=from_year, to_month=to_month, to_year=to_year),
                pillar_entry__department=dept,
                pillar_entry__pillar=code,
                sl_no=d['sl_no']
            )
            
            if kpi_values.exists():
                actual, target, benchmark = aggregate_kpi_actual(kpi_values, d['uom'], d['name'])
                remarks_list = [v.remarks.strip() for v in kpi_values if v.remarks.strip()]
                remarks = " | ".join(remarks_list) if remarks_list else ""
            else:
                actual = None
                target = d['target']
                benchmark = d['benchmark']
                remarks = ''
            
            achievement = ""
            if actual is not None and target is not None:
                ach = compute_achievement(actual, target, d['name'])
                achievement = f"{round(ach, 1)}%"
                
            ws.cell(row=p_row, column=1, value=d['sl_no']).alignment = Alignment(horizontal='center')
            ws.cell(row=p_row, column=2, value=d['name'])
            ws.cell(row=p_row, column=3, value=d['uom']).alignment = Alignment(horizontal='center')
            
            ws.cell(row=p_row, column=4, value=benchmark if benchmark is not None else "—").alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=5, value=target if target is not None else "—").alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=6, value=actual if actual is not None else "—").alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=7, value=achievement).alignment = Alignment(horizontal='right')
            ws.cell(row=p_row, column=8, value=remarks)
            
            # Format row
            for col in range(1, 9):
                c = ws.cell(row=p_row, column=col)
                c.border = thin_border
                c.font = Font(name='Segoe UI')
                
            p_row += 1
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
    # Save to IO stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_months_in_range(from_month, from_year, to_month, to_year):
    months = []
    current_month = from_month
    current_year = from_year
    while current_year < to_year or (current_year == to_year and current_month <= to_month):
        months.append((current_month, current_year))
        current_month += 1
        if current_month > 12:
            current_month = 1
            current_year += 1
    return months


def get_month_data(dept, m, y):
    pillars_meta = [
        ('KK', 'Kobetsu Kaizen'),
        ('JH', 'Jishu Hozen'),
        ('PM', 'Planned Maintenance'),
        ('QM', 'Quality Maintenance'),
        ('ET', 'Education & Training'),
        ('DM', 'Initial Flow Control / Design & Management'),
        ('SHE', 'Safety, Health & Environment'),
        ('OTPM', 'Office TPM')
    ]
    report_data = []
    
    for code, name in pillars_meta:
        entries = PillarEntry.objects.filter(
            month=m, year=y,
            department=dept,
            pillar=code
        )
        all_submitted = entries.exists() and all(e.is_locked() for e in entries)
        
        definitions = KPI_DEFINITIONS.get(code, [])
        kpi_list = []
        achievements = []
        
        for d in definitions:
            kpi_values = KPIValue.objects.filter(
                pillar_entry__month=m,
                pillar_entry__year=y,
                pillar_entry__department=dept,
                pillar_entry__pillar=code,
                sl_no=d['sl_no']
            )
            
            if kpi_values.exists():
                actual, target, benchmark = aggregate_kpi_actual(kpi_values, d['uom'], d['name'])
                remarks_list = [v.remarks.strip() for v in kpi_values if v.remarks.strip()]
                remarks = " | ".join(remarks_list) if remarks_list else ""
            else:
                actual = None
                target = d['target']
                benchmark = d['benchmark']
                remarks = ''
                
            achievement = None
            status = 'pending'
            if actual is not None and target is not None:
                achievement = compute_achievement(actual, target, d['name'])
                status = 'on-track' if achievement >= 90 else ('at-risk' if achievement >= 75 else 'behind')
                achievements.append(achievement)
                
            kpi_list.append({
                'sl_no': d['sl_no'],
                'name': d['name'],
                'uom': d['uom'],
                'benchmark': benchmark,
                'target': target,
                'actual': actual,
                'achievement': achievement,
                'remarks': remarks,
                'status': status,
            })
            
        avg_achievement = sum(achievements) / len(achievements) if achievements else 0.0
        
        report_data.append({
            'code': code,
            'name': name,
            'kpis': kpi_list,
            'achievement': round(avg_achievement, 1),
            'submitted': all_submitted,
        })
        
    return report_data


def draw_bar_chart(width, height, labels, scores, colors_list):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.HexColor('#F8FAFC'), strokeColor=colors.HexColor('#E2E8F0'), strokeWidth=1, rx=6, ry=6))
    
    chart_x = 40 if width < 400 else 45
    chart_y = 30
    chart_w = width - chart_x - 15 if width < 400 else width - chart_x - 20
    chart_h = height - chart_y - 20
    
    for i in range(6):
        val = i * 20
        y = chart_y + (val / 100.0) * chart_h
        d.add(Line(chart_x, y, chart_x + chart_w, y, strokeColor=colors.HexColor('#E2E8F0'), strokeWidth=0.5))
        d.add(String(chart_x - 6, y - 2.5, f"{val}%", fontName="Helvetica", fontSize=6.5 if width < 400 else 7, textAnchor="end", fillColor=colors.HexColor('#64748B')))
        
    n_bars = len(scores)
    bar_gap = 6 if width < 400 else 12
    total_gaps_w = bar_gap * (n_bars + 1)
    bar_w = (chart_w - total_gaps_w) / n_bars
    
    label_font_size = 6.5 if width < 400 else 8
    
    for idx, (lbl, score) in enumerate(zip(labels, scores)):
        x = chart_x + bar_gap + idx * (bar_w + bar_gap)
        val = score if score is not None else 0.0
        val_capped = max(0.0, min(val, 100.0))
        bar_h = (val_capped / 100.0) * chart_h
        color = colors_list[idx % len(colors_list)]
        
        d.add(Rect(x, chart_y, bar_w, bar_h, fillColor=colors.HexColor(color), strokeColor=None))
        d.add(String(x + bar_w/2.0, chart_y + bar_h + 4, f"{score:.1f}%" if score is not None else "—", fontName="Helvetica-Bold", fontSize=6.5 if width < 400 else 7, textAnchor="middle", fillColor=colors.HexColor('#1E293B')))
        d.add(String(x + bar_w/2.0, chart_y - 12, lbl, fontName="Helvetica-Bold", fontSize=label_font_size, textAnchor="middle", fillColor=colors.HexColor('#334155')))
        
    return d


def draw_pie_chart(width, height, on_track, at_risk, behind):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.HexColor('#F8FAFC'), strokeColor=colors.HexColor('#E2E8F0'), strokeWidth=1, rx=6, ry=6))
    
    total = on_track + at_risk + behind
    if total == 0:
        d.add(Circle(width/2.0, height/2.0, 40, fillColor=colors.HexColor('#E2E8F0'), strokeColor=colors.HexColor('#CBD5E1')))
        d.add(String(width/2.0, height/2.0 - 3, "No KPI Data", fontName="Helvetica-Bold", fontSize=8, textAnchor="middle", fillColor=colors.HexColor('#64748B')))
        return d
        
    pc = Pie()
    pc.x = 15
    pc.y = 25
    pc.width = 110
    pc.height = 110
    pc.data = []
    
    slice_colors = []
    legend_items = []
    
    if on_track > 0:
        pc.data.append(on_track)
        slice_colors.append(colors.HexColor('#16A34A'))
        legend_items.append((f"On Track: {on_track}", '#16A34A'))
    if at_risk > 0:
        pc.data.append(at_risk)
        slice_colors.append(colors.HexColor('#D97706'))
        legend_items.append((f"At Risk: {at_risk}", '#D97706'))
    if behind > 0:
        pc.data.append(behind)
        slice_colors.append(colors.HexColor('#DC2626'))
        legend_items.append((f"Behind: {behind}", '#DC2626'))
        
    pc.labels = []
    pc.slices.strokeWidth = 0.5
    pc.slices.strokeColor = colors.white
    
    for idx, color in enumerate(slice_colors):
        pc.slices[idx].fillColor = color
        
    d.add(pc)
    
    # Custom legend
    leg_x = 135
    leg_y = 50
    for text, color in legend_items:
        d.add(Rect(leg_x, leg_y, 7, 7, fillColor=colors.HexColor(color), strokeColor=None))
        d.add(String(leg_x + 11, leg_y + 0.5, text, fontName="Helvetica", fontSize=7, fillColor=colors.HexColor('#334155')))
        leg_y += 15
        
    d.add(String(10, height - 16, "KPI Status Distribution", fontName="Helvetica-Bold", fontSize=8, fillColor=colors.HexColor('#1E293B')))
    
    return d


def draw_line_chart(width, height, months_labels, overall_trends, pillar_trends):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.HexColor('#F8FAFC'), strokeColor=colors.HexColor('#E2E8F0'), strokeWidth=1, rx=6, ry=6))
    
    chart_x = 45
    chart_y = 35
    chart_w = width - chart_x - 20
    chart_h = height - chart_y - 20
    
    for i in range(6):
        val = i * 20
        y = chart_y + (val / 100.0) * chart_h
        d.add(Line(chart_x, y, chart_x + chart_w, y, strokeColor=colors.HexColor('#E2E8F0'), strokeWidth=0.5))
        d.add(String(chart_x - 8, y - 3, f"{val}%", fontName="Helvetica", fontSize=7, textAnchor="end", fillColor=colors.HexColor('#64748B')))
        
    n_points = len(months_labels)
    if n_points < 1:
        return d
        
    x_coords = []
    if n_points == 1:
        x_coords.append(chart_x + chart_w/2.0)
    else:
        for i in range(n_points):
            x_coords.append(chart_x + (i / (n_points - 1)) * chart_w)
            
    for idx, lbl in enumerate(months_labels):
        x = x_coords[idx]
        d.add(String(x, chart_y - 12, lbl, fontName="Helvetica-Bold", fontSize=7, textAnchor="middle", fillColor=colors.HexColor('#475569')))
        
    series_list = []
    series_list.append(('Overall TPM', overall_trends, '#F47920', 2.5))
    
    key_pillars = [
        ('KK', '#003478'),
        ('JH', '#0057A8'),
        ('PM', '#16a34a'),
        ('SHE', '#dc2626')
    ]
    for code, color in key_pillars:
        if code in pillar_trends:
            series_list.append((code, pillar_trends[code], color, 1.2))
            
    for label, scores, color, thickness in series_list:
        points = []
        for idx, val in enumerate(scores):
            if val is not None:
                x = x_coords[idx]
                val_capped = max(0.0, min(val, 100.0))
                y = chart_y + (val_capped / 100.0) * chart_h
                points.append((x, y))
                
        for p1, p2 in zip(points[:-1], points[1:]):
            d.add(Line(p1[0], p1[1], p2[0], p2[1], strokeColor=colors.HexColor(color), strokeWidth=thickness))
            
        for x, y in points:
            d.add(Circle(x, y, thickness + 1, fillColor=colors.HexColor(color), strokeColor=colors.white, strokeWidth=0.5))
            
    legend_x = chart_x + 10
    for label, _, color, thickness in series_list:
        d.add(Line(legend_x, 10, legend_x + 12, 10, strokeColor=colors.HexColor(color), strokeWidth=thickness))
        d.add(String(legend_x + 16, 7, label, fontName="Helvetica", fontSize=7, fillColor=colors.HexColor('#334155')))
        legend_x += 80
        
    return d


def add_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#64748B'))
    canvas.drawString(40, 25, "JINDAL STEEL LIMITED — TPM PERFORMANCE REPORT")
    canvas.drawRightString(letter[0] - 40, 25, f"Page {doc.page}")
    canvas.setStrokeColor(colors.HexColor('#E2E8F0'))
    canvas.setLineWidth(0.5)
    canvas.line(40, 35, letter[0] - 40, 35)
    canvas.restoreState()


def generate_pillar_pdf(dept, from_month, from_year, to_month, to_year, filter_type):
    """Generates a professional PDF monthly report for the department using ReportLab"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    title_style = ParagraphStyle(
        name='JSPLTitle',
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=18,
        textColor=colors.HexColor('#003478')
    )
    
    section_style = ParagraphStyle(
        name='JSPLSection',
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#0057A8'),
        spaceBefore=12,
        spaceAfter=6,
        keepWithNext=True
    )
    
    table_hdr_style = ParagraphStyle(
        name='JSPLTableHdr',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white
    )
    
    table_cell_style = ParagraphStyle(
        name='JSPLTableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#1E293B')
    )
    
    table_cell_mono = ParagraphStyle(
        name='JSPLTableCellMono',
        fontName='Courier',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#1E293B')
    )

    danger_cell_style = ParagraphStyle(
        name='JSPLDangerCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#DC2626')
    )
    
    danger_cell_mono = ParagraphStyle(
        name='JSPLDangerCellMono',
        fontName='Courier-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#DC2626')
    )

    story = []
    
    utils_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.abspath(os.path.join(utils_dir, '..', 'static', 'images', 'logo.png'))

    months_map_short = dict([
        (1, 'Jan'), (2, 'Feb'), (3, 'Mar'), (4, 'Apr'),
        (5, 'May'), (6, 'Jun'), (7, 'Jul'), (8, 'Aug'),
        (9, 'Sep'), (10, 'Oct'), (11, 'Nov'), (12, 'Dec')
    ])
    months_full = dict([
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ])
    
    if filter_type == 'range':
        period_label = f"{months_map_short.get(from_month)} {from_year} - {months_map_short.get(to_month)} {to_year}"
    else:
        period_label = f"{months_full.get(from_month)} {from_year}"

    def get_letterhead_flowables(lbl_text):
        letterhead = []
        logo_cell = ""
        if os.path.exists(logo_path):
            logo_cell = Image(logo_path, width=1.4*inch, height=0.4*inch)
            
        title_text = f"<b>JINDAL STEEL LIMITED</b><br/><font size=9 color='#475569'>Monthly TPM Performance Report — {dept.name} ({dept.code})<br/>Period: <b>{lbl_text}</b></font>"
        title_para = Paragraph(title_text, title_style)
        
        header_table = Table([[logo_cell, title_para]], colWidths=[1.5*inch, 5.58*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 0),
        ]))
        letterhead.append(header_table)
        
        divider = Table([[""]], colWidths=[510])
        divider.setStyle(TableStyle([
            ('LINEBELOW', (0,0), (-1,-1), 1.5, colors.HexColor('#003478')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('TOPPADDING', (0,0), (-1,-1), 0),
        ]))
        letterhead.append(divider)
        letterhead.append(Spacer(1, 10))
        return letterhead

    pillars_meta = [
        ('KK', 'Kobetsu Kaizen'),
        ('JH', 'Jishu Hozen'),
        ('PM', 'Planned Maintenance'),
        ('QM', 'Quality Maintenance'),
        ('ET', 'Education & Training'),
        ('DM', 'Initial Flow Control / Design & Management'),
        ('SHE', 'Safety, Health & Environment'),
        ('OTPM', 'Office TPM')
    ]

    if filter_type == 'range':
        # 1. RANGE COVER / EXECUTIVE COMPARISON OVERVIEW PAGE
        story.extend(get_letterhead_flowables(period_label))
        
        story.append(Paragraph("Executive Comparison Overview", section_style))
        story.append(Spacer(1, 5))
        
        months_in_range = get_months_in_range(from_month, from_year, to_month, to_year)
        months_labels = []
        overall_trends = []
        pillar_trends = {code: [] for code, _ in pillars_meta}
        
        monthly_reports_data = []
        
        for m, y in months_in_range:
            lbl = f"{months_map_short.get(m)} {str(y)[2:]}"
            months_labels.append(lbl)
            
            m_report = get_month_data(dept, m, y)
            monthly_reports_data.append({
                'month': m,
                'year': y,
                'label': f"{months_full.get(m)} {y}",
                'report_data': m_report
            })
            
            month_scores = []
            for p_data in m_report:
                pillar_trends[p_data['code']].append(p_data['achievement'])
                month_scores.append(p_data['achievement'])
                
            overall_score = round(sum(month_scores) / len(month_scores), 1) if month_scores else 0.0
            overall_trends.append(overall_score)
            
        chart_flowable = draw_line_chart(510, 180, months_labels, overall_trends, pillar_trends)
        story.append(chart_flowable)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Comparative Performance Summary Table", section_style))
        
        comp_hdr = [Paragraph("<b>Pillar</b>", table_hdr_style)]
        for lbl in months_labels:
            comp_hdr.append(Paragraph(f"<b>{lbl}</b>", table_hdr_style))
            
        comp_table_data = [comp_hdr]
        
        overall_row = [Paragraph("<b>Overall TPM Score</b>", table_cell_style)]
        for val in overall_trends:
            overall_row.append(Paragraph(f"<b>{val:.1f}%</b>", table_cell_style))
        comp_table_data.append(overall_row)
        
        for code, name in pillars_meta:
            row = [Paragraph(f"<b>{code}</b> — {name}", table_cell_style)]
            for val in pillar_trends[code]:
                row.append(Paragraph(f"{val:.1f}%", table_cell_style))
            comp_table_data.append(row)
            
        n_cols = len(comp_hdr)
        col_w = [210] + [300.0 / (n_cols - 1)] * (n_cols - 1)
        
        comp_table = Table(comp_table_data, colWidths=col_w)
        comp_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003478')),
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#E2E8F0')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (1,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(comp_table)
        
        # 2. DETAIL MONTH SUMMARY PAGE (Page 2)
        story.append(PageBreak())
        story.extend(get_letterhead_flowables(period_label))
        story.append(Paragraph("Monthly Performance Summaries & Action Items", section_style))
        story.append(Spacer(1, 5))
        
        for m_data in monthly_reports_data:
            m_report = m_data['report_data']
            
            # Monthly Grid
            score_headers = [Paragraph(f"<b>{p['code']}</b>", table_hdr_style) for p in m_report]
            score_values = [Paragraph(f"{p['achievement']:.1f}%", table_cell_style) for p in m_report]
            
            score_table_data = [
                [Paragraph(f"<b>{m_data['label']}</b>", table_cell_style)] + score_headers,
                [Paragraph("Pillar Score", table_cell_style)] + score_values
            ]
            score_table = Table(score_table_data, colWidths=[110] + [50] * 8)
            score_table.setStyle(TableStyle([
                ('BACKGROUND', (1,0), (-1,0), colors.HexColor('#003478')),
                ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#E2E8F0')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(KeepTogether([
                Spacer(1, 10),
                score_table
            ]))
            
            # Consolidated KPI Performance Sheet for this month
            kpi_rows_data = []
            row_index = 1
            danger_rows = []
            
            for p in m_report:
                for k in p['kpis']:
                    achievement = k['achievement']
                    is_danger = achievement is not None and achievement < 80.0
                    
                    if is_danger:
                        c_style = danger_cell_style
                        c_mono = danger_cell_mono
                        remarks_text = k['remarks'] if k['remarks'] else "No explanation provided"
                        danger_rows.append(row_index)
                    else:
                        c_style = table_cell_style
                        c_mono = table_cell_mono
                        if achievement is not None:
                            remarks_text = "On Track" if achievement >= 90.0 else "At Risk"
                        else:
                            remarks_text = "Pending"
                            
                    kpi_rows_data.append([
                        Paragraph(f"<b>{p['code']}</b>", c_style),
                        Paragraph(k['sl_no'], c_mono),
                        Paragraph(f"{k['name']} ({k['uom']})", c_style),
                        Paragraph(f"{k['target']:.1f}" if k['target'] is not None else "—", c_mono),
                        Paragraph(f"{k['actual']:.1f}" if k['actual'] is not None else "—", c_mono),
                        Paragraph(f"{achievement:.1f}%" if achievement is not None else "—", c_mono),
                        Paragraph(remarks_text, c_style)
                    ])
                    row_index += 1
            
            sheet_headers = [
                Paragraph("<b>Pillar</b>", table_hdr_style),
                Paragraph("<b>Sl</b>", table_hdr_style),
                Paragraph("<b>KPI Name</b>", table_hdr_style),
                Paragraph("<b>Target</b>", table_hdr_style),
                Paragraph("<b>Actual</b>", table_hdr_style),
                Paragraph("<b>Ach %</b>", table_hdr_style),
                Paragraph("<b>Status / Remarks</b>", table_hdr_style)
            ]
            sheet_table_data = [sheet_headers] + kpi_rows_data
            sheet_table = Table(sheet_table_data, colWidths=[40, 20, 160, 45, 45, 50, 150])
            
            t_styles = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003478')),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('ALIGN', (3,1), (5,-1), 'RIGHT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]
            
            for r in range(1, len(sheet_table_data)):
                if r in danger_rows:
                    t_styles.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#FEF2F2')))
                    t_styles.append(('GRID', (0, r), (-1, r), 0.5, colors.HexColor('#FCA5A5')))
                elif r % 2 == 0:
                    t_styles.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#F8FAFC')))
                    
            sheet_table.setStyle(TableStyle(t_styles))
            
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"📋 Monthly KPI Performance Sheet — {m_data['label']}", section_style))
            story.append(sheet_table)

    else:
        # SINGLE MONTH REPORT
        story.extend(get_letterhead_flowables(period_label))
        
        # Calculate KPI counts for Pie Chart
        m_report = get_month_data(dept, from_month, from_year)
        labels = [p['code'] for p in m_report]
        scores = [p['achievement'] for p in m_report]
        
        on_track_total = 0
        at_risk_total = 0
        behind_total = 0
        for p in m_report:
            for k in p['kpis']:
                if k['achievement'] is not None:
                    if k['achievement'] >= 90.0:
                        on_track_total += 1
                    elif k['achievement'] >= 75.0:
                        at_risk_total += 1
                    else:
                        behind_total += 1
        
        # Side-by-side charts
        bar_colors = ['#003478', '#0057A8', '#F47920', '#16a34a', '#d97706', '#dc2626', '#8b5cf6', '#06b6d4']
        bar_chart = draw_bar_chart(280, 180, labels, scores, bar_colors)
        pie_chart = draw_pie_chart(220, 180, on_track_total, at_risk_total, behind_total)
        
        charts_table = Table([[bar_chart, pie_chart]], colWidths=[290, 220])
        charts_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
        ]))
        story.append(charts_table)
        story.append(Spacer(1, 15))
        
        # Pillar Summary Overview Table
        story.append(Paragraph("Pillar Summary Overview", section_style))
        overview_hdr = [
            Paragraph("<b>Pillar</b>", table_hdr_style),
            Paragraph("<b>Pillar Score</b>", table_hdr_style),
            Paragraph("<b>Status</b>", table_hdr_style),
        ]
        overview_table_data = [overview_hdr]
        
        for p in m_report:
            status_str = "<b>Locked</b>" if p['submitted'] else "Pending"
            status_color = '#16a34a' if p['submitted'] else '#dc2626'
            overview_table_data.append([
                Paragraph(f"<b>{p['code']}</b> — {p['name']}", table_cell_style),
                Paragraph(f"{p['achievement']:.1f}%", table_cell_style),
                Paragraph(f"<font color='{status_color}'>{status_str}</font>", table_cell_style)
            ])
            
        overview_table = Table(overview_table_data, colWidths=[270, 120, 120])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003478')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1DCF0')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (1,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(overview_table)
        story.append(Spacer(1, 15))
        
        # Consolidated KPI Table for Single Month
        story.append(PageBreak())
        story.extend(get_letterhead_flowables(period_label))
        
        kpi_rows_data = []
        row_index = 1
        danger_rows = []
        
        for p in m_report:
            for k in p['kpis']:
                achievement = k['achievement']
                is_danger = achievement is not None and achievement < 80.0
                
                if is_danger:
                    c_style = danger_cell_style
                    c_mono = danger_cell_mono
                    remarks_text = k['remarks'] if k['remarks'] else "No explanation provided"
                    danger_rows.append(row_index)
                else:
                    c_style = table_cell_style
                    c_mono = table_cell_mono
                    if achievement is not None:
                        remarks_text = "On Track" if achievement >= 90.0 else "At Risk"
                    else:
                        remarks_text = "Pending"
                        
                kpi_rows_data.append([
                    Paragraph(f"<b>{p['code']}</b>", c_style),
                    Paragraph(k['sl_no'], c_mono),
                    Paragraph(f"{k['name']} ({k['uom']})", c_style),
                    Paragraph(f"{k['target']:.1f}" if k['target'] is not None else "—", c_mono),
                    Paragraph(f"{k['actual']:.1f}" if k['actual'] is not None else "—", c_mono),
                    Paragraph(f"{achievement:.1f}%" if achievement is not None else "—", c_mono),
                    Paragraph(remarks_text, c_style)
                ])
                row_index += 1
                
        sheet_headers = [
            Paragraph("<b>Pillar</b>", table_hdr_style),
            Paragraph("<b>Sl</b>", table_hdr_style),
            Paragraph("<b>KPI Name</b>", table_hdr_style),
            Paragraph("<b>Target</b>", table_hdr_style),
            Paragraph("<b>Actual</b>", table_hdr_style),
            Paragraph("<b>Ach %</b>", table_hdr_style),
            Paragraph("<b>Status / Remarks</b>", table_hdr_style)
        ]
        sheet_table_data = [sheet_headers] + kpi_rows_data
        sheet_table = Table(sheet_table_data, colWidths=[40, 20, 160, 45, 45, 50, 150])
        
        t_styles = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003478')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (3,1), (5,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]
        for r in range(1, len(sheet_table_data)):
            if r in danger_rows:
                t_styles.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#FEF2F2')))
                t_styles.append(('GRID', (0, r), (-1, r), 0.5, colors.HexColor('#FCA5A5')))
            elif r % 2 == 0:
                t_styles.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#F8FAFC')))
                
        sheet_table.setStyle(TableStyle(t_styles))
        
        story.append(Paragraph("Monthly KPI Performance Sheet", section_style))
        story.append(Spacer(1, 5))
        story.append(sheet_table)

    # 3. SIGNATURE BLOCK AT THE END OF THE REPORT
    story.append(Spacer(1, 15))
    sig_data = [
        [Paragraph("<b>Department TPM Coordinator</b><br/><font size=7 color='#64748B'>Sign / Date</font>", table_cell_style),
         Paragraph("<b>Plant TPM Administrator / Head</b><br/><font size=7 color='#64748B'>Sign / Date</font>", table_cell_style)]
    ]
    sig_table = Table(sig_data, colWidths=[250, 260])
    sig_table.setStyle(TableStyle([
        ('LINEABOVE', (0,0), (-1,-1), 1.0, colors.HexColor('#94A3B8')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(KeepTogether([sig_table]))

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)
    return buffer.getvalue()
